import hashlib
import math
import secrets as py_secrets
import sqlite3
import textwrap
from datetime import date, datetime, timedelta
from html import escape
from io import BytesIO

import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

# ====================== CONFIG ======================
st.set_page_config(
    page_title="Dashboard Financeiro",
    layout="wide",
    page_icon="📊",
    initial_sidebar_state="expanded",
)

DB_FILE = "financeiro.db"
_USANDO_DB_REMOTO = False


def _segredo(*chaves: str):
    """Lê st.secrets ou variável de ambiente."""
    import os
    for chave in chaves:
        try:
            val = st.secrets.get(chave)
            if val:
                return str(val).strip()
        except Exception:
            pass
        val = os.environ.get(chave)
        if val:
            return str(val).strip()
    return None


def get_conn():
    """
    Conexão durável:
    1) Turso (remoto) se TURSO_DATABASE_URL + TURSO_AUTH_TOKEN estiverem nos Secrets
    2) SQLite local (dev / fallback)
    """
    global _USANDO_DB_REMOTO
    url = _segredo("TURSO_DATABASE_URL", "TURSO_URL")
    token = _segredo("TURSO_AUTH_TOKEN", "TURSO_TOKEN")

    if url and token:
        try:
            import libsql_experimental as libsql
            conn = libsql.connect(url, auth_token=token)
            _USANDO_DB_REMOTO = True
            return conn
        except Exception as e:
            # Se o remoto falhar, não mascara o erro em produção configurada
            try:
                st.warning(f"Banco remoto indisponível, usando SQLite local. ({e})")
            except Exception:
                pass

    _USANDO_DB_REMOTO = False
    conn = sqlite3.connect(DB_FILE, timeout=30, check_same_thread=False)
    try:
        conn.execute("PRAGMA journal_mode=DELETE")
        conn.execute("PRAGMA busy_timeout=30000")
    except Exception:
        pass
    return conn


def banco_eh_remoto() -> bool:
    return bool(_USANDO_DB_REMOTO)


# ====================== UTILITÁRIOS ======================


def brl(valor) -> str:
    try:
        if valor is None:
            numero = 0.0
        else:
            numero = float(valor)
        if numero != numero:  # NaN check
            numero = 0.0
    except Exception:
        numero = 0.0
    texto = f"R$ {numero:,.2f}"
    return texto.replace(",", "X").replace(".", ",").replace("X", ".")


def brl_curto(valor) -> str:
    try:
        numero = float(valor or 0)
    except (TypeError, ValueError):
        numero = 0.0
    absoluto = abs(numero)
    if absoluto >= 1_000_000:
        texto = f"R$ {numero / 1_000_000:.1f} mi"
    elif absoluto >= 1_000:
        texto = f"R$ {numero / 1_000:.1f} mil"
    else:
        texto = f"R$ {numero:.0f}"
    return texto.replace(".", ",")


def pct(valor, casas: int = 0) -> str:
    try:
        numero = float(valor)
    except (TypeError, ValueError):
        numero = 0.0
    return f"{numero:.{casas}f}%".replace(".", ",")


def data_br(valor) -> str:
    if valor is None:
        return "Sem data"
    try:
        texto = str(valor).strip()
        # ISO dates (YYYY-MM-DD) from SQLite — no dayfirst needed
        if len(texto) >= 10 and texto[4] == "-" and texto[7] == "-":
            data_convertida = pd.to_datetime(texto[:10], errors="coerce", format="%Y-%m-%d")
        else:
            data_convertida = pd.to_datetime(valor, errors="coerce", dayfirst=True)
    except Exception:
        return "Sem data"
    try:
        if data_convertida is None or pd.isna(data_convertida):
            return "Sem data"
    except Exception:
        return "Sem data"
    return data_convertida.strftime("%d/%m/%Y")


def limpar_texto(valor, padrao: str = "") -> str:
    if valor is None:
        return padrao
    try:
        # Avoid pd.isna on exotic objects that raise
        if isinstance(valor, float) and math.isnan(valor):
            return padrao
        if isinstance(valor, (pd.Series, pd.DataFrame)):
            return padrao
        # Scalar check without exploding
        try:
            if valor is pd.NA or valor is pd.NaT:
                return padrao
        except Exception:
            pass
        if str(valor).lower() in ("nan", "nat", "none", "<na>"):
            return padrao
    except Exception:
        return padrao
    texto = str(valor).strip()
    if not texto or texto.lower() == "nan":
        return padrao
    return texto


def normalizar(valor) -> str:
    texto = limpar_texto(valor).lower()
    trocas = {
        "á": "a", "à": "a", "ã": "a", "â": "a",
        "é": "e", "ê": "e", "í": "i",
        "ó": "o", "ô": "o", "õ": "o",
        "ú": "u", "ç": "c",
    }
    for origem, destino in trocas.items():
        texto = texto.replace(origem, destino)
    return " ".join(texto.replace("_", " ").replace("-", " ").split())


def texto_meses(meses: int) -> str:
    if meses <= 0:
        return "agora"
    anos = meses // 12
    meses_restantes = meses % 12
    partes = []
    if anos:
        partes.append(f"{anos} ano" if anos == 1 else f"{anos} anos")
    if meses_restantes:
        partes.append(f"{meses_restantes} mês" if meses_restantes == 1 else f"{meses_restantes} meses")
    return " e ".join(partes) if partes else "menos de 1 mês"


def limitar_percentual(valor) -> float:
    try:
        numero = float(valor)
    except (TypeError, ValueError):
        numero = 0.0
    return max(0.0, min(numero, 100.0))


def mensagem_erro_usuario(erro) -> str:
    texto = str(erro or "").strip()
    if not texto or len(texto) > 180:
        return "Não conseguimos concluir agora. Tente novamente."
    return texto


# ====================== ESTILO (otimizado + responsivo) ======================
st.markdown(
    """
<style>
/* ========== TOKENS ========== */
:root {
    --ink: #1c1f26;
    --muted: #8a90a0;
    --lime: #d9ff00;
    --mint: #b8f0d8;
    --mint-soft: #e8faf3;
    --mint-mid: #7ed9b0;
    --blue: #a8c4ff;
    --bg: #f3f4f7;
    --card: #ffffff;
    --line: rgba(28, 31, 38, 0.08);
    --shadow: 0 12px 40px rgba(28, 31, 38, 0.07);
    --shadow-soft: 0 6px 18px rgba(28, 31, 38, 0.05);
    --radius: 18px;
    --radius-sm: 14px;
    --pad-page: 1.25rem 1.4rem 2.5rem;
    --gap-grid: 0.95rem;
    --card-pad: 1.15rem 1.2rem;
    --font-value: clamp(1.25rem, 2.4vw, 1.9rem);
    --touch: 2.7rem;
}

/* ========== BASE ========== */
html, body, [class*="css"] {
    font-family: Inter, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    -webkit-text-size-adjust: 100%;
    -webkit-tap-highlight-color: transparent;
}

[data-testid="stAppViewContainer"] {
    background:
        radial-gradient(circle at 0% 0%, rgba(217, 255, 0, 0.28), transparent 22rem),
        radial-gradient(circle at 100% 0%, rgba(184, 240, 216, 0.42), transparent 24rem),
        radial-gradient(circle at 15% 85%, rgba(168, 196, 255, 0.30), transparent 22rem),
        radial-gradient(circle at 90% 90%, rgba(217, 255, 0, 0.14), transparent 18rem),
        linear-gradient(180deg, #f7fff0 0%, #f3f4f7 45%, #eef8ff 100%) !important;
    color: var(--ink);
}
section.main {
    background: transparent !important;
}

[data-testid="stHeader"],
[data-testid="stToolbar"] {
    background: transparent !important;
}

.main .block-container {
    max-width: 1200px;
    padding: var(--pad-page) !important;
}

h1, h2, h3, h4,
.stSubheader,
[data-testid="stCaptionContainer"],
.stCaption {
    color: var(--ink) !important;
    opacity: 1 !important;
    letter-spacing: -0.02em;
    word-wrap: break-word;
    overflow-wrap: break-word;
}

h1, h2, h3 {
    font-weight: 800 !important;
}

/* ========== SIDEBAR ========== */
[data-testid="stSidebar"] {
    background: var(--card) !important;
    border-right: 1px solid var(--line);
    box-shadow: 4px 0 24px rgba(28, 31, 38, 0.03);
}

[data-testid="stSidebar"] .stRadio > label {
    display: none !important;
}

[data-testid="stSidebar"] .stRadio [role="radiogroup"] label {
    padding: 0.55rem 0.95rem !important;
    border-radius: 12px !important;
    font-weight: 650 !important;
    color: #6b7280 !important;
    margin-bottom: 0.25rem !important;
    min-height: var(--touch);
    transition: background 0.15s ease, color 0.15s ease;
}

[data-testid="stSidebar"] .stRadio [role="radiogroup"] label:hover {
    background: #f6f7f9 !important;
    color: var(--ink) !important;
}

[data-testid="stSidebar"] .stRadio [role="radiogroup"] label[data-checked="true"],
[data-testid="stSidebar"] .stRadio [aria-checked="true"] {
    background: #1c1f26 !important;
    color: var(--ink) !important;
    font-weight: 800 !important;
    box-shadow: 0 4px 14px rgba(217, 255, 0, 0.25);
}

/* ========== GRIDS ========== */
.metric-grid,
.indicator-grid,
.answer-grid,
.goal-grid,
.debt-grid,
.investment-grid {
    display: grid;
    grid-template-columns: repeat(4, minmax(0, 1fr));
    gap: var(--gap-grid);
    margin: 1rem 0 1.15rem;
}

/* ========== CARDS ========== */
.metric-card,
.indicator-card,
.answer-card,
.goal-card,
.history-item,
.investment-item,
.debt-item {
    position: relative;
    min-height: 7.4rem;
    padding: var(--card-pad);
    border-radius: var(--radius);
    background: var(--card);
    border: 1px solid var(--line);
    box-shadow: var(--shadow-soft);
    overflow: visible !important;
    transition: transform 0.15s ease, box-shadow 0.15s ease;
}

.metric-card:hover,
.indicator-card:hover {
    transform: translateY(-2px);
    box-shadow: var(--shadow);
}

.metric-card.accent {
    background: linear-gradient(145deg, #d9ff00 0%, #e8ff4a 100%);
    border: none;
    box-shadow: 0 10px 28px rgba(217, 255, 0, 0.28);
}

.metric-card.mint {
    background: linear-gradient(145deg, #e8faf3 0%, #d4f5e8 100%);
    border: 1px solid rgba(126, 217, 176, 0.25);
}

.metric-card.dark {
    background: linear-gradient(145deg, #1c1f26 0%, #2a2e38 100%);
    color: #fff;
    border: none;
}

.metric-card.dark .metric-label,
.metric-card.dark .metric-foot,
.metric-card.dark .metric-value {
    color: #fff !important;
}

.metric-label,
.indicator-top,
.answer-question {
    color: var(--muted);
    font-size: 0.76rem;
    font-weight: 700;
    letter-spacing: 0.01em;
    white-space: normal !important;
    overflow: visible !important;
    word-break: break-word !important;
}

.metric-value,
.indicator-value,
.answer-value {
    margin-top: 0.65rem;
    color: var(--ink);
    font-size: var(--font-value);
    line-height: 1.2 !important;
    font-weight: 800;
    letter-spacing: -0.03em;
    white-space: normal !important;
    overflow: visible !important;
    text-overflow: unset !important;
    word-break: break-word !important;
    overflow-wrap: anywhere !important;
    max-width: 100% !important;
}

.metric-foot,
.indicator-note,
.answer-action,
.goal-meta,
.debt-meta,
.investment-meta {
    margin-top: 0.35rem;
    color: var(--muted);
    font-size: 0.8rem;
    line-height: 1.4;
    font-weight: 500;
    white-space: normal !important;
    word-break: break-word !important;
}

.positive { color: #0d9f6e; font-weight: 800; }
.negative { color: #e04b5a; font-weight: 800; }

.goal-progress,
.progress-track {
    overflow: hidden;
    height: 0.48rem;
    margin-top: 0.65rem;
    border-radius: 999px;
    background: rgba(28, 31, 38, 0.07);
}

.goal-progress span,
.progress-track span {
    display: block;
    height: 100%;
    border-radius: inherit;
    background: linear-gradient(90deg, var(--lime), var(--mint-mid));
}

.chart-intro,
.history-summary {
    margin: 0.85rem 0;
    padding: 1rem 1.1rem;
    border-radius: var(--radius-sm);
    background: var(--card);
    border: 1px solid var(--line);
    box-shadow: var(--shadow-soft);
    color: var(--ink);
    font-size: 0.93rem;
}

.answer-good {
    background: #e8faf3;
    border: 1px solid rgba(126, 217, 176, 0.35);
}
.answer-care {
    background: #f7ffc8;
    border: 1px solid rgba(217, 255, 0, 0.45);
}
.answer-risk {
    background: #eef0f4;
    border: 1px solid rgba(28, 31, 38, 0.12);
}
.answer-card .answer-question {
    color: #6b7280 !important;
}
.answer-card .answer-value {
    color: #1c1f26 !important;
}
.answer-card .answer-action {
    color: #6b7280 !important;
}

/* Header */
.page-header {
    display: flex;
    justify-content: space-between;
    align-items: flex-start;
    gap: 1rem;
    margin-bottom: 1.1rem;
    flex-wrap: wrap;
}

.page-header h1 {
    margin: 0.1rem 0 0.2rem !important;
    font-size: clamp(1.55rem, 5vw, 2.25rem) !important;
    font-weight: 800 !important;
}

.page-header p {
    margin: 0;
    color: var(--muted);
    font-size: 0.92rem;
    font-weight: 500;
}

.user-chip {
    display: flex;
    align-items: center;
    gap: 0.55rem;
    padding: 0.32rem 0.85rem 0.32rem 0.38rem;
    border-radius: 999px;
    background: var(--card);
    box-shadow: var(--shadow-soft);
    border: 1px solid var(--line);
    font-size: 0.84rem;
    font-weight: 700;
    color: var(--ink);
    white-space: nowrap;
    flex-shrink: 0;
}

.avatar-dot {
    width: 1.85rem;
    height: 1.85rem;
    display: grid;
    place-items: center;
    border-radius: 50%;
    background: linear-gradient(135deg, #1c1f26, #3a404c);
    color: #fff;
    font-size: 0.72rem;
    font-weight: 800;
    flex-shrink: 0;
}

.badge-new,
.badge-lime {
    display: inline-flex;
    align-items: center;
    padding: 0.22rem 0.65rem;
    border-radius: 999px;
    font-size: 0.74rem;
    font-weight: 800;
}
.badge-new { background: var(--mint); color: #0d5c3d; }
.badge-lime { background: var(--lime); color: var(--ink); }

.card-icon {
    position: absolute;
    top: 0.9rem;
    right: 0.9rem;
    width: 2rem;
    height: 2rem;
    display: grid;
    place-items: center;
    border-radius: 50%;
    border: 1.5px solid rgba(28, 31, 38, 0.1);
    font-size: 0.85rem;
    color: var(--ink);
    background: rgba(255, 255, 255, 0.5);
    flex-shrink: 0;
}

.metric-card.accent .card-icon {
    border-color: rgba(28, 31, 38, 0.18);
    background: rgba(255, 255, 255, 0.35);
}

.metric-card.dark .card-icon {
    border-color: rgba(255, 255, 255, 0.2);
    color: #fff;
    background: rgba(255, 255, 255, 0.08);
}

.status-pill {
    display: inline-flex;
    align-items: center;
    padding: 0.2rem 0.6rem;
    border-radius: 999px;
    font-size: 0.72rem;
    font-weight: 750;
}
.status-ok { background: var(--mint-soft); color: #0d5c3d; }
.status-warn { background: #fff4d6; color: #8a5a00; }
.status-bad { background: #ffe4e8; color: #a31d2e; }

/* List items */
.history-item,
.investment-item,
.debt-item {
    display: flex;
    justify-content: space-between;
    align-items: center;
    gap: 0.85rem;
    min-height: auto;
    margin-bottom: 0.65rem;
}

.history-item > div:first-child,
.investment-item > div:first-child,
.debt-item > div:first-child {
    min-width: 0;
    flex: 1;
}

/* ========== TABS ========== */
div[data-testid="stTabs"] {
    margin-top: 0.35rem;
}

div[data-testid="stTabs"] [data-baseweb="tab-list"] {
    gap: 0.4rem;
    background: transparent !important;
    flex-wrap: wrap !important;
    row-gap: 0.4rem !important;
    overflow-x: visible !important;
    -webkit-overflow-scrolling: touch;
}

div[data-testid="stTabs"] [data-baseweb="tab-highlight"],
div[data-testid="stTabs"] [data-baseweb="tab-border"] {
    display: none !important;
}

div[data-testid="stTabs"] button {
    border-radius: 999px !important;
    color: #1c1f26 !important;
    font-weight: 700 !important;
    background: #ffffff !important;
    border: 1px solid var(--line) !important;
    box-shadow: var(--shadow-soft) !important;
    padding: 0.45rem 1rem !important;
    min-height: 2.5rem;
    height: auto !important;
    font-size: 0.88rem !important;
    white-space: normal !important;
    line-height: 1.25 !important;
    max-width: 100%;
    opacity: 1 !important;
    transition: background 0.15s ease, color 0.15s ease, box-shadow 0.15s ease;
}
div[data-testid="stTabs"] button p,
div[data-testid="stTabs"] button span,
div[data-testid="stTabs"] button div {
    color: #1c1f26 !important;
    opacity: 1 !important;
    -webkit-text-fill-color: #1c1f26 !important;
}

div[data-testid="stTabs"] button:hover {
    background: #f8f9fb !important;
    color: var(--ink) !important;
}

div[data-testid="stTabs"] button[aria-selected="true"] {
    color: #1c1f26 !important;
    background: var(--lime) !important;
    border-color: transparent !important;
    font-weight: 800 !important;
    box-shadow: 0 4px 16px rgba(217, 255, 0, 0.3) !important;
    opacity: 1 !important;
}
div[data-testid="stTabs"] button[aria-selected="true"] * {
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
    opacity: 1 !important;
}

/* ========== CHARTS / SURFACES ========== */
div[data-testid="stPlotlyChart"],
div[data-testid="stDataFrame"],
div[data-testid="stExpander"],
div[data-testid="stForm"],
[data-testid="stAlert"] {
    border-radius: var(--radius) !important;
    background: var(--card) !important;
    border: 1px solid var(--line) !important;
    box-shadow: var(--shadow-soft) !important;
}

div[data-testid="stPlotlyChart"],
div[data-testid="stDataFrame"] {
    overflow-x: auto !important;
    max-width: 100% !important;
}

.js-plotly-plot,
.plotly {
    max-width: 100% !important;
}

/* ========== BUTTONS ========== */
.stButton > button,
[data-testid="stFormSubmitButton"] button,
div[data-testid="stDownloadButton"] button {
    min-height: var(--touch) !important;
    height: auto !important;
    border-radius: 999px !important;
    font-weight: 750 !important;
    box-shadow: var(--shadow-soft) !important;
    width: 100%;
    padding: 0.55rem 1.1rem !important;
    white-space: normal !important;
    line-height: 1.25 !important;
    word-break: break-word !important;
    transition: background 0.15s ease, transform 0.15s ease, box-shadow 0.15s ease;
}

.stButton > button,
[data-testid="stFormSubmitButton"] button,
div[data-testid="stForm"] button,
button[kind="primary"],
button[kind="primaryFormSubmit"],
button[data-testid="baseButton-primary"],
button[data-testid="baseButton-secondaryFormSubmit"],
button[data-testid="baseButton-primaryFormSubmit"] {
    border: 0 !important;
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
    background: #1c1f26 !important;
    opacity: 1 !important;
}

.stButton > button *,
[data-testid="stFormSubmitButton"] button *,
div[data-testid="stForm"] button *,
button[kind="primary"] *,
button[kind="primaryFormSubmit"] *,
button[data-testid="baseButton-primary"] p,
button[data-testid="baseButton-primaryFormSubmit"] p,
button[data-testid="baseButton-secondaryFormSubmit"] p {
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
    opacity: 1 !important;
    fill: #ffffff !important;
}

.stButton > button:hover,
[data-testid="stFormSubmitButton"] button:hover,
div[data-testid="stForm"] button:hover,
button[kind="primary"]:hover,
button[kind="primaryFormSubmit"]:hover {
    background: #2d323c !important;
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
    transform: translateY(-1px);
    box-shadow: 0 8px 20px rgba(28, 31, 38, 0.12) !important;
}

div[data-testid="stDownloadButton"] button {
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
    background: #1c1f26 !important;
    border: none !important;
    font-weight: 800 !important;
}
div[data-testid="stDownloadButton"] button p,
div[data-testid="stDownloadButton"] button span {
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
}

div[data-testid="stDownloadButton"] button:hover {
    background: #2d323c !important;
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
}

/* ========== FORMS / LABELS ========== */
label,
[data-testid="stWidgetLabel"],
[data-testid="stWidgetLabel"] p,
.stSelectbox label,
.stTextInput label,
.stNumberInput label,
.stDateInput label,
.stTextArea label,
.stRadio label,
.stMultiSelect label {
    color: var(--ink) !important;
    opacity: 1 !important;
    visibility: visible !important;
    font-weight: 650 !important;
    font-size: 0.88rem !important;
    margin-bottom: 0.25rem !important;
}

div[data-baseweb="input"],
div[data-baseweb="input"] > div,
div[data-baseweb="select"] > div,
div[data-baseweb="base-input"],
input,
textarea,
[data-baseweb="textarea"],
.stTextInput input,
.stNumberInput input,
.stDateInput input,
.stTextArea textarea {
    background: #fff !important;
    color: var(--ink) !important;
    border: 1.5px solid rgba(28, 31, 38, 0.12) !important;
    border-radius: 12px !important;
    min-height: var(--touch) !important;
    font-size: 0.95rem !important;
    caret-color: var(--ink) !important;
    box-shadow: none !important;
}

div[data-baseweb="input"] input,
div[data-baseweb="select"] input {
    color: var(--ink) !important;
    background: transparent !important;
    -webkit-text-fill-color: var(--ink) !important;
}

input::placeholder,
textarea::placeholder {
    color: var(--muted) !important;
    opacity: 1 !important;
    -webkit-text-fill-color: var(--muted) !important;
}

div[data-baseweb="select"] span,
div[data-baseweb="select"] div {
    color: var(--ink) !important;
}

div[data-baseweb="input"]:focus-within,
div[data-baseweb="select"] > div:focus-within,
textarea:focus {
    border-color: rgba(126, 217, 176, 0.75) !important;
    box-shadow: 0 0 0 3px rgba(184, 240, 216, 0.35) !important;
}

div[role="radiogroup"] {
    flex-wrap: wrap !important;
    gap: 0.45rem !important;
}

div[role="radiogroup"] label {
    white-space: normal !important;
}

[data-testid="column"] {
    min-width: 0 !important;
    overflow: visible !important;
}

[data-testid="stMetricValue"] {
    white-space: normal !important;
    overflow: visible !important;
    word-break: break-word !important;
}

[data-testid="stMetricLabel"] {
    color: var(--ink) !important;
    opacity: 1 !important;
}

div[data-testid="stExpander"] summary,
div[data-testid="stExpander"] summary p,
div[data-testid="stExpander"] summary span {
    color: var(--ink) !important;
    opacity: 1 !important;
    font-weight: 700 !important;
}

/* ========== TABLET ========== */
@media (max-width: 980px) {
    :root {
        --pad-page: 1rem 1rem 2rem;
        --gap-grid: 0.75rem;
        --card-pad: 1rem;
        --font-value: clamp(1.2rem, 4vw, 1.65rem);
    }

    .metric-grid,
    .indicator-grid,
    .answer-grid,
    .goal-grid,
    .debt-grid,
    .investment-grid {
        grid-template-columns: repeat(2, minmax(0, 1fr));
    }

    .metric-card,
    .indicator-card,
    .answer-card,
    .goal-card {
        min-height: 6.8rem;
    }
}

/* ========== MOBILE ========== */
@media (max-width: 640px) {
    :root {
        --pad-page: 0.75rem 0.7rem 1.5rem;
        --gap-grid: 0.65rem;
        --card-pad: 0.95rem 1rem;
        --font-value: 1.35rem;
        --radius: 16px;
        --touch: 2.85rem;
    }

    .metric-grid,
    .indicator-grid,
    .answer-grid,
    .goal-grid,
    .debt-grid,
    .investment-grid {
        grid-template-columns: 1fr !important;
        margin: 0.75rem 0 1rem;
    }

    .metric-card,
    .indicator-card,
    .answer-card,
    .goal-card,
    .history-item,
    .investment-item,
    .debt-item {
        min-height: auto;
    }

    .metric-label,
    .indicator-top,
    .answer-question {
        font-size: 0.72rem;
    }

    .metric-foot,
    .indicator-note,
    .answer-action,
    .goal-meta,
    .debt-meta,
    .investment-meta {
        font-size: 0.76rem;
    }

    .card-icon {
        width: 1.75rem;
        height: 1.75rem;
        top: 0.8rem;
        right: 0.8rem;
        font-size: 0.78rem;
    }

    .page-header {
        flex-direction: column;
        align-items: stretch;
        gap: 0.65rem;
        margin-bottom: 0.85rem;
    }

    .page-header h1 {
        font-size: 1.55rem !important;
    }

    .page-header p {
        font-size: 0.85rem;
    }

    .user-chip {
        align-self: flex-start;
        font-size: 0.8rem;
    }

    .badge-new,
    .badge-lime {
        font-size: 0.7rem;
        padding: 0.18rem 0.55rem;
    }

    /* Tabs: 2 por linha */
    div[data-testid="stTabs"] [data-baseweb="tab-list"] {
        justify-content: flex-start !important;
    }

    div[data-testid="stTabs"] button {
        flex: 1 1 calc(50% - 0.3rem) !important;
        min-width: calc(50% - 0.3rem) !important;
        max-width: 100% !important;
        text-align: center !important;
        padding: 0.45rem 0.55rem !important;
        font-size: 0.76rem !important;
    }

    .chart-intro,
    .history-summary {
        padding: 0.85rem 0.95rem;
        font-size: 0.88rem;
    }

    .stButton > button,
    [data-testid="stFormSubmitButton"] button,
    div[data-testid="stDownloadButton"] button {
        min-height: 3rem !important;
        font-size: 0.95rem !important;
    }

    /* List items stack */
    .history-item,
    .investment-item,
    .debt-item {
        flex-direction: column;
        align-items: flex-start;
        gap: 0.45rem;
    }

    .history-item > div:last-child,
    .investment-item > div:last-child,
    .debt-item > div:last-child {
        align-self: flex-end;
        font-size: 1.05rem;
    }

    /* Sem lift no touch */
    .metric-card:hover,
    .indicator-card:hover,
    .stButton > button:hover {
        transform: none !important;
    }

    /* Labels / inputs mobile */
    label,
    [data-testid="stWidgetLabel"] p {
        font-size: 0.9rem !important;
    }

    div[data-baseweb="input"],
    div[data-baseweb="select"] > div,
    input,
    textarea {
        font-size: 16px !important; /* evita zoom iOS */
    }

    h2, h3, .stSubheader {
        font-size: 1.15rem !important;
    }

    .stCaption,
    [data-testid="stCaptionContainer"] {
        font-size: 0.78rem !important;
    }

    div[data-testid="stExpander"] summary {
        min-height: 2.8rem;
    }
}

/* ========== PHONE PEQUENO ========== */
@media (max-width: 380px) {
    :root {
        --pad-page: 0.65rem 0.55rem 1.25rem;
        --font-value: 1.28rem;
    }

    .page-header h1 {
        font-size: 1.4rem !important;
    }

    div[data-testid="stTabs"] button {
        flex: 1 1 100% !important;
        min-width: 100% !important;
        font-size: 0.74rem !important;
        padding: 0.4rem 0.5rem !important;
    }
}

/* ========== ACESSIBILIDADE ========== */
@media (prefers-reduced-motion: reduce) {
    .metric-card,
    .indicator-card,
    .stButton > button,
    [data-testid="stFormSubmitButton"] button {
        transition: none !important;
    }

    .metric-card:hover,
    .indicator-card:hover,
    .stButton > button:hover {
        transform: none !important;
    }
}

/* Landscape phones: 2 cols when short height */
@media (max-width: 900px) and (orientation: landscape) {
    :root {
        --pad-page: 0.6rem 0.9rem 1.2rem;
    }

    .metric-grid,
    .indicator-grid,
    .answer-grid {
        grid-template-columns: repeat(2, minmax(0, 1fr));
    }
}

/* ========== CONTRASTE OBRIGATÓRIO (nunca branco em branco) ========== */

/* Tabs: texto sempre escuro */
div[data-testid="stTabs"] button,
div[data-testid="stTabs"] button * {
    color: #1c1f26 !important;
    opacity: 1 !important;
    -webkit-text-fill-color: #1c1f26 !important;
}
div[data-testid="stTabs"] button[aria-selected="true"],
div[data-testid="stTabs"] button[aria-selected="true"] * {
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
    opacity: 1 !important;
}

/* Radio horizontal (Tipo Entrada/Saída) */
div[role="radiogroup"] label,
div[role="radiogroup"] label p,
div[role="radiogroup"] label span,
div[role="radiogroup"] label div,
[data-testid="stRadio"] label,
[data-testid="stRadio"] label p,
[data-testid="stRadio"] label span,
[data-testid="stRadio"] label div,
[data-testid="stRadio"] p,
.stRadio label,
.stRadio p,
.stRadio span {
    color: #1c1f26 !important;
    opacity: 1 !important;
    visibility: visible !important;
    -webkit-text-fill-color: #1c1f26 !important;
    font-weight: 650 !important;
}

/* Baseweb radio text specifically */
div[data-baseweb="radio"],
div[data-baseweb="radio"] + div,
div[data-baseweb="radio"] ~ div,
label[data-baseweb="radio"],
[data-baseweb="radio"] span {
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
    opacity: 1 !important;
}

/* Textos de widget no main (sem forçar todos os spans — preserva .positive/.negative) */
.main label,
.main [data-testid="stWidgetLabel"],
.main [data-testid="stWidgetLabel"] p,
.main [data-testid="stWidgetLabel"] span,
.main .stRadio label,
.main .stRadio p,
.main [data-testid="stRadio"] [data-testid="stMarkdownContainer"] p,
section.main label,
[data-testid="stMain"] label,
[data-testid="stMain"] [data-testid="stWidgetLabel"] p {
    color: #1c1f26 !important;
    opacity: 1 !important;
    -webkit-text-fill-color: #1c1f26 !important;
}

/* Checkbox / multiselect options */
[data-baseweb="checkbox"] + div,
[data-baseweb="checkbox"] ~ span {
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
}

/* Placeholder pode ser cinza, mas não branco */
input::placeholder,
textarea::placeholder {
    color: #6b7280 !important;
    -webkit-text-fill-color: #6b7280 !important;
    opacity: 1 !important;
}

/* Força contraste em radios desmarcados (BaseWeb usa opacity baixa) */
div[role="radiogroup"] label[data-checked="false"],
div[role="radiogroup"] label:not([data-checked="true"]) {
    opacity: 1 !important;
}
div[role="radiogroup"] label[data-checked="false"] *,
div[role="radiogroup"] label:not([data-checked="true"]) * {
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
    opacity: 1 !important;
}

/* Streamlit 1.3x radio structure */
[data-testid="stRadio"] [data-testid="stMarkdownContainer"] p {
    color: #1c1f26 !important;
    opacity: 1 !important;
}

/* Tabs inativas: fundo cinza bem claro, texto escuro (não transparente) */
div[data-testid="stTabs"] button[aria-selected="false"] {
    background: #ffffff !important;
    color: #1c1f26 !important;
    opacity: 1 !important;
    border: 1px solid rgba(28, 31, 38, 0.12) !important;
}
div[data-testid="stTabs"] button[aria-selected="false"] * {
    color: #1c1f26 !important;
    opacity: 1 !important;
    -webkit-text-fill-color: #1c1f26 !important;
}

/* Active tab: lime + texto escuro */
div[data-testid="stTabs"] button[aria-selected="true"] {
    background: var(--lime) !important;
    color: #1c1f26 !important;
    opacity: 1 !important;
}


/* ========== OVERRIDE FINAL DE CONTRASTE (alta especificidade) ========== */

/* Streamlit tabs (várias versões do DOM) */
button[data-baseweb="tab"],
button[role="tab"],
div[data-testid="stTabs"] button,
div[data-testid="stTabs"] [role="tab"] {
    color: #1c1f26 !important;
    opacity: 1 !important;
    -webkit-text-fill-color: #1c1f26 !important;
    background-color: #ffffff !important;
    border: 1px solid rgba(28,31,38,0.12) !important;
    border-radius: 999px !important;
    box-shadow: 0 4px 12px rgba(28,31,38,0.05) !important;
}

button[data-baseweb="tab"] *,
button[role="tab"] *,
div[data-testid="stTabs"] button *,
div[data-testid="stTabs"] [role="tab"] * {
    color: #1c1f26 !important;
    opacity: 1 !important;
    -webkit-text-fill-color: #1c1f26 !important;
}

button[data-baseweb="tab"][aria-selected="true"],
button[role="tab"][aria-selected="true"],
div[data-testid="stTabs"] button[aria-selected="true"] {
    background-color: #d9ff00 !important;
    color: #1c1f26 !important;
    border-color: transparent !important;
    font-weight: 800 !important;
}

/* esconde underline vermelho padrão do Streamlit */
div[data-testid="stTabs"] [data-baseweb="tab-highlight"],
div[data-testid="stTabs"] [data-baseweb="tab-border"],
div[data-testid="stTabs"] hr {
    display: none !important;
    opacity: 0 !important;
    height: 0 !important;
}

/* Radio labels — todas as estruturas conhecidas do Streamlit/Baseweb */
[data-testid="stRadio"] label,
[data-testid="stRadio"] label *,
[data-testid="stRadio"] [data-testid="stMarkdownContainer"],
[data-testid="stRadio"] [data-testid="stMarkdownContainer"] *,
[data-testid="stRadio"] p,
[data-testid="stRadio"] span,
div[role="radiogroup"] label,
div[role="radiogroup"] label *,
div[role="radiogroup"] p,
div[role="radiogroup"] span,
label[data-baseweb="radio"],
label[data-baseweb="radio"] * {
    color: #1c1f26 !important;
    opacity: 1 !important;
    visibility: visible !important;
    -webkit-text-fill-color: #1c1f26 !important;
    font-weight: 650 !important;
}

/* Texto "Entrada"/"Saída" ao lado do círculo */
[data-testid="stRadio"] > div > label > div:last-child,
[data-testid="stRadio"] label > div:last-child,
div[role="radiogroup"] label > div:last-child {
    color: #1c1f26 !important;
    opacity: 1 !important;
    -webkit-text-fill-color: #1c1f26 !important;
}

/* Checkbox text */
[data-testid="stCheckbox"] label,
[data-testid="stCheckbox"] label *,
[data-testid="stCheckbox"] p {
    color: #1c1f26 !important;
    opacity: 1 !important;
    -webkit-text-fill-color: #1c1f26 !important;
}


/* ========== POLISH ESTÉTICO ========== */

/* Números alinhados e legíveis */
.metric-value,
.indicator-value,
.answer-value,
.positive,
.negative {
    font-variant-numeric: tabular-nums;
    font-feature-settings: "tnum" 1;
}

/* Cards: espaço reservado para ícone (não sobrepõe texto) */
.metric-card,
.indicator-card,
.answer-card {
    padding-right: 3.1rem !important;
}

.metric-card .metric-label,
.metric-card .metric-value,
.metric-card .metric-foot {
    max-width: calc(100% - 0.25rem);
    padding-right: 0.15rem;
}

.card-icon {
    z-index: 2;
    pointer-events: none;
}

/* Hierarquia: saldo maior */
.metric-card.accent .metric-value {
    font-size: clamp(1.55rem, 3vw, 2.15rem) !important;
}

/* Badges de score semânticos */
.badge-score {
    display: inline-flex;
    align-items: center;
    gap: 0.3rem;
    padding: 0.28rem 0.75rem;
    border-radius: 999px;
    font-size: 0.78rem;
    font-weight: 800;
    letter-spacing: 0.01em;
}
.badge-score.good {
    background: #c8f5df;
    color: #0b6b45;
}
.badge-score.mid {
    background: #fff0c2;
    color: #8a5a00;
}
.badge-score.bad {
    background: #ffd6db;
    color: #9b1c2e;
}

/* Segmented control visual para radio horizontal */
[data-testid="stRadio"] > div[role="radiogroup"] {
    display: flex !important;
    flex-direction: row !important;
    flex-wrap: wrap !important;
    gap: 0.5rem !important;
    background: #eef0f4;
    padding: 0.35rem;
    border-radius: 999px;
    width: fit-content;
    max-width: 100%;
}
[data-testid="stRadio"] > div[role="radiogroup"] > label {
    margin: 0 !important;
    padding: 0.5rem 1.15rem !important;
    border-radius: 999px !important;
    background: transparent !important;
    border: none !important;
    min-height: 2.5rem !important;
    display: inline-flex !important;
    align-items: center !important;
    gap: 0.4rem !important;
}
[data-testid="stRadio"] > div[role="radiogroup"] > label[data-checked="true"],
[data-testid="stRadio"] > div[role="radiogroup"] > label:has(input:checked) {
    background: #d9ff00 !important;
    box-shadow: 0 2px 10px rgba(217,255,0,0.35) !important;
}
[data-testid="stRadio"] label,
[data-testid="stRadio"] label *,
[data-testid="stRadio"] p,
[data-testid="stRadio"] span {
    color: #1c1f26 !important;
    opacity: 1 !important;
    -webkit-text-fill-color: #1c1f26 !important;
    font-weight: 700 !important;
    visibility: visible !important;
}

/* Tabs mais limpas: inativas cinza, ativas lime */
div[data-testid="stTabs"] button[aria-selected="false"],
button[role="tab"][aria-selected="false"] {
    background: #eef0f4 !important;
    color: #1c1f26 !important;
    border: none !important;
    opacity: 1 !important;
}
div[data-testid="stTabs"] button[aria-selected="true"],
button[role="tab"][aria-selected="true"] {
    background: #d9ff00 !important;
    color: #1c1f26 !important;
    border: none !important;
    font-weight: 800 !important;
}

/* Subtítulos de seção */
.section-title {
    margin: 1.1rem 0 0.55rem;
    font-size: 1.05rem;
    font-weight: 800;
    color: #1c1f26;
    letter-spacing: -0.02em;
}
.section-sub {
    margin: 0 0 0.9rem;
    color: #8a90a0;
    font-size: 0.88rem;
    font-weight: 500;
}

/* Empty-ish intro cards */
.chart-intro {
    border-left: 3px solid #d9ff00;
}

/* Mobile: ícone menor e mais espaço para texto */
@media (max-width: 640px) {
    .metric-card,
    .indicator-card,
    .answer-card {
        padding-right: 2.6rem !important;
        padding-top: 0.95rem !important;
        padding-bottom: 0.95rem !important;
    }
    .card-icon {
        width: 1.55rem !important;
        height: 1.55rem !important;
        top: 0.75rem !important;
        right: 0.7rem !important;
        font-size: 0.7rem !important;
    }
    .metric-card.accent .metric-value {
        font-size: 1.55rem !important;
    }
    [data-testid="stRadio"] > div[role="radiogroup"] {
        width: 100%;
    }
    [data-testid="stRadio"] > div[role="radiogroup"] > label {
        flex: 1 1 auto !important;
        justify-content: center !important;
        padding: 0.55rem 0.75rem !important;
    }
    /* Esconde emoji longo nas tabs no CSS se possível — tabs usam texto do st.tabs */
    div[data-testid="stTabs"] button {
        font-size: 0.78rem !important;
        padding: 0.42rem 0.5rem !important;
    }
}


/* ========== FIX MOBILE HEADER / ICONS / TABS ========== */

.page-header {
    display: block !important;
    margin: 0 0 1rem !important;
    padding: 0 !important;
}
.page-header-main { min-width: 0; }
.page-header-top {
    display: flex;
    align-items: center;
    flex-wrap: wrap;
    gap: 0.45rem;
    margin-bottom: 0.35rem;
}
.page-kicker {
    color: #6b7280 !important;
    font-weight: 700;
    font-size: 0.82rem;
}
.page-sub {
    margin: 0.15rem 0 0 !important;
    color: #6b7280 !important;
    font-size: 0.9rem !important;
    font-weight: 500 !important;
    line-height: 1.4 !important;
}
.page-header h1 {
    margin: 0 !important;
    line-height: 1.15 !important;
}

/* Ícones dos cards: cor escura legível, nunca quadrado preto sólido */
.card-icon {
    position: absolute !important;
    top: 0.85rem !important;
    right: 0.85rem !important;
    width: 1.85rem !important;
    height: 1.85rem !important;
    display: grid !important;
    place-items: center !important;
    border-radius: 50% !important;
    border: 1.5px solid rgba(28, 31, 38, 0.12) !important;
    background: #f3f4f7 !important;
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
    font-size: 0.8rem !important;
    font-weight: 700 !important;
    line-height: 1 !important;
    z-index: 1 !important;
    overflow: hidden !important;
    box-shadow: none !important;
}
.metric-card.accent .card-icon {
    background: rgba(255,255,255,0.55) !important;
    border-color: rgba(28,31,38,0.14) !important;
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
}
.metric-card.mint .card-icon {
    background: rgba(255,255,255,0.7) !important;
    color: #0b6b45 !important;
    -webkit-text-fill-color: #0b6b45 !important;
}

/* Texto do card nunca sob o ícone */
.metric-card {
    padding-right: 3rem !important;
}
.metric-card .metric-label,
.metric-card .metric-value,
.metric-card .metric-foot {
    position: relative;
    z-index: 2;
    max-width: calc(100% - 0.5rem) !important;
    padding-right: 0 !important;
}

/* Tabs: texto sempre legível, sem fill em filhos (quebra emoji/ícone) */
div[data-testid="stTabs"] button,
button[role="tab"],
button[data-baseweb="tab"] {
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
    opacity: 1 !important;
    overflow: visible !important;
    text-overflow: unset !important;
    white-space: nowrap !important;
}
div[data-testid="stTabs"] button *,
button[role="tab"] * {
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
    opacity: 1 !important;
    /* NÃO usar fill — transforma emoji em quadrado preto */
}

div[data-testid="stTabs"] [data-baseweb="tab-list"] {
    gap: 0.4rem !important;
    flex-wrap: wrap !important;
    overflow-x: auto !important;
    -webkit-overflow-scrolling: touch;
    scrollbar-width: none;
    padding-bottom: 0.25rem !important;
}
div[data-testid="stTabs"] [data-baseweb="tab-list"]::-webkit-scrollbar {
    display: none;
}

/* Esconde tooltip "?" estranho no mobile se sobrar */
[data-testid="stTooltipHoverTarget"],
[data-testid="stTooltipIcon"] {
    opacity: 0.85;
}

@media (max-width: 640px) {
    .page-header {
        margin-bottom: 0.75rem !important;
    }
    .page-header h1 {
        font-size: 1.45rem !important;
    }
    .page-sub {
        font-size: 0.84rem !important;
    }
    .page-kicker {
        font-size: 0.76rem !important;
    }
    .badge-score {
        font-size: 0.7rem !important;
        padding: 0.2rem 0.55rem !important;
    }

    /* Ícones menores e fora do texto */
    .metric-card {
        padding: 0.95rem 2.6rem 0.95rem 1rem !important;
    }
    .card-icon {
        top: 0.7rem !important;
        right: 0.65rem !important;
        width: 1.5rem !important;
        height: 1.5rem !important;
        font-size: 0.68rem !important;
        background: #eef0f4 !important;
        color: #374151 !important;
        -webkit-text-fill-color: #374151 !important;
        border-color: rgba(28,31,38,0.1) !important;
    }
    .metric-card.accent .card-icon {
        background: rgba(255,255,255,0.65) !important;
        color: #1c1f26 !important;
        -webkit-text-fill-color: #1c1f26 !important;
    }
    .metric-card.mint .card-icon {
        color: #0b6b45 !important;
        -webkit-text-fill-color: #0b6b45 !important;
    }
    .metric-foot {
        color: #6b7280 !important;
        max-width: 100% !important;
        padding-right: 0 !important;
    }

    /* Tabs mobile: largura automática, scroll horizontal, sem cortar texto */
    div[data-testid="stTabs"] [data-baseweb="tab-list"] {
        flex-wrap: nowrap !important;
        overflow-x: auto !important;
        justify-content: flex-start !important;
        gap: 0.35rem !important;
    }
    div[data-testid="stTabs"] button,
    button[role="tab"] {
        flex: 0 0 auto !important;
        min-width: auto !important;
        max-width: none !important;
        width: auto !important;
        padding: 0.48rem 0.9rem !important;
        font-size: 0.82rem !important;
        white-space: nowrap !important;
    }
}

@media (max-width: 380px) {
    /* Em telas muito estreitas: esconde ícone do card para zero sobreposição */
    .card-icon {
        display: none !important;
    }
    .metric-card {
        padding: 0.9rem 1rem !important;
    }
}


/* ========== HEADER QUOTE + TABS + INPUTS CLEAN ========== */

.page-header {
    display: flex !important;
    justify-content: space-between !important;
    align-items: flex-start !important;
    gap: 1.25rem !important;
    flex-wrap: wrap !important;
    margin: 0 0 1.15rem !important;
}

.page-quote {
    max-width: 16.5rem;
    padding: 0.85rem 1rem;
    border-radius: 16px;
    background: rgba(255, 255, 255, 0.78);
    border: 1px solid rgba(28, 31, 38, 0.06);
    box-shadow: 0 6px 18px rgba(28, 31, 38, 0.04);
    align-self: center;
}

.page-quote-mark {
    display: block;
    font-size: 1.4rem;
    line-height: 1;
    color: #d9ff00;
    font-weight: 900;
    margin-bottom: 0.15rem;
}

.page-quote-text {
    display: block;
    color: #374151 !important;
    font-size: 0.88rem;
    font-weight: 600;
    line-height: 1.45;
}

/* Abas / balões com espaçamento claro */
div[data-testid="stTabs"] {
    margin-top: 0.35rem !important;
    margin-bottom: 0.75rem !important;
}

div[data-testid="stTabs"] [data-baseweb="tab-list"] {
    gap: 0.55rem !important;
    padding: 0.2rem 0 0.55rem !important;
    flex-wrap: wrap !important;
    border-bottom: none !important;
}

div[data-testid="stTabs"] button,
button[role="tab"],
button[data-baseweb="tab"] {
    margin: 0 !important;
    padding: 0.55rem 1.15rem !important;
    min-height: 2.55rem !important;
    border-radius: 999px !important;
    background: #eef0f4 !important;
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
    border: 1px solid transparent !important;
    box-shadow: none !important;
    font-weight: 700 !important;
    font-size: 0.9rem !important;
    letter-spacing: 0.01em;
}

div[data-testid="stTabs"] button[aria-selected="true"],
button[role="tab"][aria-selected="true"] {
    background: #d9ff00 !important;
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
    border-color: transparent !important;
    box-shadow: 0 4px 14px rgba(217, 255, 0, 0.28) !important;
    font-weight: 800 !important;
}

div[data-testid="stTabs"] [data-baseweb="tab-highlight"],
div[data-testid="stTabs"] [data-baseweb="tab-border"],
div[data-testid="stTabs"] hr {
    display: none !important;
    height: 0 !important;
    opacity: 0 !important;
}

/* Remove underline vermelha padrão Streamlit */
div[data-testid="stTabs"] [data-baseweb="tab-border"],
div[data-testid="stTabs"] button::after {
    display: none !important;
}

/* ========== INPUTS: uma linha só (sem borda duplicada) ========== */

/* Container Baseweb sem borda extra */
div[data-baseweb="input"],
div[data-baseweb="base-input"],
div[data-baseweb="select"] > div,
div[data-baseweb="select"] > div > div {
    border: none !important;
    box-shadow: none !important;
    background: transparent !important;
    outline: none !important;
}

/* Campo real com UMA borda */
.stTextInput input,
.stNumberInput input,
.stDateInput input,
.stTextArea textarea,
div[data-baseweb="input"] input,
div[data-baseweb="select"] [data-baseweb="select"] ,
div[data-baseweb="select"] > div {
    border: 1.5px solid rgba(28, 31, 38, 0.12) !important;
    border-radius: 12px !important;
    background: #ffffff !important;
    box-shadow: none !important;
    outline: none !important;
    min-height: 2.7rem !important;
}

/* Number input wrapper Streamlit */
[data-testid="stNumberInput"] > div,
[data-testid="stTextInput"] > div,
[data-testid="stDateInput"] > div,
[data-testid="stSelectbox"] > div {
    border: none !important;
    box-shadow: none !important;
    background: transparent !important;
}

[data-testid="stNumberInput"] input,
[data-testid="stTextInput"] input,
[data-testid="stDateInput"] input {
    border: 1.5px solid rgba(28, 31, 38, 0.12) !important;
    border-radius: 12px !important;
    background: #fff !important;
    box-shadow: none !important;
}

/* Remove anel/outline duplicado no focus do wrapper */
[data-testid="stNumberInput"] div:focus-within,
[data-testid="stTextInput"] div:focus-within,
[data-testid="stDateInput"] div:focus-within,
[data-testid="stSelectbox"] div:focus-within,
div[data-baseweb="input"]:focus-within,
div[data-baseweb="select"] > div:focus-within {
    border: none !important;
    box-shadow: none !important;
    outline: none !important;
}

[data-testid="stNumberInput"] input:focus,
[data-testid="stTextInput"] input:focus,
[data-testid="stDateInput"] input:focus,
.stTextArea textarea:focus {
    border-color: rgba(126, 217, 176, 0.85) !important;
    box-shadow: 0 0 0 3px rgba(184, 240, 216, 0.35) !important;
    outline: none !important;
}

/* Form surface */
div[data-testid="stForm"] {
    border: 1px solid rgba(28, 31, 38, 0.06) !important;
    box-shadow: 0 6px 18px rgba(28, 31, 38, 0.04) !important;
    padding: 1rem 1.1rem 1.15rem !important;
}

@media (max-width: 640px) {
    .page-header {
        flex-direction: column !important;
        gap: 0.75rem !important;
    }
    .page-quote {
        max-width: 100%;
        width: 100%;
    }
    div[data-testid="stTabs"] [data-baseweb="tab-list"] {
        gap: 0.4rem !important;
        flex-wrap: nowrap !important;
        overflow-x: auto !important;
    }
    div[data-testid="stTabs"] button,
    button[role="tab"] {
        flex: 0 0 auto !important;
        padding: 0.5rem 0.95rem !important;
        font-size: 0.84rem !important;
    }
}


/* ========== NUMBER INPUT + FORM POLISH ========== */

/* Streamlit number input: evita “linha dupla” do stepper */
[data-testid="stNumberInput"] {
    width: 100%;
}
[data-testid="stNumberInput"] > div {
    border: none !important;
    box-shadow: none !important;
    background: transparent !important;
}
[data-testid="stNumberInput"] [data-baseweb="input"] {
    border: none !important;
    background: transparent !important;
    box-shadow: none !important;
}
[data-testid="stNumberInput"] input {
    border: 1.5px solid rgba(28, 31, 38, 0.12) !important;
    border-radius: 12px !important;
    background: #fff !important;
    box-shadow: none !important;
    min-height: 2.75rem !important;
    padding: 0.55rem 0.85rem !important;
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
    font-variant-numeric: tabular-nums;
}
/* botões +/- do number input */
[data-testid="stNumberInput"] button {
    border: none !important;
    background: #f3f4f7 !important;
    color: #1c1f26 !important;
    border-radius: 8px !important;
    min-height: 1.6rem !important;
    box-shadow: none !important;
}

/* Date / text: mesma regra limpa */
[data-testid="stDateInput"] input,
[data-testid="stTextInput"] input {
    border: 1.5px solid rgba(28, 31, 38, 0.12) !important;
    border-radius: 12px !important;
    background: #fff !important;
    box-shadow: none !important;
    min-height: 2.75rem !important;
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
}
[data-testid="stDateInput"] > div,
[data-testid="stTextInput"] > div,
[data-testid="stSelectbox"] > div {
    border: none !important;
    box-shadow: none !important;
    background: transparent !important;
}

/* Select */
[data-testid="stSelectbox"] [data-baseweb="select"] > div {
    border: 1.5px solid rgba(28, 31, 38, 0.12) !important;
    border-radius: 12px !important;
    background: #fff !important;
    box-shadow: none !important;
    min-height: 2.75rem !important;
    color: #1c1f26 !important;
}

/* Caption de ajuda mais discreta */
[data-testid="stCaptionContainer"] p,
.stCaption {
    color: #6b7280 !important;
    font-size: 0.84rem !important;
}

/* Botão primary do form */
div[data-testid="stForm"] [data-testid="stFormSubmitButton"] button,
div[data-testid="stForm"] [data-testid="stFormSubmitButton"] button[kind="primary"],
div[data-testid="stForm"] button[kind="primaryFormSubmit"],
div[data-testid="stForm"] button[data-testid="baseButton-primaryFormSubmit"],
div[data-testid="stForm"] button[data-testid="baseButton-secondaryFormSubmit"] {
    background: #1c1f26 !important;
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
    border: none !important;
    min-height: 3rem !important;
    font-weight: 800 !important;
    opacity: 1 !important;
}
div[data-testid="stForm"] [data-testid="stFormSubmitButton"] button *,
div[data-testid="stForm"] button[kind="primaryFormSubmit"] *,
div[data-testid="stForm"] button[data-testid="baseButton-primaryFormSubmit"] * {
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
    opacity: 1 !important;
}

/* Espaço entre radio e formulário */
div[data-testid="stForm"] {
    margin-top: 0.65rem !important;
}


/* ========== SIDEBAR MENU ========== */
[data-testid="stSidebar"] {
    background: #ffffff !important;
    border-right: 1px solid rgba(28,31,38,0.06);
}
[data-testid="stSidebar"] [data-testid="stRadio"] > div[role="radiogroup"] {
    display: flex !important;
    flex-direction: column !important;
    gap: 0.35rem !important;
    background: transparent !important;
    padding: 0 !important;
    border-radius: 0 !important;
    width: 100% !important;
}
[data-testid="stSidebar"] [data-testid="stRadio"] > div[role="radiogroup"] > label {
    width: 100% !important;
    margin: 0 !important;
    padding: 0.7rem 0.95rem !important;
    border-radius: 12px !important;
    background: transparent !important;
    border: none !important;
    min-height: 2.7rem !important;
    justify-content: flex-start !important;
}
[data-testid="stSidebar"] [data-testid="stRadio"] > div[role="radiogroup"] > label[data-checked="true"],
[data-testid="stSidebar"] [data-testid="stRadio"] > div[role="radiogroup"] > label:has(input:checked) {
    background: #d9ff00 !important;
    box-shadow: 0 4px 14px rgba(217,255,0,0.25) !important;
}
[data-testid="stSidebar"] [data-testid="stRadio"] label,
[data-testid="stSidebar"] [data-testid="stRadio"] label *,
[data-testid="stSidebar"] [data-testid="stRadio"] p,
[data-testid="stSidebar"] [data-testid="stRadio"] span {
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
    opacity: 1 !important;
    font-weight: 700 !important;
    visibility: visible !important;
}
/* Esconde o círculo do radio na sidebar — parece menu */
[data-testid="stSidebar"] [data-testid="stRadio"] input,
[data-testid="stSidebar"] [data-baseweb="radio"] > div:first-child,
[data-testid="stSidebar"] [data-baseweb="radio"] > span:first-child {
    display: none !important;
}

/* Garante texto do botão visível em qualquer tema Streamlit */
[data-testid="stFormSubmitButton"] button,
[data-testid="stFormSubmitButton"] button p,
[data-testid="stFormSubmitButton"] button span,
[data-testid="stFormSubmitButton"] button div {
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
    opacity: 1 !important;
}


/* ========== MENU LABEL + COLLAPSE VERMELHO + SELECT ICON ========== */

.side-brand {
    padding: 0.2rem 0.1rem 0.9rem;
}
.side-menu-label {
    display: inline-flex;
    align-items: center;
    gap: 0.4rem;
    font-size: 0.78rem;
    font-weight: 800;
    letter-spacing: 0.06em;
    text-transform: uppercase;
    color: #1c1f26;
}
.side-menu-dot {
    width: 0.55rem;
    height: 0.55rem;
    border-radius: 50%;
    background: #e11d48;
    box-shadow: 0 0 0 3px rgba(225, 29, 72, 0.18);
    flex-shrink: 0;
}
.side-title {
    font-size: 1.2rem;
    font-weight: 800;
    color: #1c1f26;
    margin-top: 0.3rem;
    letter-spacing: -0.02em;
}

/* Botão de recolher sidebar — ícone/vermelho */
[data-testid="stSidebarCollapsedControl"] button,
[data-testid="stSidebarCollapseButton"] button,
[data-testid="collapsedControl"] button,
button[kind="headerNoPadding"],
[data-testid="stBaseButton-headerNoPadding"],
[data-testid="stSidebar"] button[kind="header"],
section[data-testid="stSidebar"] > div:first-child button {
    color: #e11d48 !important;
}
[data-testid="stSidebarCollapsedControl"] button svg,
[data-testid="stSidebarCollapseButton"] button svg,
[data-testid="collapsedControl"] button svg,
button[kind="headerNoPadding"] svg,
[data-testid="stBaseButton-headerNoPadding"] svg,
section[data-testid="stSidebar"] > div:first-child button svg,
[data-testid="stSidebarCollapsedControl"] svg,
[data-testid="collapsedControl"] svg {
    fill: #e11d48 !important;
    color: #e11d48 !important;
    stroke: #e11d48 !important;
}

/* Select: corrige ícone □ preto — usa chevron legível */
[data-testid="stSelectbox"] svg,
div[data-baseweb="select"] svg,
[data-baseweb="select"] svg {
    width: 1rem !important;
    height: 1rem !important;
    color: #1c1f26 !important;
    fill: #1c1f26 !important;
    opacity: 1 !important;
    visibility: visible !important;
}
/* Remove fill agressivo em path que vira quadrado */
[data-testid="stSelectbox"] svg path,
div[data-baseweb="select"] svg path {
    fill: #1c1f26 !important;
    stroke: none !important;
}

/* Dropdown do select: texto escuro */
div[data-baseweb="popover"] li,
div[data-baseweb="menu"] li,
ul[role="listbox"] li,
[role="option"] {
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
}

/* Value do select legível */
[data-testid="stSelectbox"] [data-baseweb="select"] > div,
[data-testid="stSelectbox"] [data-baseweb="select"] span {
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
    opacity: 1 !important;
}

/* Number input steppers: ícones legíveis */
[data-testid="stNumberInput"] button svg {
    fill: #1c1f26 !important;
    color: #1c1f26 !important;
}


/* ========== COLLAPSED CONTROL "MENU" VERMELHO + SELECT SETA BRANCA ========== */

/* Controle da sidebar recolhida (canto superior esquerdo) */
[data-testid="stSidebarCollapsedControl"],
[data-testid="collapsedControl"] {
    display: flex !important;
    align-items: center !important;
    gap: 0.35rem !important;
}
[data-testid="stSidebarCollapsedControl"] button,
[data-testid="collapsedControl"] button,
[data-testid="stBaseButton-headerNoPadding"],
button[kind="headerNoPadding"] {
    color: #e11d48 !important;
    background: transparent !important;
    border: none !important;
}
[data-testid="stSidebarCollapsedControl"] button svg,
[data-testid="collapsedControl"] button svg,
[data-testid="stBaseButton-headerNoPadding"] svg,
button[kind="headerNoPadding"] svg,
[data-testid="stSidebarCollapsedControl"] svg,
[data-testid="collapsedControl"] svg {
    fill: #e11d48 !important;
    color: #e11d48 !important;
    stroke: #e11d48 !important;
}

/* Escreve "Menu" ao lado do ícone >> quando a sidebar está recolhida */
[data-testid="stSidebarCollapsedControl"]::after,
[data-testid="collapsedControl"]::after {
    content: "Menu";
    color: #e11d48 !important;
    font-weight: 800 !important;
    font-size: 0.92rem !important;
    letter-spacing: 0.02em;
    font-family: Inter, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    line-height: 1;
    margin-left: 0.15rem;
    user-select: none;
}

/* Também pinta o controle expandido (dentro da sidebar) de vermelho */
[data-testid="stSidebar"] [data-testid="stBaseButton-header"],
[data-testid="stSidebar"] button[kind="header"],
[data-testid="stSidebarCollapseButton"] button {
    color: #e11d48 !important;
}
[data-testid="stSidebar"] [data-testid="stBaseButton-header"] svg,
[data-testid="stSidebar"] button[kind="header"] svg,
[data-testid="stSidebarCollapseButton"] svg {
    fill: #e11d48 !important;
    color: #e11d48 !important;
    stroke: #e11d48 !important;
}

/* Select: seta branca em fundo preto (pill) */
[data-testid="stSelectbox"] [data-baseweb="select"] > div {
    position: relative !important;
    padding-right: 2.6rem !important;
}

/* Esconde o ícone SVG padrão quebrado (vira □) */
[data-testid="stSelectbox"] [data-baseweb="select"] svg,
div[data-baseweb="select"] svg {
    opacity: 0 !important;
    width: 0 !important;
    height: 0 !important;
    position: absolute !important;
}

/* Pseudo-seta branca no bloco preto */
[data-testid="stSelectbox"] [data-baseweb="select"] > div::after {
    content: "";
    position: absolute;
    top: 50%;
    right: 0.45rem;
    transform: translateY(-50%);
    width: 1.7rem;
    height: 1.7rem;
    border-radius: 0.55rem;
    background: #1c1f26;
    pointer-events: none;
    z-index: 2;
    /* chevron branco via mask */
    -webkit-mask: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24'%3E%3Cpath fill='black' d='M7.4 8.6 12 13.2l4.6-4.6L18 10l-6 6-6-6z'/%3E%3C/svg%3E") center / 0.95rem 0.95rem no-repeat;
    mask: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24'%3E%3Cpath fill='black' d='M7.4 8.6 12 13.2l4.6-4.6L18 10l-6 6-6-6z'/%3E%3C/svg%3E") center / 0.95rem 0.95rem no-repeat;
    /* fallback: se mask falhar, ainda fica o bloco preto */
    box-shadow: inset 0 0 0 999px #1c1f26;
    background-color: #1c1f26;
    background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24'%3E%3Cpath fill='white' d='M7.4 8.6 12 13.2l4.6-4.6L18 10l-6 6-6-6z'/%3E%3C/svg%3E");
    background-repeat: no-repeat;
    background-position: center;
    background-size: 0.95rem 0.95rem;
}


/* Score no sidebar — sem duplicar marca */
.side-score {
    margin-top: 1rem;
    padding: 0.85rem 0.9rem;
    border-radius: 14px;
    background: #f3f4f7;
    border: 1px solid rgba(28,31,38,0.06);
}
.side-score-label {
    font-size: 0.75rem;
    font-weight: 700;
    color: #8a90a0;
}
.side-score-value {
    font-size: 1.25rem;
    font-weight: 800;
    color: #1c1f26;
    margin-top: 0.2rem;
}
.side-score-note {
    font-size: 0.78rem;
    color: #6b7280;
    margin-top: 0.15rem;
}

/* NÃO injetar "Menu" dentro da sidebar expandida */
[data-testid="stSidebar"]::after,
[data-testid="stSidebar"]::before {
    content: none !important;
}

/* Só o controle COLLAPSED (fora da sidebar) ganha texto Menu */
[data-testid="stSidebar"] [data-testid="stSidebarCollapsedControl"]::after,
[data-testid="stSidebar"] [data-testid="collapsedControl"]::after {
    content: none !important;
}


/* ========== MENU MOBILE VISÍVEL + TEXTO SEMPRE LEGÍVEL ========== */

/* Controle da sidebar recolhida: botão vermelho sólido + texto Menu */
[data-testid="stSidebarCollapsedControl"],
[data-testid="collapsedControl"] {
    position: fixed !important;
    top: 0.5rem !important;
    left: 0.5rem !important;
    z-index: 1000000 !important;
    display: flex !important;
    align-items: center !important;
    gap: 0.3rem !important;
    padding: 0.35rem 0.7rem 0.35rem 0.45rem !important;
    border-radius: 999px !important;
    background: #1c1f26 !important;
    box-shadow: 0 4px 14px rgba(225, 29, 72, 0.35), 0 0 0 3px rgba(225, 29, 72, 0.12) !important;
    opacity: 1 !important;
    visibility: visible !important;
    pointer-events: auto !important;
}

[data-testid="stSidebarCollapsedControl"] button,
[data-testid="collapsedControl"] button,
[data-testid="stSidebarCollapsedControl"] > button,
[data-testid="collapsedControl"] > button {
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
    background: transparent !important;
    border: none !important;
    box-shadow: none !important;
    min-width: 2.4rem !important;
    min-height: 2.4rem !important;
    opacity: 1 !important;
    visibility: visible !important;
}

[data-testid="stSidebarCollapsedControl"] button svg,
[data-testid="collapsedControl"] button svg,
[data-testid="stSidebarCollapsedControl"] svg,
[data-testid="collapsedControl"] svg,
[data-testid="stSidebarCollapsedControl"] path,
[data-testid="collapsedControl"] path {
    fill: #1c1f26 !important;
    color: #1c1f26 !important;
    stroke: #1c1f26 !important;
    opacity: 1 !important;
    visibility: visible !important;
    width: 1.25rem !important;
    height: 1.25rem !important;
}

[data-testid="stSidebarCollapsedControl"]::after,
[data-testid="collapsedControl"]::after {
    content: "Menu" !important;
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
    font-weight: 800 !important;
    font-size: 0.9rem !important;
    letter-spacing: 0.02em;
    font-family: Inter, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    line-height: 1;
    margin-right: 0.25rem;
    opacity: 1 !important;
    visibility: visible !important;
}

/* Também quando Streamlit usa outros testids no mobile */
section[data-testid="stSidebar"] + div [data-testid="stBaseButton-headerNoPadding"],
button[data-testid="baseButton-headerNoPadding"],
[data-testid="stBaseButton-headerNoPadding"] {
    color: #ffffff !important;
    background: #e11d48 !important;
    border-radius: 999px !important;
    min-width: 2.5rem !important;
    min-height: 2.5rem !important;
    box-shadow: 0 6px 18px rgba(225, 29, 72, 0.35) !important;
    opacity: 1 !important;
}
[data-testid="stBaseButton-headerNoPadding"] svg,
button[data-testid="baseButton-headerNoPadding"] svg {
    fill: #ffffff !important;
    color: #ffffff !important;
    stroke: #ffffff !important;
}

/* Mobile: área de toque maior */
@media (max-width: 768px) {
    [data-testid="stSidebarCollapsedControl"],
    [data-testid="collapsedControl"] {
        top: 0.55rem !important;
        left: 0.55rem !important;
        padding: 0.4rem 0.75rem 0.4rem 0.45rem !important;
        background: #e11d48 !important;
        border: 2px solid #ffffff !important;
    }
    [data-testid="stSidebarCollapsedControl"] button,
    [data-testid="collapsedControl"] button {
        min-width: 2.6rem !important;
        min-height: 2.6rem !important;
    }
    [data-testid="stSidebarCollapsedControl"]::after,
    [data-testid="collapsedControl"]::after {
        font-size: 0.95rem !important;
        color: #ffffff !important;
    }
    /* Empurra o header um pouco para não colidir com o botão Menu */
    .page-header {
        padding-top: 0.35rem !important;
    }
}

/* ========== TEXTO SEMPRE VISÍVEL (claro e escuro) ========== */

/* Força contraste no app principal — não some no dark mode do sistema/Streamlit */
[data-testid="stAppViewContainer"] p,
[data-testid="stAppViewContainer"] label,
[data-testid="stAppViewContainer"] h1,
[data-testid="stAppViewContainer"] h2,
[data-testid="stAppViewContainer"] h3,
[data-testid="stAppViewContainer"] h4,
.main .block-container p,
.main .block-container label,
[data-testid="stMarkdownContainer"] p,
[data-testid="stWidgetLabel"],
[data-testid="stWidgetLabel"] p {
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
    opacity: 1 !important;
}
/* spans genéricos: não forçar dentro de botões */
[data-testid="stAppViewContainer"] span:not(button span):not([data-testid="stFormSubmitButton"] *) {
    color: inherit;
}

/* Secundário permanece legível (cinza escuro, não claro) */
.page-sub,
.page-kicker,
.metric-label,
.metric-foot,
.indicator-top,
.indicator-note,
.answer-question,
.answer-action,
.side-score-label,
.side-score-note,
[data-testid="stCaptionContainer"],
[data-testid="stCaptionContainer"] p,
.stCaption {
    color: #4b5563 !important;
    -webkit-text-fill-color: #4b5563 !important;
    opacity: 1 !important;
}

/* Valores e títulos bem escuros */
.metric-value,
.indicator-value,
.answer-value,
.page-header h1,
.side-title {
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
}

/* Inputs legíveis em qualquer tema */
input, textarea, select,
[data-baseweb="input"] input,
[data-baseweb="select"] span,
[data-baseweb="select"] div {
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
    background-color: #ffffff !important;
}

/* Sidebar texto legível */
[data-testid="stSidebar"],
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] span {
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
    opacity: 1 !important;
}

/* Dark mode do sistema: mantém fundo claro do app (evita texto sumir) */
@media (prefers-color-scheme: dark) {
    [data-testid="stAppViewContainer"] {
        color-scheme: light !important;
    }
    .main, .main .block-container,
    [data-testid="stAppViewContainer"] {
        color: #1c1f26 !important;
    }
    /* Botão menu continua vermelho com texto branco */
    [data-testid="stSidebarCollapsedControl"],
    [data-testid="collapsedControl"] {
        background: #e11d48 !important;
    }
    [data-testid="stSidebarCollapsedControl"]::after,
    [data-testid="collapsedControl"]::after,
    [data-testid="stSidebarCollapsedControl"] button,
    [data-testid="collapsedControl"] button {
        color: #ffffff !important;
        -webkit-text-fill-color: #ffffff !important;
    }
    [data-testid="stSidebarCollapsedControl"] svg,
    [data-testid="collapsedControl"] svg {
        fill: #ffffff !important;
        stroke: #ffffff !important;
    }
}

/* Streamlit theme dark class fallback */
[data-theme="dark"] .main,
[data-theme="dark"] [data-testid="stAppViewContainer"],
.stApp[data-theme="dark"] {
    color: #1c1f26 !important;
}


/* ========== DROPDOWN SELECT LEGÍVEL (mobile + desktop) ========== */

/* Painel do menu aberto (Baseweb popover/listbox) */
div[data-baseweb="popover"],
div[data-baseweb="popover"] > div,
div[data-baseweb="menu"],
ul[role="listbox"],
ul[role="listbox"] ul,
[data-baseweb="popover"] [role="listbox"],
div[data-baseweb="popover"] ul {
    background: #ffffff !important;
    background-color: #ffffff !important;
    color: #1c1f26 !important;
    border: 1px solid rgba(28, 31, 38, 0.12) !important;
    border-radius: 14px !important;
    box-shadow: 0 12px 32px rgba(28, 31, 38, 0.14) !important;
    opacity: 1 !important;
}

/* Cada opção */
ul[role="listbox"] li,
[role="option"],
div[data-baseweb="menu"] li,
div[data-baseweb="popover"] li,
div[data-baseweb="popover"] [role="option"],
ul[role="listbox"] li > div,
[role="option"] > div,
[role="option"] span,
[role="option"] p {
    background: #ffffff !important;
    background-color: #ffffff !important;
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
    opacity: 1 !important;
    visibility: visible !important;
    font-weight: 600 !important;
    font-size: 0.95rem !important;
    min-height: 2.6rem !important;
}

/* Hover / selected */
ul[role="listbox"] li:hover,
[role="option"]:hover,
div[data-baseweb="menu"] li:hover,
[role="option"][aria-selected="true"],
ul[role="listbox"] li[aria-selected="true"] {
    background: #f3ff9a !important;
    background-color: #f3ff9a !important;
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
}

/* Highlighted (teclado/foco) */
[role="option"][data-highlighted="true"],
li[aria-selected="true"] {
    background: #d9ff00 !important;
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
}

/* Garante que nenhum tema escuro pinte o popover de preto */
div[data-baseweb="popover"] *,
div[data-baseweb="menu"] *,
ul[role="listbox"] * {
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
}

/* Campo fechado do select */
[data-testid="stSelectbox"] [data-baseweb="select"] > div {
    background: #ffffff !important;
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
    border: 1.5px solid rgba(28, 31, 38, 0.12) !important;
}

[data-testid="stSelectbox"] [data-baseweb="select"] span,
[data-testid="stSelectbox"] [data-baseweb="select"] div {
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
    opacity: 1 !important;
}

/* Seta: branco no preto (mantém) */
[data-testid="stSelectbox"] [data-baseweb="select"] > div::after {
    background-color: #1c1f26 !important;
    background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24'%3E%3Cpath fill='white' d='M7.4 8.6 12 13.2l4.6-4.6L18 10l-6 6-6-6z'/%3E%3C/svg%3E") !important;
    background-repeat: no-repeat !important;
    background-position: center !important;
    background-size: 0.95rem 0.95rem !important;
}


/* ========== BOTÕES: TEXTO SEMPRE LEGÍVEL ========== */

/* Primário / submit: fundo escuro + texto BRANCO */
button[kind="primary"],
button[kind="primaryFormSubmit"],
button[data-testid="baseButton-primary"],
button[data-testid="baseButton-primaryFormSubmit"],
[data-testid="stFormSubmitButton"] button,
div[data-testid="stForm"] button,
.stButton > button[kind="primary"] {
    background: #1c1f26 !important;
    background-color: #1c1f26 !important;
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
    border: none !important;
    opacity: 1 !important;
    visibility: visible !important;
}

button[kind="primary"] *,
button[kind="primaryFormSubmit"] *,
button[data-testid="baseButton-primary"] *,
button[data-testid="baseButton-primaryFormSubmit"] *,
[data-testid="stFormSubmitButton"] button *,
div[data-testid="stForm"] button *,
.stButton > button[kind="primary"] * {
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
    fill: #ffffff !important;
    opacity: 1 !important;
    visibility: visible !important;
}

/* Secundário / apagar: fundo claro + texto ESCURO (nunca preto no preto) */
button[kind="secondary"],
button[kind="secondaryFormSubmit"],
button[data-testid="baseButton-secondary"],
button[data-testid="baseButton-secondaryFormSubmit"],
.stButton > button[kind="secondary"],
.stButton > button:not([kind="primary"]) {
    background: #ffffff !important;
    background-color: #ffffff !important;
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
    border: 1.5px solid rgba(28, 31, 38, 0.16) !important;
    box-shadow: 0 4px 12px rgba(28, 31, 38, 0.06) !important;
    opacity: 1 !important;
}

button[kind="secondary"] *,
button[kind="secondaryFormSubmit"] *,
button[data-testid="baseButton-secondary"] *,
button[data-testid="baseButton-secondaryFormSubmit"] *,
.stButton > button[kind="secondary"] *,
.stButton > button:not([kind="primary"]) * {
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
    fill: #1c1f26 !important;
    opacity: 1 !important;
    visibility: visible !important;
}

/* Botões de apagar: vermelho suave com texto legível */
.stButton > button[kind="secondary"]:hover,
button[data-testid="baseButton-secondary"]:hover {
    background: #ffe4e8 !important;
    color: #9b1c2e !important;
    -webkit-text-fill-color: #9b1c2e !important;
    border-color: rgba(225, 29, 72, 0.35) !important;
}

/* Fallback geral: qualquer botão do main */
.main button,
[data-testid="stAppViewContainer"] button {
    opacity: 1 !important;
}
.main button p,
.main button span,
.main button div,
[data-testid="stAppViewContainer"] button p,
[data-testid="stAppViewContainer"] button span {
    opacity: 1 !important;
    visibility: visible !important;
}


/* ÚLTIMA CAMADA: botões nunca ficam com texto invisível */
button[kind="primary"],
button[kind="primaryFormSubmit"],
button[data-testid="baseButton-primary"],
button[data-testid="baseButton-primaryFormSubmit"],
[data-testid="stFormSubmitButton"] button {
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
    background: #1c1f26 !important;
}
button[kind="primary"] span,
button[kind="primary"] p,
button[kind="primaryFormSubmit"] span,
button[kind="primaryFormSubmit"] p,
button[data-testid="baseButton-primary"] span,
button[data-testid="baseButton-primary"] p,
button[data-testid="baseButton-primaryFormSubmit"] span,
button[data-testid="baseButton-primaryFormSubmit"] p,
[data-testid="stFormSubmitButton"] button span,
[data-testid="stFormSubmitButton"] button p {
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
}

button[kind="secondary"],
button[data-testid="baseButton-secondary"],
.stButton > button[kind="secondary"] {
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
    background: #ffffff !important;
    border: 1.5px solid rgba(28,31,38,0.16) !important;
}
button[kind="secondary"] span,
button[kind="secondary"] p,
button[data-testid="baseButton-secondary"] span,
button[data-testid="baseButton-secondary"] p,
.stButton > button[kind="secondary"] span,
.stButton > button[kind="secondary"] p {
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
}


/* ========== BOTÃO PDF + MENU MOBILE VISÍVEIS ========== */

/* Download / primary: fundo escuro, texto branco FORÇADO */
div[data-testid="stDownloadButton"] button,
div[data-testid="stDownloadButton"] button[kind="primary"],
button[data-testid="baseButton-primary"],
button[kind="primary"],
.stDownloadButton > button {
    background: #1c1f26 !important;
    background-color: #1c1f26 !important;
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
    border: none !important;
    border-radius: 999px !important;
    font-weight: 800 !important;
    min-height: 2.85rem !important;
    opacity: 1 !important;
    visibility: visible !important;
    box-shadow: 0 6px 16px rgba(28, 31, 38, 0.14) !important;
}
div[data-testid="stDownloadButton"] button *,
div[data-testid="stDownloadButton"] button p,
div[data-testid="stDownloadButton"] button span,
button[data-testid="baseButton-primary"] *,
button[kind="primary"] *,
.stDownloadButton > button * {
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
    opacity: 1 !important;
    visibility: visible !important;
}
div[data-testid="stDownloadButton"] button:hover,
button[kind="primary"]:hover {
    background: #2d323c !important;
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
}

/* Alternativa se primary ficar ilegível: versão lime */
div[data-testid="stDownloadButton"] button:focus {
    outline: 2px solid #d9ff00 !important;
}

/* MENU recolhido no mobile — pílula vermelha bem visível */
[data-testid="stSidebarCollapsedControl"],
[data-testid="collapsedControl"],
section[data-testid="stSidebar"] ~ div [data-testid="stSidebarCollapsedControl"] {
    position: fixed !important;
    top: 0.55rem !important;
    left: 0.55rem !important;
    z-index: 2147483647 !important;
    display: inline-flex !important;
    align-items: center !important;
    justify-content: center !important;
    gap: 0.25rem !important;
    min-width: 2.75rem !important;
    min-height: 2.75rem !important;
    padding: 0.35rem 0.75rem !important;
    border-radius: 999px !important;
    background: #e11d48 !important;
    border: 2px solid #ffffff !important;
    box-shadow: 0 8px 22px rgba(225, 29, 72, 0.4) !important;
    opacity: 1 !important;
    visibility: visible !important;
    pointer-events: auto !important;
}

[data-testid="stSidebarCollapsedControl"] button,
[data-testid="collapsedControl"] button,
[data-testid="stSidebarCollapsedControl"] > button,
[data-testid="collapsedControl"] > button {
    background: transparent !important;
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
    border: none !important;
    box-shadow: none !important;
    min-width: 1.75rem !important;
    min-height: 1.75rem !important;
    opacity: 1 !important;
    visibility: visible !important;
}

[data-testid="stSidebarCollapsedControl"] svg,
[data-testid="collapsedControl"] svg,
[data-testid="stSidebarCollapsedControl"] path,
[data-testid="collapsedControl"] path,
[data-testid="stSidebarCollapsedControl"] button svg,
[data-testid="collapsedControl"] button svg {
    fill: #ffffff !important;
    color: #ffffff !important;
    stroke: #ffffff !important;
    opacity: 1 !important;
    visibility: visible !important;
    width: 1.35rem !important;
    height: 1.35rem !important;
}

/* Texto "Menu" ao lado do >> */
[data-testid="stSidebarCollapsedControl"]::after,
[data-testid="collapsedControl"]::after {
    content: "Menu" !important;
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
    font-weight: 800 !important;
    font-size: 0.9rem !important;
    font-family: Inter, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif !important;
    letter-spacing: 0.02em;
    margin-left: 0.1rem;
    opacity: 1 !important;
    visibility: visible !important;
}

@media (max-width: 768px) {
    [data-testid="stSidebarCollapsedControl"],
    [data-testid="collapsedControl"] {
        top: 0.45rem !important;
        left: 0.45rem !important;
        min-height: 3rem !important;
        padding: 0.4rem 0.85rem 0.4rem 0.55rem !important;
        background: #e11d48 !important;
        border: 2px solid #fff !important;
    }
}


/* ========== MENU MOBILE: ícone preto + sombra vermelha leve ========== */
[data-testid="stSidebarCollapsedControl"],
[data-testid="collapsedControl"] {
    position: fixed !important;
    top: 0.5rem !important;
    left: 0.5rem !important;
    z-index: 2147483647 !important;
    display: inline-flex !important;
    align-items: center !important;
    justify-content: center !important;
    gap: 0.3rem !important;
    min-width: 2.6rem !important;
    min-height: 2.6rem !important;
    padding: 0.35rem 0.7rem 0.35rem 0.45rem !important;
    border-radius: 999px !important;
    background: #ffffff !important;
    border: 1.5px solid rgba(225, 29, 72, 0.55) !important;
    box-shadow: 0 0 0 2px rgba(225, 29, 72, 0.10), 0 3px 10px rgba(225, 29, 72, 0.18) !important;
    opacity: 1 !important;
    visibility: visible !important;
    pointer-events: auto !important;
}

[data-testid="stSidebarCollapsedControl"] button,
[data-testid="collapsedControl"] button,
[data-testid="stSidebarCollapsedControl"] > button,
[data-testid="collapsedControl"] > button {
    background: transparent !important;
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
    border: none !important;
    box-shadow: none !important;
    min-width: 1.6rem !important;
    min-height: 1.6rem !important;
    opacity: 1 !important;
    visibility: visible !important;
}

[data-testid="stSidebarCollapsedControl"] svg,
[data-testid="collapsedControl"] svg,
[data-testid="stSidebarCollapsedControl"] path,
[data-testid="collapsedControl"] path,
[data-testid="stSidebarCollapsedControl"] button svg,
[data-testid="collapsedControl"] button svg {
    fill: #ffffff !important;
    color: #ffffff !important;
    stroke: #ffffff !important;
    opacity: 1 !important;
    visibility: visible !important;
    width: 1.25rem !important;
    height: 1.25rem !important;
}

[data-testid="stSidebarCollapsedControl"]::after,
[data-testid="collapsedControl"]::after {
    content: "Menu" !important;
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
    font-weight: 800 !important;
    font-size: 0.88rem !important;
    font-family: Inter, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif !important;
    letter-spacing: 0.02em;
    opacity: 1 !important;
    visibility: visible !important;
}

@media (max-width: 768px) {
    [data-testid="stSidebarCollapsedControl"],
    [data-testid="collapsedControl"] {
        top: 0.4rem !important;
        left: 0.4rem !important;
        min-height: 2.85rem !important;
        background: #1c1f26 !important;
        box-shadow: 0 5px 16px rgba(225, 29, 72, 0.4), 0 0 0 3px rgba(225, 29, 72, 0.14) !important;
    }
}



/* ========== ÍCONE MENU: fundo branco + contorno vermelho + >> preto ========== */
[data-testid="stSidebarCollapsedControl"],
[data-testid="collapsedControl"],
[data-testid="stSidebarCollapsedControl"] > div,
[data-testid="collapsedControl"] > div {
    position: fixed !important;
    top: 0.45rem !important;
    left: 0.45rem !important;
    z-index: 2147483647 !important;
    display: inline-flex !important;
    align-items: center !important;
    justify-content: center !important;
    background: #ffffff !important;
    background-color: #ffffff !important;
    border-radius: 999px !important;
    border: 1.5px solid rgba(225, 29, 72, 0.55) !important;
    box-shadow: 0 0 0 2px rgba(225, 29, 72, 0.10), 0 3px 10px rgba(225, 29, 72, 0.18) !important;
    min-width: 2.75rem !important;
    min-height: 2.75rem !important;
    padding: 0.3rem 0.65rem !important;
    opacity: 1 !important;
    visibility: visible !important;
}

[data-testid="stSidebarCollapsedControl"] button,
[data-testid="collapsedControl"] button,
[data-testid="stSidebarCollapsedControl"] button[kind="header"],
[data-testid="stSidebarCollapsedControl"] button[kind="headerNoPadding"],
[data-testid="stBaseButton-headerNoPadding"],
button[data-testid="baseButton-headerNoPadding"],
button[kind="headerNoPadding"] {
    background: #ffffff !important;
    background-color: #ffffff !important;
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
    border: none !important;
    border-radius: 999px !important;
    box-shadow: none !important;
    opacity: 1 !important;
    visibility: visible !important;
    min-width: 2.4rem !important;
    min-height: 2.4rem !important;
}

/* >> preto no fundo branco */
[data-testid="stSidebarCollapsedControl"] svg,
[data-testid="collapsedControl"] svg,
[data-testid="stSidebarCollapsedControl"] path,
[data-testid="collapsedControl"] path,
[data-testid="stSidebarCollapsedControl"] button svg,
[data-testid="collapsedControl"] button svg,
[data-testid="stBaseButton-headerNoPadding"] svg,
button[kind="headerNoPadding"] svg,
[data-testid="stSidebarCollapsedControl"] span,
[data-testid="collapsedControl"] span {
    fill: #1c1f26 !important;
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
    stroke: #1c1f26 !important;
    opacity: 1 !important;
    visibility: visible !important;
    width: 1.35rem !important;
    height: 1.35rem !important;
    font-weight: 800 !important;
}

[data-testid="stSidebarCollapsedControl"]::after,
[data-testid="collapsedControl"]::after {
    content: "Menu" !important;
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
    font-weight: 800 !important;
    font-size: 0.88rem !important;
    font-family: Inter, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif !important;
    margin-left: 0.15rem;
    opacity: 1 !important;
    visibility: visible !important;
}

[data-testid="stSidebarCollapsedControl"]::before,
[data-testid="collapsedControl"]::before {
    content: none !important;
    display: none !important;
}

@media (max-width: 768px) {
    [data-testid="stSidebarCollapsedControl"],
    [data-testid="collapsedControl"] {
        top: 0.35rem !important;
        left: 0.35rem !important;
        min-width: 3rem !important;
        min-height: 3rem !important;
        background: #ffffff !important;
        border: 1.5px solid rgba(225, 29, 72, 0.55) !important;
        box-shadow: 0 0 0 2px rgba(225, 29, 72, 0.10), 0 3px 10px rgba(225, 29, 72, 0.18) !important;
    }
}


/* ========== MENU ESTILO REFERÊNCIA (rail / painel claro) ========== */

[data-testid="stSidebar"] {
    background: #f4f5f7 !important;
    border-right: none !important;
}
[data-testid="stSidebar"] > div:first-child {
    background: #f4f5f7 !important;
    padding-top: 0.75rem !important;
}
[data-testid="stSidebar"] [data-testid="stVerticalBlock"] {
    gap: 0.35rem !important;
}

/* Card do menu — painel branco arredondado */
[data-testid="stSidebar"] [data-testid="stVerticalBlockBorderWrapper"],
[data-testid="stSidebar"] section {
    background: transparent !important;
}

.side-brand {
    padding: 0.55rem 0.35rem 0.85rem 0.45rem;
}
.side-menu-label {
    display: inline-flex;
    align-items: center;
    gap: 0.45rem;
    font-size: 1.05rem;
    font-weight: 800;
    color: #1c1f26;
    letter-spacing: -0.02em;
}
.side-spark {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 1.55rem;
    height: 1.55rem;
    border-radius: 50%;
    background: #ffffff;
    color: #1c1f26;
    font-size: 0.85rem;
    box-shadow: 0 2px 8px rgba(28, 31, 38, 0.08);
}
.side-title { display: none !important; }
.side-menu-dot { display: none !important; }

/* Radio do menu = lista limpa */
[data-testid="stSidebar"] [data-testid="stRadio"] > div[role="radiogroup"] {
    display: flex !important;
    flex-direction: column !important;
    gap: 0.28rem !important;
    background: transparent !important;
    padding: 0.15rem 0 !important;
    border: none !important;
    border-radius: 0 !important;
    width: 100% !important;
    box-shadow: none !important;
}

[data-testid="stSidebar"] [data-testid="stRadio"] > div[role="radiogroup"] > label {
    width: 100% !important;
    margin: 0 !important;
    padding: 0.72rem 0.95rem !important;
    border-radius: 999px !important;
    background: transparent !important;
    border: none !important;
    min-height: 2.65rem !important;
    justify-content: flex-start !important;
    box-shadow: none !important;
    transition: background 0.15s ease, color 0.15s ease, box-shadow 0.15s ease;
}

/* Item inativo */
[data-testid="stSidebar"] [data-testid="stRadio"] > div[role="radiogroup"] > label p,
[data-testid="stSidebar"] [data-testid="stRadio"] > div[role="radiogroup"] > label span,
[data-testid="stSidebar"] [data-testid="stRadio"] label,
[data-testid="stSidebar"] [data-testid="stRadio"] label * {
    color: #3a3f4b !important;
    -webkit-text-fill-color: #3a3f4b !important;
    font-weight: 600 !important;
    font-size: 0.95rem !important;
    opacity: 1 !important;
    visibility: visible !important;
}

/* Hover */
[data-testid="stSidebar"] [data-testid="stRadio"] > div[role="radiogroup"] > label:hover {
    background: #ffffff !important;
    box-shadow: 0 2px 10px rgba(28, 31, 38, 0.06) !important;
}

/* Item ATIVO — pílula preta (como Threads na referência) */
[data-testid="stSidebar"] [data-testid="stRadio"] > div[role="radiogroup"] > label[data-checked="true"],
[data-testid="stSidebar"] [data-testid="stRadio"] > div[role="radiogroup"] > label:has(input:checked),
[data-testid="stSidebar"] [data-testid="stRadio"] [aria-checked="true"] {
    background: #1c1f26 !important;
    box-shadow: 0 6px 18px rgba(28, 31, 38, 0.16) !important;
}
[data-testid="stSidebar"] [data-testid="stRadio"] > div[role="radiogroup"] > label[data-checked="true"] p,
[data-testid="stSidebar"] [data-testid="stRadio"] > div[role="radiogroup"] > label[data-checked="true"] span,
[data-testid="stSidebar"] [data-testid="stRadio"] > div[role="radiogroup"] > label[data-checked="true"] *,
[data-testid="stSidebar"] [data-testid="stRadio"] > div[role="radiogroup"] > label:has(input:checked) p,
[data-testid="stSidebar"] [data-testid="stRadio"] > div[role="radiogroup"] > label:has(input:checked) span,
[data-testid="stSidebar"] [data-testid="stRadio"] > div[role="radiogroup"] > label:has(input:checked) *,
[data-testid="stSidebar"] [data-testid="stRadio"] [aria-checked="true"] p,
[data-testid="stSidebar"] [data-testid="stRadio"] [aria-checked="true"] * {
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
    font-weight: 700 !important;
}

/* Esconde círculo do radio */
[data-testid="stSidebar"] [data-testid="stRadio"] input,
[data-testid="stSidebar"] [data-baseweb="radio"] > div:first-child,
[data-testid="stSidebar"] [data-baseweb="radio"] > span:first-child {
    display: none !important;
}

/* Score card no estilo da referência */
.side-score {
    margin-top: 1rem;
    padding: 0.9rem 1rem;
    border-radius: 18px;
    background: #ffffff;
    border: none;
    box-shadow: 0 4px 16px rgba(28, 31, 38, 0.06);
}
.side-score-label {
    font-size: 0.72rem;
    font-weight: 700;
    color: #8a90a0;
    text-transform: uppercase;
    letter-spacing: 0.04em;
}
.side-score-value {
    font-size: 1.35rem;
    font-weight: 800;
    color: #1c1f26;
    margin-top: 0.25rem;
}
.side-score-note {
    font-size: 0.8rem;
    color: #6b7280;
    margin-top: 0.15rem;
}

/* Card da conta */
[data-testid="stSidebar"] .stMarkdown + div {
    border-radius: 16px;
}

/* Botão Sair — secundário limpo */
[data-testid="stSidebar"] .stButton > button {
    background: #ffffff !important;
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
    border: none !important;
    border-radius: 999px !important;
    box-shadow: 0 2px 10px rgba(28, 31, 38, 0.06) !important;
    font-weight: 700 !important;
}
[data-testid="stSidebar"] .stButton > button * {
    color: #1c1f26 !important;
    -webkit-text-fill-color: #1c1f26 !important;
}
[data-testid="stSidebar"] .stButton > button:hover {
    background: #1c1f26 !important;
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
}
[data-testid="stSidebar"] .stButton > button:hover * {
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
}

/* Remove fundo lime antigo do item ativo na sidebar */
[data-testid="stSidebar"] .stRadio [role="radiogroup"] label[data-checked="true"],
[data-testid="stSidebar"] .stRadio [aria-checked="true"] {
    background: #1c1f26 !important;
}

</style>
""",
    unsafe_allow_html=True,
)

# ====================== BANCO DE DADOS ======================
def _ensure_column(cur, tabela: str, coluna: str, tipo_sql: str = "INTEGER"):
    cur.execute(f"PRAGMA table_info({tabela})")
    cols = {row[1] for row in cur.fetchall()}
    if coluna not in cols:
        cur.execute(f"ALTER TABLE {tabela} ADD COLUMN {coluna} {tipo_sql}")


def init_db():
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT NOT NULL UNIQUE,
            nome TEXT NOT NULL,
            senha_hash TEXT NOT NULL,
            salt TEXT NOT NULL,
            criado_em TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS transacoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            data TEXT NOT NULL,
            descricao TEXT,
            categoria TEXT,
            valor REAL NOT NULL,
            tipo TEXT,
            cartao TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS investimentos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            data TEXT,
            tipo TEXT,
            valor REAL,
            rentabilidade TEXT,
            descricao TEXT,
            status TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS dividas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            data TEXT,
            credor TEXT,
            tipo TEXT,
            saldo_original REAL,
            desconto REAL,
            saldo_negociado REAL,
            parcela_possivel REAL,
            vencimento TEXT,
            prioridade TEXT,
            consequencia TEXT,
            status TEXT,
            proxima_acao TEXT,
            anotacoes TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS metas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            data TEXT,
            nome TEXT,
            valor_meta REAL,
            valor_atual REAL,
            aporte_mensal REAL,
            prazo TEXT,
            status TEXT,
            anotacoes TEXT
        )
    """)

    for tabela in ("transacoes", "investimentos", "dividas", "metas"):
        try:
            _ensure_column(cur, tabela, "user_id", "INTEGER")
        except Exception:
            pass

    cur.execute("""
        CREATE TABLE IF NOT EXISTS sessoes (
            token TEXT PRIMARY KEY,
            user_id INTEGER NOT NULL,
            criado_em TEXT NOT NULL,
            expira_em TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS redefinicoes_senha (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            codigo_hash TEXT NOT NULL,
            salt TEXT NOT NULL,
            criado_em TEXT NOT NULL,
            expira_em TEXT NOT NULL,
            usado INTEGER NOT NULL DEFAULT 0
        )
    """)

    conn.commit()
    conn.close()


# ====================== AUTENTICAÇÃO ======================
def _hash_senha(senha: str, salt: str | None = None) -> tuple[str, str]:
    if not salt:
        salt = py_secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", senha.encode("utf-8"), salt.encode("utf-8"), 120_000)
    return digest.hex(), salt


def criar_usuario(email: str, nome: str, senha: str) -> tuple[bool, str]:
    email_n = limpar_texto(email).lower()
    nome_n = limpar_texto(nome)
    if not email_n or "@" not in email_n:
        return False, "Informe um e-mail válido."
    if len(limpar_texto(senha)) < 6:
        return False, "A senha precisa ter pelo menos 6 caracteres."
    if not nome_n:
        return False, "Informe seu nome."
    senha_hash, salt = _hash_senha(senha)
    conn = get_conn()
    try:
        cur = conn.execute(
            "INSERT INTO usuarios (email, nome, senha_hash, salt, criado_em) VALUES (?, ?, ?, ?, ?)",
            (email_n, nome_n, senha_hash, salt, date.today().isoformat()),
        )
        novo_id = cur.lastrowid
        # Se for o primeiro usuário, associa dados antigos (sem dono) a esta conta
        total_users = conn.execute("SELECT COUNT(*) FROM usuarios").fetchone()[0]
        if total_users == 1 and novo_id:
            for tabela in ("transacoes", "investimentos", "dividas", "metas"):
                try:
                    conn.execute(
                        f"UPDATE {tabela} SET user_id = ? WHERE user_id IS NULL",
                        (novo_id,),
                    )
                except Exception:
                    pass
        conn.commit()
        return True, "Conta criada com sucesso."
    except sqlite3.IntegrityError:
        return False, "Este e-mail já está cadastrado."
    finally:
        conn.close()


def autenticar_usuario(email: str, senha: str):
    email_n = limpar_texto(email).lower()
    conn = get_conn()
    cur = conn.execute(
        "SELECT id, email, nome, senha_hash, salt FROM usuarios WHERE email = ?",
        (email_n,),
    )
    row = cur.fetchone()
    conn.close()
    if not row:
        return None
    uid, em, nome, senha_hash, salt = row
    teste, _ = _hash_senha(senha, salt)
    if py_secrets.compare_digest(teste, senha_hash):
        return {"id": int(uid), "email": em, "nome": nome}
    return None


def contar_usuarios() -> int:
    conn = get_conn()
    try:
        n = conn.execute("SELECT COUNT(*) FROM usuarios").fetchone()[0]
    except Exception:
        n = 0
    conn.close()
    return int(n)


def usuario_atual():
    return st.session_state.get("usuario")


def user_id_atual() -> int | None:
    u = usuario_atual()
    return int(u["id"]) if u else None


COOKIE_SESSAO = "df_sessao"
DIAS_SESSAO = 90


def _cookie_manager():
    """Cookie no navegador para manter login entre reloads / sono do app."""
    try:
        import extra_streamlit_components as stx
        # key estável evita recriar o component a cada rerun
        return stx.CookieManager(key="df_cookie_mgr")
    except Exception:
        return None


def criar_sessao(user_id: int, dias: int = DIAS_SESSAO) -> str:
    token = py_secrets.token_urlsafe(32)
    agora = datetime.now()
    expira = agora + timedelta(days=dias)
    conn = get_conn()
    conn.execute(
        "INSERT INTO sessoes (token, user_id, criado_em, expira_em) VALUES (?, ?, ?, ?)",
        (token, int(user_id), agora.isoformat(timespec="seconds"), expira.isoformat(timespec="seconds")),
    )
    # limpa sessões vencidas
    conn.execute("DELETE FROM sessoes WHERE expira_em < ?", (agora.isoformat(timespec="seconds"),))
    conn.commit()
    conn.close()
    return token


def usuario_por_token(token: str | None):
    if not token:
        return None
    agora = datetime.now().isoformat(timespec="seconds")
    conn = get_conn()
    row = conn.execute(
        """
        SELECT u.id, u.email, u.nome
        FROM sessoes s
        JOIN usuarios u ON u.id = s.user_id
        WHERE s.token = ? AND s.expira_em >= ?
        """,
        (token, agora),
    ).fetchone()
    conn.close()
    if not row:
        return None
    return {"id": int(row[0]), "email": row[1], "nome": row[2]}


def revogar_sessao(token: str | None):
    if not token:
        return
    conn = get_conn()
    conn.execute("DELETE FROM sessoes WHERE token = ?", (token,))
    conn.commit()
    conn.close()


def _salvar_token_cookie(token: str):
    cm = _cookie_manager()
    if cm is not None:
        try:
            cm.set(COOKIE_SESSAO, token, expires_at=datetime.now() + timedelta(days=DIAS_SESSAO))
        except Exception:
            pass
    # fallback: query param (sobrevive a refresh na mesma aba)
    try:
        st.query_params[COOKIE_SESSAO] = token
    except Exception:
        pass
    st.session_state["_token_sessao"] = token


def _ler_token_cookie() -> str | None:
    if st.session_state.get("_token_sessao"):
        return st.session_state.get("_token_sessao")
    # query params
    try:
        qp = st.query_params.get(COOKIE_SESSAO)
        if qp:
            return qp
    except Exception:
        pass
    cm = _cookie_manager()
    if cm is not None:
        try:
            val = cm.get(COOKIE_SESSAO)
            if val:
                return val
        except Exception:
            pass
    return None


def _limpar_token_cookie():
    st.session_state.pop("_token_sessao", None)
    try:
        if COOKIE_SESSAO in st.query_params:
            del st.query_params[COOKIE_SESSAO]
    except Exception:
        pass
    cm = _cookie_manager()
    if cm is not None:
        try:
            cm.delete(COOKIE_SESSAO)
        except Exception:
            pass


def renovar_sessao(token: str | None, dias: int = DIAS_SESSAO):
    """Estende a validade a cada acesso (sessão deslizante)."""
    if not token:
        return
    nova_expira = (datetime.now() + timedelta(days=dias)).isoformat(timespec="seconds")
    conn = get_conn()
    try:
        conn.execute(
            "UPDATE sessoes SET expira_em = ? WHERE token = ?",
            (nova_expira, token),
        )
        conn.commit()
    except Exception:
        pass
    finally:
        conn.close()
    # renova cookie também
    cm = _cookie_manager()
    if cm is not None:
        try:
            cm.set(COOKIE_SESSAO, token, expires_at=datetime.now() + timedelta(days=dias))
        except Exception:
            pass


def restaurar_sessao_se_houver() -> bool:
    """Se não houver usuário na memória, tenta cookie/token persistente."""
    if st.session_state.get("usuario"):
        # renova em acessos seguintes
        token = st.session_state.get("_token_sessao") or _ler_token_cookie()
        if token:
            renovar_sessao(token)
        return True
    token = _ler_token_cookie()
    user = usuario_por_token(token)
    if user:
        st.session_state.usuario = user
        st.session_state["_token_sessao"] = token
        if "pagina_atual" not in st.session_state:
            st.session_state.pagina_atual = "Dashboard"
        renovar_sessao(token)
        return True
    return False


def iniciar_sessao_usuario(user: dict):
    token = criar_sessao(user["id"])
    st.session_state.usuario = user
    st.session_state.pagina_atual = "Dashboard"
    _salvar_token_cookie(token)


def fazer_logout():
    token = st.session_state.get("_token_sessao") or _ler_token_cookie()
    revogar_sessao(token)
    _limpar_token_cookie()
    for k in ("usuario", "pagina_atual", "menu_mounted", "_token_sessao"):
        if k in st.session_state:
            del st.session_state[k]


def _usuario_por_email(email: str):
    email_n = limpar_texto(email).lower()
    if not email_n or "@" not in email_n:
        return None
    conn = get_conn()
    row = conn.execute(
        "SELECT id, email, nome FROM usuarios WHERE email = ?",
        (email_n,),
    ).fetchone()
    conn.close()
    if not row:
        return None
    return {"id": int(row[0]), "email": row[1], "nome": row[2]}


def _gerar_codigo_recuperacao() -> str:
    return f"{py_secrets.randbelow(1_000_000):06d}"


def solicitar_redefinicao(email: str) -> tuple[bool, str, str | None]:
    """Gera código de 6 dígitos válido por 15 minutos."""
    user = _usuario_por_email(email)
    msg_padrao = "Se este e-mail estiver cadastrado, um código de recuperação foi gerado."
    if not user:
        return True, msg_padrao, None

    codigo = _gerar_codigo_recuperacao()
    codigo_hash, salt = _hash_senha(codigo)
    agora = datetime.now()
    expira = agora + timedelta(minutes=15)

    conn = get_conn()
    conn.execute(
        "UPDATE redefinicoes_senha SET usado = 1 WHERE user_id = ? AND usado = 0",
        (user["id"],),
    )
    conn.execute(
        """
        INSERT INTO redefinicoes_senha (user_id, codigo_hash, salt, criado_em, expira_em, usado)
        VALUES (?, ?, ?, ?, ?, 0)
        """,
        (
            user["id"],
            codigo_hash,
            salt,
            agora.isoformat(timespec="seconds"),
            expira.isoformat(timespec="seconds"),
        ),
    )
    conn.commit()
    conn.close()

    enviado = _enviar_email_recuperacao(user["email"], user["nome"], codigo)
    if enviado:
        return True, "Enviamos um código de 6 dígitos para o seu e-mail. Ele vale por 15 minutos.", None
    return True, "Código gerado. Use-o abaixo para criar uma nova senha (válido por 15 minutos).", codigo


def _enviar_email_recuperacao(email: str, nome: str, codigo: str) -> bool:
    """Tenta enviar por Resend (RESEND_API_KEY nos Secrets)."""
    api_key = _segredo("RESEND_API_KEY")
    de = _segredo("EMAIL_FROM") or "Meu Dinheiro <onboarding@resend.dev>"
    if not api_key:
        return False
    try:
        import urllib.request
        import json as _json
        payload = _json.dumps({
            "from": de,
            "to": [email],
            "subject": "Código para redefinir sua senha",
            "text": (
                f"Olá, {nome}!\n\n"
                f"Seu código de recuperação é: {codigo}\n\n"
                f"Ele vale por 15 minutos. Se você não pediu isso, ignore este e-mail.\n"
            ),
        }).encode("utf-8")
        req = urllib.request.Request(
            "https://api.resend.com/emails",
            data=payload,
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            },
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=12) as resp:
            return 200 <= getattr(resp, "status", 200) < 300
    except Exception:
        return False


def redefinir_senha_com_codigo(email: str, codigo: str, nova_senha: str) -> tuple[bool, str]:
    if len(limpar_texto(nova_senha)) < 6:
        return False, "A nova senha precisa ter pelo menos 6 caracteres."
    user = _usuario_por_email(email)
    if not user:
        return False, "Não foi possível redefinir. Verifique o e-mail e o código."

    agora = datetime.now().isoformat(timespec="seconds")
    conn = get_conn()
    rows = conn.execute(
        """
        SELECT id, codigo_hash, salt FROM redefinicoes_senha
        WHERE user_id = ? AND usado = 0 AND expira_em >= ?
        ORDER BY id DESC LIMIT 5
        """,
        (user["id"], agora),
    ).fetchall()

    match_id = None
    for rid, codigo_hash, salt in rows:
        teste, _ = _hash_senha(limpar_texto(codigo), salt)
        if py_secrets.compare_digest(teste, codigo_hash):
            match_id = rid
            break

    if match_id is None:
        conn.close()
        return False, "Código inválido ou expirado. Solicite um novo."

    nova_hash, novo_salt = _hash_senha(nova_senha)
    conn.execute(
        "UPDATE usuarios SET senha_hash = ?, salt = ? WHERE id = ?",
        (nova_hash, novo_salt, user["id"]),
    )
    conn.execute("UPDATE redefinicoes_senha SET usado = 1 WHERE user_id = ?", (user["id"],))
    conn.execute("DELETE FROM sessoes WHERE user_id = ?", (user["id"],))
    conn.commit()
    conn.close()
    return True, "Senha alterada com sucesso. Entre com a nova senha."


def render_login_page():
    """Tela de entrada — login, criação de conta e recuperação de senha."""
    _cookie_manager()

    st.markdown(
        """
        <div style="max-width:420px;margin:1.5rem auto 0;">
            <div style="text-align:center;margin-bottom:1.25rem;">
                <div style="display:inline-flex;align-items:center;gap:0.4rem;
                    background:#fff;border:1px solid rgba(28,31,38,0.08);
                    border-radius:999px;padding:0.35rem 0.85rem;font-size:0.78rem;
                    font-weight:800;color:#6b7280;letter-spacing:0.04em;text-transform:uppercase;">
                    <span style="width:0.5rem;height:0.5rem;border-radius:50%;background:#d9ff00;display:inline-block;"></span>
                    Financeiro pessoal
                </div>
                <h1 style="font-size:1.85rem;font-weight:800;color:#1c1f26;margin:0.85rem 0 0.35rem;letter-spacing:-0.03em;">
                    Meu dinheiro
                </h1>
                <p style="color:#6b7280;font-size:0.95rem;margin:0;">
                    Entre na sua conta. A sessão fica salva por 90 dias e renova a cada acesso.
                </p>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    col_l, col_c, col_r = st.columns([1, 1.35, 1])
    with col_c:
        tem_usuarios = contar_usuarios() > 0
        if "login_modo_ui" not in st.session_state:
            st.session_state.login_modo_ui = "Criar conta" if not tem_usuarios else "Entrar"
        if st.session_state.get("recup_etapa") in ("codigo", "pedido"):
            st.session_state.login_modo_ui = "Esqueci a senha"

        opcoes = ["Entrar", "Criar conta", "Esqueci a senha"]
        if st.session_state.login_modo_ui not in opcoes:
            st.session_state.login_modo_ui = opcoes[0]

        modo = st.radio(
            "Acesso",
            opcoes,
            index=opcoes.index(st.session_state.login_modo_ui),
            horizontal=True,
            label_visibility="collapsed",
            key="login_modo_radio",
        )
        st.session_state.login_modo_ui = modo

        if modo != "Esqueci a senha":
            st.session_state.pop("recup_etapa", None)
            st.session_state.pop("recup_email", None)
            st.session_state.pop("recup_codigo_demo", None)
            st.session_state.pop("recup_msg", None)

        if modo == "Entrar":
            with st.form("form_login", clear_on_submit=False):
                email = st.text_input("E-mail", placeholder="voce@email.com")
                senha = st.text_input("Senha", type="password", placeholder="Sua senha")
                ok = st.form_submit_button("Entrar", use_container_width=True)
                if ok:
                    user = autenticar_usuario(email, senha)
                    if user:
                        iniciar_sessao_usuario(user)
                        st.success(f"Olá, {user['nome'].split()[0]}!")
                        st.rerun()
                    else:
                        st.error("E-mail ou senha incorretos.")
            if st.button("Esqueci minha senha", key="btn_esqueci_senha", use_container_width=True):
                st.session_state.login_modo_ui = "Esqueci a senha"
                st.session_state.recup_etapa = "pedido"
                st.rerun()

        elif modo == "Criar conta":
            with st.form("form_registro", clear_on_submit=False):
                nome = st.text_input("Seu nome", placeholder="Como quer ser chamado")
                email = st.text_input("E-mail", placeholder="voce@email.com")
                senha = st.text_input("Senha", type="password", placeholder="Mínimo 6 caracteres")
                senha2 = st.text_input("Confirmar senha", type="password", placeholder="Repita a senha")
                ok = st.form_submit_button("Criar minha conta", use_container_width=True)
                if ok:
                    if senha != senha2:
                        st.error("As senhas não coincidem.")
                    else:
                        ok_c, msg = criar_usuario(email, nome, senha)
                        if ok_c:
                            user = autenticar_usuario(email, senha)
                            if user:
                                iniciar_sessao_usuario(user)
                                st.success(msg)
                                st.rerun()
                        else:
                            st.error(msg)

        else:
            st.markdown("##### Recuperar acesso")
            st.caption("Informe o e-mail da conta. Um código de 6 dígitos será gerado (válido por 15 minutos).")

            if st.session_state.get("recup_etapa") != "codigo":
                with st.form("form_recup_pedido", clear_on_submit=False):
                    email_r = st.text_input("E-mail da conta", placeholder="voce@email.com")
                    ok_r = st.form_submit_button("Enviar código", use_container_width=True)
                    if ok_r:
                        _ok_s, msg_s, codigo_demo = solicitar_redefinicao(email_r)
                        st.session_state.recup_email = limpar_texto(email_r).lower()
                        st.session_state.recup_etapa = "codigo"
                        st.session_state.recup_msg = msg_s
                        st.session_state.recup_codigo_demo = codigo_demo
                        st.rerun()

            if st.session_state.get("recup_etapa") == "codigo":
                if st.session_state.get("recup_msg"):
                    st.info(st.session_state.get("recup_msg"))
                if st.session_state.get("recup_codigo_demo"):
                    st.warning(
                        f"Seu código de recuperação: **{st.session_state['recup_codigo_demo']}**"
                    )
                    st.caption("Anote o código. Sem e-mail configurado, ele aparece aqui por 15 minutos.")
                with st.form("form_recup_nova", clear_on_submit=False):
                    email_c = st.text_input(
                        "E-mail",
                        value=st.session_state.get("recup_email", ""),
                        placeholder="voce@email.com",
                    )
                    codigo_c = st.text_input("Código de 6 dígitos", placeholder="000000")
                    nova = st.text_input("Nova senha", type="password", placeholder="Mínimo 6 caracteres")
                    nova2 = st.text_input("Confirmar nova senha", type="password", placeholder="Repita a senha")
                    ok_n = st.form_submit_button("Redefinir senha", use_container_width=True)
                    if ok_n:
                        if nova != nova2:
                            st.error("As senhas não coincidem.")
                        else:
                            ok_x, msg_x = redefinir_senha_com_codigo(email_c, codigo_c, nova)
                            if ok_x:
                                for k in ("recup_etapa", "recup_email", "recup_msg", "recup_codigo_demo"):
                                    st.session_state.pop(k, None)
                                st.session_state.login_modo_ui = "Entrar"
                                st.success(msg_x)
                                st.rerun()
                            else:
                                st.error(msg_x)
                if st.button("Pedir novo código", key="btn_recup_voltar"):
                    for k in ("recup_etapa", "recup_email", "recup_msg", "recup_codigo_demo"):
                        st.session_state.pop(k, None)
                    st.rerun()

            if st.button("Voltar para Entrar", key="btn_voltar_entrar"):
                st.session_state.login_modo_ui = "Entrar"
                for k in ("recup_etapa", "recup_email", "recup_msg", "recup_codigo_demo"):
                    st.session_state.pop(k, None)
                st.rerun()

    st.caption("Cada conta tem seus próprios lançamentos, dívidas e metas. Login válido por 90 dias (renova ao usar).")


def carregar_dados(uid: int | None = None) -> pd.DataFrame:
    uid = uid if uid is not None else user_id_atual()
    conn = get_conn()
    try:
        if uid is None:
            df = pd.DataFrame(columns=["id", "user_id", "data", "descricao", "categoria", "valor", "tipo", "cartao"])
        else:
            df = pd.read_sql_query(
                "SELECT * FROM transacoes WHERE user_id = ? ORDER BY data DESC, id DESC",
                conn, params=(uid,),
            )
    except Exception:
        df = pd.DataFrame(columns=["id", "user_id", "data", "descricao", "categoria", "valor", "tipo", "cartao"])
    conn.close()
    if len(df) and "tipo" in df.columns:
        df["tipo"] = df["tipo"].replace({"Saida": "Saída", "saida": "Saída", "Entrada": "Entrada", "entrada": "Entrada"})
    return df


def carregar_investimentos(uid: int | None = None) -> pd.DataFrame:
    uid = uid if uid is not None else user_id_atual()
    conn = get_conn()
    try:
        if uid is None:
            df = pd.DataFrame(columns=["id", "user_id", "data", "tipo", "valor", "rentabilidade", "descricao", "status"])
        else:
            df = pd.read_sql_query(
                "SELECT * FROM investimentos WHERE user_id = ? ORDER BY data DESC, id DESC",
                conn, params=(uid,),
            )
    except Exception:
        df = pd.DataFrame(columns=["id", "user_id", "data", "tipo", "valor", "rentabilidade", "descricao", "status"])
    conn.close()
    return df


def carregar_dividas(uid: int | None = None) -> pd.DataFrame:
    uid = uid if uid is not None else user_id_atual()
    conn = get_conn()
    try:
        if uid is None:
            df = pd.DataFrame(columns=[
                "id", "user_id", "data", "credor", "tipo", "saldo_original", "desconto",
                "saldo_negociado", "parcela_possivel", "vencimento", "prioridade",
                "consequencia", "status", "proxima_acao", "anotacoes",
            ])
        else:
            df = pd.read_sql_query(
                "SELECT * FROM dividas WHERE user_id = ? ORDER BY data DESC, id DESC",
                conn, params=(uid,),
            )
    except Exception:
        df = pd.DataFrame(columns=[
            "id", "user_id", "data", "credor", "tipo", "saldo_original", "desconto",
            "saldo_negociado", "parcela_possivel", "vencimento", "prioridade",
            "consequencia", "status", "proxima_acao", "anotacoes",
        ])
    conn.close()
    return df


def carregar_metas(uid: int | None = None) -> pd.DataFrame:
    uid = uid if uid is not None else user_id_atual()
    conn = get_conn()
    try:
        if uid is None:
            df = pd.DataFrame(columns=[
                "id", "user_id", "data", "nome", "valor_meta", "valor_atual",
                "aporte_mensal", "prazo", "status", "anotacoes",
            ])
        else:
            df = pd.read_sql_query(
                "SELECT * FROM metas WHERE user_id = ? ORDER BY data DESC, id DESC",
                conn, params=(uid,),
            )
    except Exception:
        df = pd.DataFrame(columns=[
            "id", "user_id", "data", "nome", "valor_meta", "valor_atual",
            "aporte_mensal", "prazo", "status", "anotacoes",
        ])
    conn.close()
    return df


def salvar_transacao(data, descricao, categoria, valor, tipo, cartao, uid: int | None = None):
    uid = uid if uid is not None else user_id_atual()
    if uid is None:
        raise ValueError("Usuário não autenticado.")
    conn = get_conn()
    conn.execute(
        "INSERT INTO transacoes (user_id, data, descricao, categoria, valor, tipo, cartao) VALUES (?, ?, ?, ?, ?, ?, ?)",
        (uid, str(data), descricao, categoria, float(valor), tipo, cartao),
    )
    conn.commit()
    conn.close()


def excluir_transacao(id_, uid: int | None = None):
    uid = uid if uid is not None else user_id_atual()
    conn = get_conn()
    conn.execute("DELETE FROM transacoes WHERE id = ? AND user_id = ?", (int(id_), uid))
    conn.commit()
    conn.close()


def limpar_historico(uid: int | None = None):
    uid = uid if uid is not None else user_id_atual()
    conn = get_conn()
    conn.execute("DELETE FROM transacoes WHERE user_id = ?", (uid,))
    conn.commit()
    conn.close()


def salvar_investimento(data, tipo, valor, rentabilidade, descricao, status, uid: int | None = None):
    uid = uid if uid is not None else user_id_atual()
    if uid is None:
        raise ValueError("Usuário não autenticado.")
    conn = get_conn()
    conn.execute(
        "INSERT INTO investimentos (user_id, data, tipo, valor, rentabilidade, descricao, status) VALUES (?, ?, ?, ?, ?, ?, ?)",
        (uid, str(data), tipo, float(valor), rentabilidade, descricao, status),
    )
    conn.commit()
    conn.close()


def excluir_investimento(id_, uid: int | None = None):
    uid = uid if uid is not None else user_id_atual()
    conn = get_conn()
    conn.execute("DELETE FROM investimentos WHERE id = ? AND user_id = ?", (int(id_), uid))
    conn.commit()
    conn.close()


def salvar_divida(
    data, credor, tipo, saldo_original, desconto, saldo_negociado,
    parcela_possivel, vencimento, prioridade, consequencia, status,
    proxima_acao, anotacoes, uid: int | None = None,
):
    uid = uid if uid is not None else user_id_atual()
    if uid is None:
        raise ValueError("Usuário não autenticado.")
    conn = get_conn()
    conn.execute(
        """
        INSERT INTO dividas (
            user_id, data, credor, tipo, saldo_original, desconto, saldo_negociado,
            parcela_possivel, vencimento, prioridade, consequencia, status,
            proxima_acao, anotacoes
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            uid, str(data), credor, tipo, float(saldo_original), float(desconto),
            float(saldo_negociado), float(parcela_possivel), str(vencimento),
            prioridade, consequencia, status, proxima_acao, anotacoes,
        ),
    )
    conn.commit()
    conn.close()


def excluir_divida(id_, uid: int | None = None):
    uid = uid if uid is not None else user_id_atual()
    conn = get_conn()
    conn.execute("DELETE FROM dividas WHERE id = ? AND user_id = ?", (int(id_), uid))
    conn.commit()
    conn.close()


def salvar_meta(data, nome, valor_meta, valor_atual, aporte_mensal, prazo, status, anotacoes, uid: int | None = None):
    uid = uid if uid is not None else user_id_atual()
    if uid is None:
        raise ValueError("Usuário não autenticado.")
    conn = get_conn()
    conn.execute(
        """
        INSERT INTO metas (user_id, data, nome, valor_meta, valor_atual, aporte_mensal, prazo, status, anotacoes)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (uid, str(data), nome, float(valor_meta), float(valor_atual), float(aporte_mensal), prazo, status, anotacoes),
    )
    conn.commit()
    conn.close()


def excluir_meta(id_, uid: int | None = None):
    uid = uid if uid is not None else user_id_atual()
    conn = get_conn()
    conn.execute("DELETE FROM metas WHERE id = ? AND user_id = ?", (int(id_), uid))
    conn.commit()
    conn.close()


# ====================== FUNÇÕES DE DADOS ======================
def converter_valor(valor):
    if pd.isna(valor):
        return None
    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip()
    if not texto:
        return None

    negativo = (texto.startswith("(") and texto.endswith(")")) or texto.startswith("-")
    texto = (
        texto.replace("R$", "")
        .replace(" ", "")
        .replace("\u00a0", "")
        .replace("(", "")
        .replace(")", "")
    )
    texto = "".join(c for c in texto if c.isdigit() or c in ",.-")

    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    elif "," in texto:
        texto = texto.replace(",", ".")

    try:
        numero = float(texto)
    except ValueError:
        return None
    return -abs(numero) if negativo else numero


def calcular_score_financeiro(entradas, saidas, saldo, investimentos, total_dividas, parcelas_dividas, progresso_metas):
    if entradas <= 0 and saidas <= 0:
        return 50

    taxa_sobra = (saldo / entradas) * 100 if entradas > 0 else -35
    comprometimento = ((saidas + parcelas_dividas) / entradas) * 100 if entradas > 0 else 100
    meses_reserva = investimentos / (saidas / 3) if saidas > 0 else (3 if investimentos > 0 else 0)

    score = 52
    score += max(-24, min(taxa_sobra, 24))
    score -= max(0, min(comprometimento - 70, 22))
    score += min(meses_reserva, 3) * 4
    score += min(progresso_metas / 10, 10)
    if total_dividas > 0:
        score -= min((total_dividas / max(entradas, 1)) * 8, 16)
    if saldo >= 0:
        score += 5
    return int(max(0, min(round(score), 100)))


def preparar_fluxo_mensal(df_transacoes: pd.DataFrame) -> pd.DataFrame:
    if df_transacoes.empty:
        return pd.DataFrame(columns=["mes", "entradas", "saidas", "saldo"])

    df_fluxo = df_transacoes.copy()
    df_fluxo["data_convertida"] = pd.to_datetime(df_fluxo["data"], errors="coerce")
    df_fluxo = df_fluxo.dropna(subset=["data_convertida"])
    if df_fluxo.empty:
        return pd.DataFrame(columns=["mes", "entradas", "saidas", "saldo"])

    df_fluxo["mes"] = df_fluxo["data_convertida"].dt.to_period("M").dt.to_timestamp()
    df_fluxo["entradas"] = df_fluxo["valor"].clip(lower=0)
    df_fluxo["saidas"] = df_fluxo["valor"].clip(upper=0).abs()
    return (
        df_fluxo.groupby("mes", as_index=False)[["entradas", "saidas"]]
        .sum()
        .assign(saldo=lambda d: d["entradas"] - d["saidas"])
        .sort_values("mes")
    )


def resumo_mes_recente(df_transacoes: pd.DataFrame):
    df_fluxo = df_transacoes.copy()
    df_fluxo["data_convertida"] = pd.to_datetime(df_fluxo["data"], errors="coerce")
    df_fluxo = df_fluxo.dropna(subset=["data_convertida"])
    if df_fluxo.empty:
        return pd.DataFrame(), "Sem mês"

    mes_recente = df_fluxo["data_convertida"].max().to_period("M")
    df_mes = df_fluxo[df_fluxo["data_convertida"].dt.to_period("M") == mes_recente]
    return df_mes, mes_recente.strftime("%m/%Y")


def style_plot(fig, height=320, show_legend=True):
    """Estilo compacto, legivel no mobile, sem toolbar."""
    fig.update_layout(
        paper_bgcolor="rgba(255,255,255,0)",
        plot_bgcolor="rgba(255,255,255,0)",
        font=dict(color="#1c1f26", size=11, family="Inter, sans-serif"),
        title=dict(
            font=dict(size=14, color="#1c1f26", family="Inter, sans-serif"),
            x=0.0,
            xanchor="left",
            pad=dict(t=0, b=8),
        ),
        margin=dict(l=28, r=16, t=44, b=64),
        height=height,
        autosize=True,
        legend=dict(
            orientation="h",
            yanchor="top",
            y=-0.18,
            xanchor="center",
            x=0.5,
            bgcolor="rgba(255,255,255,0)",
            font=dict(size=10, color="#1c1f26"),
            itemwidth=30,
            tracegroupgap=4,
        ) if show_legend else dict(visible=False),
        hoverlabel=dict(bgcolor="#1c1f26", font_color="#ffffff", font_size=12),
        separators=",.",
        uniformtext_minsize=9,
        uniformtext_mode="hide",
        showlegend=show_legend,
    )
    fig.update_xaxes(
        gridcolor="rgba(28,31,38,0.06)",
        tickfont=dict(color="#8a90a0", size=10),
        zeroline=False,
        automargin=True,
    )
    fig.update_yaxes(
        gridcolor="rgba(28,31,38,0.06)",
        tickfont=dict(color="#8a90a0", size=10),
        zeroline=False,
        automargin=True,
    )
    # remove modebar chrome
    fig.update_layout(modebar=dict(remove=["zoom", "pan", "select", "lasso", "zoomIn", "zoomOut", "autoScale", "resetScale"]))
    return fig


PLOTLY_CONFIG = {
    "displayModeBar": False,
    "displaylogo": False,
    "responsive": True,
    "staticPlot": False,
}


# ====================== IMPORTAÇÃO DE PLANILHA ======================
MESES_PLANILHA = {
    "janeiro": 1, "fevereiro": 2, "marco": 3, "abril": 4,
    "maio": 5, "junho": 6, "julho": 7, "agosto": 8,
    "setembro": 9, "outubro": 10, "novembro": 11, "dezembro": 12,
}

PALAVRAS_ENTRADA = {"entrada", "receita", "recebimento", "salario", "renda", "freelance", "reembolso", "venda"}
PALAVRAS_SAIDA = {"saida", "despesa", "debito", "credito", "cartao", "gasto", "fixo", "parcelado"}


def excel_serial_para_data(valor, ano_padrao: int | None = None, mes_padrao: int | None = None) -> date:
    """Converte data do Excel (serial, datetime ou texto) para date."""
    if valor is None or (isinstance(valor, float) and math.isnan(valor)):
        if ano_padrao and mes_padrao:
            return date(ano_padrao, mes_padrao, 1)
        return date.today()

    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor

    # Serial do Excel (número de dias desde 1899-12-30)
    if isinstance(valor, (int, float)) and not isinstance(valor, bool):
        try:
            n = float(valor)
            if 20000 < n < 80000:  # faixa típica de datas Excel modernas
                return (datetime(1899, 12, 30) + timedelta(days=int(n))).date()
        except Exception:
            pass

    texto = str(valor).strip()
    if not texto:
        if ano_padrao and mes_padrao:
            return date(ano_padrao, mes_padrao, 1)
        return date.today()

    dt = pd.to_datetime(texto, errors="coerce", dayfirst=True)
    if not pd.isna(dt):
        return dt.date()

    if ano_padrao and mes_padrao:
        return date(ano_padrao, mes_padrao, 1)
    return date.today()


def _cel(df: pd.DataFrame, r: int, c: int):
    if r < 0 or c < 0 or r >= len(df.index) or c >= len(df.columns):
        return None
    try:
        return df.iat[r, c]
    except Exception:
        return None


def _linha_norm(df: pd.DataFrame, r: int) -> list[str]:
    return [normalizar(_cel(df, r, c)) for c in range(len(df.columns))]


def _achar_cabecalho(
    df: pd.DataFrame,
    obrigatorios: set[str],
    linha_inicio: int = 0,
    linha_fim: int | None = None,
    col_min: int = 0,
    col_max: int | None = None,
):
    """Retorna (linha, mapa col_nome_normalizado -> índice) limitado a um intervalo de colunas."""
    fim = linha_fim if linha_fim is not None else len(df.index)
    cmax = col_max if col_max is not None else len(df.columns)
    for r in range(linha_inicio, min(fim, len(df.index))):
        vals = _linha_norm(df, r)
        mapa = {}
        for c, v in enumerate(vals):
            if c < col_min or c >= cmax or not v:
                continue
            if v not in mapa:
                mapa[v] = c
        if obrigatorios.issubset(set(mapa.keys())):
            return r, mapa
    return None, {}


def _ler_bloco_tabela(
    df: pd.DataFrame,
    linha_header: int,
    mapa: dict,
    col_nome: str,
    col_valor: str,
    ano: int,
    mes: int,
    tipo_forcado: str | None = None,
    forma_pagamento: str = "Débito",
    parar_em: set[str] | None = None,
) -> list[dict]:
    """Lê linhas de dados abaixo de um cabeçalho até linha vazia ou marcador de fim."""
    parar_em = parar_em or {
        "total", "totais", "saidas", "entradas", "investimentos", "reserva",
        "cartao de credito", "fixos", "gastos do mes", "resumo",
    }
    idx_nome = mapa.get(col_nome, mapa.get("nome", mapa.get("descricao")))
    idx_valor = mapa.get(col_valor, mapa.get("valor"))
    idx_data = mapa.get("data")
    idx_tipo = mapa.get("tipo")
    idx_cat = mapa.get("categoria")
    idx_obs = mapa.get("observacao")
    idx_status = mapa.get("status")
    idx_pago = mapa.get("pago?")

    if idx_nome is None or idx_valor is None:
        return []

    linhas = []
    vazias = 0
    for r in range(linha_header + 1, len(df.index)):
        nome = limpar_texto(_cel(df, r, idx_nome))
        valor = converter_valor(_cel(df, r, idx_valor))
        nome_n = normalizar(nome)

        if not nome and (valor is None or valor == 0):
            vazias += 1
            if vazias >= 3:
                break
            continue
        vazias = 0

        if nome_n.startswith("total") or nome_n in parar_em:
            break
        # Evita capturar cabeçalhos repetidos
        if nome_n in ("nome", "descricao", "objetivo", "categoria"):
            continue
        if valor is None or valor == 0:
            continue

        tipo = tipo_forcado
        if not tipo:
            tipo_txt = normalizar(_cel(df, r, idx_tipo)) if idx_tipo is not None else ""
            if any(p in tipo_txt for p in PALAVRAS_ENTRADA) or tipo_txt in ("recebido",):
                tipo = "Entrada"
            elif any(p in tipo_txt for p in PALAVRAS_SAIDA) or "credito" in tipo_txt:
                tipo = "Saída"
            elif any(p in nome_n for p in PALAVRAS_ENTRADA):
                tipo = "Entrada"
            else:
                tipo = "Saída"

        categoria = limpar_texto(_cel(df, r, idx_cat), "Importado") if idx_cat is not None else "Importado"
        if tipo == "Entrada" and categoria in ("Importado", "Outros", ""):
            categoria = limpar_texto(nome, "Receita")

        data_mov = excel_serial_para_data(
            _cel(df, r, idx_data) if idx_data is not None else None,
            ano_padrao=ano,
            mes_padrao=mes,
        )

        # Forma de pagamento a partir do Tipo da planilha (Débito / Crédito 1...)
        cartao = forma_pagamento
        if idx_tipo is not None:
            tipo_plan = limpar_texto(_cel(df, r, idx_tipo))
            if tipo_plan:
                cartao = tipo_plan

        obs = limpar_texto(_cel(df, r, idx_obs)) if idx_obs is not None else ""
        status = limpar_texto(_cel(df, r, idx_status)) if idx_status is not None else ""
        pago = limpar_texto(_cel(df, r, idx_pago)) if idx_pago is not None else ""

        descricao = nome
        extras = []
        if obs and normalizar(obs) not in ("recebido", "sim", "nao", "pago"):
            extras.append(obs)
        if pago and normalizar(pago) in ("sim", "nao", "pendente"):
            extras.append(f"Pago: {pago}")
        if status and normalizar(status) not in ("recebido",) and tipo == "Entrada":
            extras.append(status)
        if extras:
            descricao = f"{nome} — " + " | ".join(extras)

        linhas.append({
            "data": data_mov.isoformat(),
            "descricao": descricao,
            "categoria": categoria if categoria else ("Receita" if tipo == "Entrada" else "Outros"),
            "valor": abs(float(valor)) if tipo == "Entrada" else -abs(float(valor)),
            "tipo": tipo,
            "cartao": cartao or "Planilha",
            "origem": "mes",
        })
    return linhas


def parse_aba_mensal(nome_aba: str, df_aba: pd.DataFrame, ano: int) -> list[dict]:
    """Interpreta abas Janeiro–Dezembro da Planilha Financeira Inteligente."""
    mes = MESES_PLANILHA.get(normalizar(nome_aba))
    if not mes or df_aba is None or df_aba.empty:
        return []

    linhas: list[dict] = []

    # --- FIXOS: Nome + Valor (+ Data, Tipo, Categoria) nas colunas da esquerda ---
    r, mapa = _achar_cabecalho(df_aba, {"nome", "valor"}, 0, 15, col_min=0, col_max=8)
    if r is not None:
        linhas.extend(_ler_bloco_tabela(
            df_aba, r, mapa, "nome", "valor", ano, mes,
            tipo_forcado="Saída", forma_pagamento="Débito",
        ))

    # --- GASTOS DO MÊS: segundo bloco com Nome/Data/Tipo/Categoria/Valor ---
    # Busca cabeçalho cujo "nome" está em colunas intermediárias (por volta de 9)
    for r in range(0, min(15, len(df_aba.index))):
        vals = _linha_norm(df_aba, r)
        cols_nome = [c for c, v in enumerate(vals) if v == "nome"]
        cols_valor = [c for c, v in enumerate(vals) if v == "valor"]
        cols_cat = [c for c, v in enumerate(vals) if v == "categoria"]
        for cn in cols_nome:
            if cn < 5:
                continue  # já tratado como fixos
            cv = next((c for c in cols_valor if c > cn), None)
            cc = next((c for c in cols_cat if c > cn), None)
            if cv is None:
                continue
            mapa_g = {"nome": cn, "valor": cv}
            if cc is not None:
                mapa_g["categoria"] = cc
            for label, key in (("data", "data"), ("tipo", "tipo"), ("observacao", "observacao")):
                for c, v in enumerate(vals):
                    if v == label and cn < c < cv + 2:
                        mapa_g[key] = c
            linhas.extend(_ler_bloco_tabela(
                df_aba, r, mapa_g, "nome", "valor", ano, mes,
                tipo_forcado="Saída", forma_pagamento="Débito",
            ))
            break

    # --- ENTRADAS: Descrição + Valor (+ Status) ---
    for r in range(0, min(20, len(df_aba.index))):
        vals = _linha_norm(df_aba, r)
        if "descricao" in vals and "valor" in vals:
            cd = vals.index("descricao")
            # Preferir bloco da direita (entradas)
            if cd < 10:
                continue
            cv = next((c for c, v in enumerate(vals) if v == "valor" and c > cd), None)
            if cv is None:
                continue
            mapa_e = {"descricao": cd, "valor": cv, "nome": cd}
            if "status" in vals:
                mapa_e["status"] = vals.index("status")
            linhas.extend(_ler_bloco_tabela(
                df_aba, r, mapa_e, "descricao", "valor", ano, mes,
                tipo_forcado="Entrada", forma_pagamento="Conta",
            ))
            break

    # --- CARTÃO DE CRÉDITO: Nome, Parcelas, Data, Tipo, Categoria, Valor ---
    for r in range(20, min(40, len(df_aba.index))):
        vals = _linha_norm(df_aba, r)
        if "nome" in vals and "valor" in vals and ("parcelas" in vals or "categoria" in vals):
            cn = vals.index("nome")
            if cn > 5:
                continue
            cv = next((c for c, v in enumerate(vals) if v == "valor" and c > cn), None)
            if cv is None:
                continue
            mapa_c = {"nome": cn, "valor": cv}
            for label in ("data", "tipo", "categoria", "observacao", "parcelas"):
                if label in vals:
                    mapa_c[label] = vals.index(label)
            bloc = _ler_bloco_tabela(
                df_aba, r, mapa_c, "nome", "valor", ano, mes,
                tipo_forcado="Saída", forma_pagamento="Cartão",
            )
            for item in bloc:
                item["cartao"] = limpar_texto(item.get("cartao"), "Cartão")
                if "Crédito" not in item["cartao"] and "Cartão" not in item["cartao"]:
                    item["cartao"] = f"Cartão — {item['cartao']}"
            linhas.extend(bloc)
            break

    # --- INVESTIMENTOS / RESERVA: Objetivo + Valor ---
    for r in range(20, min(40, len(df_aba.index))):
        vals = _linha_norm(df_aba, r)
        if "objetivo" in vals and "valor" in vals:
            co = vals.index("objetivo")
            cv = next((c for c, v in enumerate(vals) if v == "valor" and c > co), None)
            if cv is None:
                continue
            mapa_i = {"nome": co, "descricao": co, "valor": cv}
            if "data" in vals:
                mapa_i["data"] = vals.index("data")
            inv = _ler_bloco_tabela(
                df_aba, r, mapa_i, "nome", "valor", ano, mes,
                tipo_forcado="Entrada", forma_pagamento="Investimento",
                parar_em={"total", "debito", "credito 1", "resumo"},
            )
            for item in inv:
                item["categoria"] = "Investimento/Reserva"
                item["origem"] = "investimento"
                # Mantém como saída de caixa para reserva (dinheiro separado)
                # mas registra valor positivo em categoria especial — trata como saída de gasto livre
                item["tipo"] = "Saída"
                item["valor"] = -abs(float(item["valor"]))
                item["cartao"] = "Reserva/Investimento"
            linhas.extend(inv)
            break

    return linhas


def parse_aba_dividas(df_aba: pd.DataFrame) -> list[dict]:
    """Lê a aba Dividas da planilha inteligente."""
    if df_aba is None or df_aba.empty:
        return []
    r, mapa = _achar_cabecalho(df_aba, {"credor", "saldo original"}, 0, 20)
    # fallback: procura "credor" e "saldo negociado"
    if r is None:
        r, mapa = _achar_cabecalho(df_aba, {"credor"}, 0, 20)
    if r is None:
        return []

    def col(*nomes):
        for n in nomes:
            if n in mapa:
                return mapa[n]
        return None

    idx = {
        "data": col("data"),
        "credor": col("credor"),
        "tipo": col("tipo"),
        "saldo_original": col("saldo original", "saldo"),
        "desconto": col("desconto / abatimento", "desconto", "desconto abatimento"),
        "saldo_negociado": col("saldo negociado"),
        "parcela": col("parcela possivel", "parcela possível", "parcela"),
        "vencimento": col("vencimento"),
        "prioridade": col("prioridade"),
        "consequencia": col("consequencia se atrasar", "consequencia"),
        "status": col("status"),
        "proxima": col("proxima acao", "próxima ação", "proxima acao"),
    }
    if idx["credor"] is None:
        return []

    out = []
    for row in range(r + 1, len(df_aba.index)):
        credor = limpar_texto(_cel(df_aba, row, idx["credor"]))
        if not credor or normalizar(credor) in ("credor", "total"):
            if not credor:
                continue
            break
        saldo_neg = converter_valor(_cel(df_aba, row, idx["saldo_negociado"])) if idx["saldo_negociado"] is not None else None
        saldo_orig = converter_valor(_cel(df_aba, row, idx["saldo_original"])) if idx["saldo_original"] is not None else None
        if (saldo_neg is None or saldo_neg == 0) and (saldo_orig is None or saldo_orig == 0):
            continue
        out.append({
            "data": excel_serial_para_data(_cel(df_aba, row, idx["data"]) if idx["data"] is not None else None).isoformat(),
            "credor": credor,
            "tipo": limpar_texto(_cel(df_aba, row, idx["tipo"]), "Outro") if idx["tipo"] is not None else "Outro",
            "saldo_original": float(saldo_orig or saldo_neg or 0),
            "desconto": float(converter_valor(_cel(df_aba, row, idx["desconto"])) or 0) if idx["desconto"] is not None else 0.0,
            "saldo_negociado": float(saldo_neg or saldo_orig or 0),
            "parcela_possivel": float(converter_valor(_cel(df_aba, row, idx["parcela"])) or 0) if idx["parcela"] is not None else 0.0,
            "vencimento": excel_serial_para_data(_cel(df_aba, row, idx["vencimento"]) if idx["vencimento"] is not None else None).isoformat(),
            "prioridade": limpar_texto(_cel(df_aba, row, idx["prioridade"]), "Média") if idx["prioridade"] is not None else "Média",
            "consequencia": limpar_texto(_cel(df_aba, row, idx["consequencia"])) if idx["consequencia"] is not None else "",
            "status": limpar_texto(_cel(df_aba, row, idx["status"]), "Mapear") if idx["status"] is not None else "Mapear",
            "proxima_acao": limpar_texto(_cel(df_aba, row, idx["proxima"])) if idx["proxima"] is not None else "",
            "anotacoes": "Importado da planilha",
        })
    return out


def parse_aba_metas(df_aba: pd.DataFrame) -> list[dict]:
    """Lê a aba Metas da planilha inteligente."""
    if df_aba is None or df_aba.empty:
        return []
    r, mapa = _achar_cabecalho(df_aba, {"meta", "valor alvo"}, 0, 15)
    if r is None:
        r, mapa = _achar_cabecalho(df_aba, {"meta"}, 0, 15)
    if r is None:
        return []

    def col(*nomes):
        for n in nomes:
            if n in mapa:
                return mapa[n]
        return None

    idx_meta = col("meta")
    idx_alvo = col("valor alvo", "valor meta", "alvo")
    idx_atual = col("valor atual", "atual")
    idx_prazo = col("prazo")
    idx_prio = col("prioridade")
    idx_acao = col("proxima acao", "próxima ação", "proxima acao")
    if idx_meta is None:
        return []

    out = []
    for row in range(r + 1, len(df_aba.index)):
        nome = limpar_texto(_cel(df_aba, row, idx_meta))
        if not nome or normalizar(nome) in ("meta", "indicador"):
            if not nome:
                continue
            if normalizar(nome) in ("meta", "indicador"):
                break
            continue
        alvo = converter_valor(_cel(df_aba, row, idx_alvo)) if idx_alvo is not None else None
        if alvo is None or alvo == 0:
            continue
        atual = converter_valor(_cel(df_aba, row, idx_atual)) if idx_atual is not None else 0
        prazo_raw = _cel(df_aba, row, idx_prazo) if idx_prazo is not None else None
        try:
            prazo = excel_serial_para_data(prazo_raw).isoformat() if prazo_raw is not None else ""
        except Exception:
            prazo = limpar_texto(prazo_raw)
        prio = limpar_texto(_cel(df_aba, row, idx_prio), "Média") if idx_prio is not None else "Média"
        acao = limpar_texto(_cel(df_aba, row, idx_acao)) if idx_acao is not None else ""
        progresso = (float(atual or 0) / float(alvo)) if alvo else 0
        status = "Concluída" if progresso >= 1 else ("Em andamento" if progresso > 0 else "Planejada")
        out.append({
            "data": date.today().isoformat(),
            "nome": nome,
            "valor_meta": float(alvo),
            "valor_atual": float(atual or 0),
            "aporte_mensal": 0.0,
            "prazo": prazo,
            "status": status,
            "anotacoes": f"{prio}. {acao}".strip(". "),
        })
    return out


def gerar_modelo_excel() -> bytes:
    arquivo = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimentações"

    cabecalhos = ["Data", "Descrição", "Categoria", "Valor", "Tipo", "Forma de Pagamento"]
    exemplos = [
        [date.today(), "Salário mensal", "Salário", 3500.00, "Entrada", "Conta corrente"],
        [date.today(), "Compras do mês", "Mercado", 420.00, "Saída", "Cartão"],
        [date.today(), "Internet", "Contas", 99.90, "Saída", "Pix"],
    ]
    ws.append(cabecalhos)
    for ex in exemplos:
        ws.append(ex)

    tabela = Table(displayName="TabelaMovimentacoes", ref=f"A1:F{len(exemplos)+1}")
    tabela.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tabela)
    ws.freeze_panes = "A2"

    for col, largura in {"A": 14, "B": 28, "C": 18, "D": 16, "E": 12, "F": 22}.items():
        ws.column_dimensions[col].width = largura

    for cel in ws["A"][1:]:
        cel.number_format = "dd/mm/yyyy"
    for cel in ws["D"][1:]:
        cel.number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'

    for cel in ws[1]:
        cel.font = Font(bold=True, color="FFFFFF")
        cel.fill = PatternFill("solid", fgColor="111318")
        cel.alignment = Alignment(horizontal="center")

    wb.save(arquivo)
    arquivo.seek(0)
    return arquivo.getvalue()


def ler_planilha(arquivo):
    nome = getattr(arquivo, "name", "").lower()
    if nome.endswith(".csv"):
        try:
            return pd.read_csv(arquivo, sep=None, engine="python")
        except UnicodeDecodeError:
            arquivo.seek(0)
            return pd.read_csv(arquivo, sep=None, engine="python", encoding="latin-1")

    dados = arquivo.read()
    if nome.endswith(".xlsx"):
        return pd.read_excel(BytesIO(dados), sheet_name=None, header=None, engine="openpyxl")
    if nome.endswith(".xls"):
        return pd.read_excel(BytesIO(dados), sheet_name=None, header=None)
    raise ValueError("Envie uma planilha CSV, XLSX ou XLS.")


def preparar_importacao(dados, ano: int | None = None):
    """
    Suporta:
    1) Planilha Financeira Inteligente (abas Janeiro–Dezembro + Dividas + Metas)
    2) Planilha simples tabular (Data/Descrição/Valor/Tipo)
    Retorna dict com dataframes e contagens.
    """
    ano = ano or date.today().year
    resultado = {
        "movimentacoes": pd.DataFrame(),
        "dividas": [],
        "metas": [],
        "ignoradas": 0,
        "abas_lidas": [],
        "modelo": "desconhecido",
    }

    if isinstance(dados, dict):
        nomes_norm = {normalizar(k): k for k in dados.keys()}
        tem_meses = any(m in nomes_norm for m in MESES_PLANILHA)
        tem_dividas = any(n in nomes_norm for n in ("dividas", "dívidas", "divida"))
        tem_metas = "metas" in nomes_norm

        if tem_meses or tem_dividas or tem_metas:
            resultado["modelo"] = "planilha_inteligente"
            movs = []
            for mes_nome, num in MESES_PLANILHA.items():
                if mes_nome not in nomes_norm:
                    continue
                aba = dados[nomes_norm[mes_nome]]
                parsed = parse_aba_mensal(mes_nome, aba, ano)
                if parsed:
                    resultado["abas_lidas"].append(mes_nome.capitalize())
                    movs.extend(parsed)

            if tem_dividas:
                chave = next(nomes_norm[n] for n in nomes_norm if n in ("dividas", "dívidas", "divida"))
                resultado["dividas"] = parse_aba_dividas(dados[chave])
                resultado["abas_lidas"].append("Dividas")

            if tem_metas:
                resultado["metas"] = parse_aba_metas(dados[nomes_norm["metas"]])
                resultado["abas_lidas"].append("Metas")

            if movs:
                df_m = pd.DataFrame(movs)
                # remove coluna auxiliar
                if "origem" in df_m.columns:
                    df_m = df_m.drop(columns=["origem"])
                resultado["movimentacoes"] = df_m.drop_duplicates(
                    subset=["data", "descricao", "valor", "tipo"], keep="first"
                ).reset_index(drop=True)
            return resultado

        # Fallback: primeira aba com header tabular
        primeira = next(iter(dados.values()), pd.DataFrame())
        if primeira.empty:
            return resultado
        primeira = primeira.copy()
        primeira.columns = primeira.iloc[0]
        dados = primeira.iloc[1:].reset_index(drop=True)

    # Planilha tabular normal
    resultado["modelo"] = "tabular"
    aliases = {
        "data": ["data", "dt", "dia", "date"],
        "descricao": ["descricao", "descrição", "historico", "histórico", "nome", "lancamento"],
        "categoria": ["categoria", "grupo", "classificacao"],
        "valor": ["valor", "valor r$", "valor rs", "amount", "preco"],
        "tipo": ["tipo", "natureza", "entrada saida"],
        "cartao": ["forma de pagamento", "pagamento", "cartao", "cartão", "conta", "banco"],
    }
    colunas = {normalizar(c): c for c in dados.columns}
    mapa = {}
    for destino, opcoes in aliases.items():
        for opcao in opcoes:
            if normalizar(opcao) in colunas:
                mapa[destino] = colunas[normalizar(opcao)]
                break

    if "valor" not in mapa:
        raise ValueError("A planilha precisa ter uma coluna de valor ou abas mensais (Janeiro–Dezembro).")

    linhas = []
    ignoradas = 0
    for _, row in dados.iterrows():
        valor_original = converter_valor(row.get(mapa["valor"]))
        if valor_original is None or valor_original == 0:
            ignoradas += 1
            continue

        data_conv = pd.to_datetime(row.get(mapa.get("data"), date.today()), errors="coerce", dayfirst=True)
        if pd.isna(data_conv):
            data_conv = pd.Timestamp(date.today())

        tipo_texto = normalizar(row.get(mapa.get("tipo"), ""))
        if any(p in tipo_texto for p in PALAVRAS_ENTRADA):
            tipo = "Entrada"
        elif any(p in tipo_texto for p in PALAVRAS_SAIDA):
            tipo = "Saída"
        else:
            tipo = "Entrada" if valor_original > 0 else "Saída"

        linhas.append({
            "data": data_conv.date().isoformat(),
            "descricao": limpar_texto(row.get(mapa.get("descricao")), "Importado da planilha"),
            "categoria": limpar_texto(row.get(mapa.get("categoria")), "Importado"),
            "valor": abs(valor_original) if tipo == "Entrada" else -abs(valor_original),
            "tipo": tipo,
            "cartao": limpar_texto(row.get(mapa.get("cartao")), "Planilha"),
        })

    resultado["movimentacoes"] = pd.DataFrame(linhas)
    resultado["ignoradas"] = ignoradas
    return resultado


def chave_movimentacao(registro) -> tuple:
    data_mov = pd.to_datetime(registro.get("data"), errors="coerce", dayfirst=True)
    data_norm = data_mov.date().isoformat() if not pd.isna(data_mov) else limpar_texto(registro.get("data"))
    descricao = normalizar(limpar_texto(registro.get("descricao")))
    tipo = normalizar(limpar_texto(registro.get("tipo")))
    valor = round(float(converter_valor(registro.get("valor")) or 0), 2)
    return data_norm, descricao, valor, tipo


def conciliar_movimentacoes(df_importado: pd.DataFrame, df_existente: pd.DataFrame):
    if df_importado.empty:
        return df_importado.copy(), 0

    chaves_existentes = {chave_movimentacao(r) for r in df_existente.to_dict(orient="records")}
    chaves_novas = set()
    indices = []
    duplicadas = 0

    for idx, reg in df_importado.iterrows():
        chave = chave_movimentacao(reg)
        if chave in chaves_existentes or chave in chaves_novas:
            duplicadas += 1
            continue
        chaves_novas.add(chave)
        indices.append(idx)

    return df_importado.loc[indices].reset_index(drop=True), duplicadas


def importar_movimentacoes(df_importado: pd.DataFrame):
    df_novo, duplicadas = conciliar_movimentacoes(df_importado, carregar_dados())
    if df_novo.empty:
        return 0, duplicadas

    conn = get_conn()
    uid = user_id_atual()
    if uid is None:
        return 0, duplicadas
    rows = [
        (uid, r.data, r.descricao, r.categoria, r.valor, r.tipo, r.cartao)
        for r in df_novo[["data", "descricao", "categoria", "valor", "tipo", "cartao"]].itertuples(index=False)
    ]
    conn.executemany(
        "INSERT INTO transacoes (user_id, data, descricao, categoria, valor, tipo, cartao) VALUES (?, ?, ?, ?, ?, ?, ?)",
        rows,
    )
    conn.commit()
    conn.close()
    return len(df_novo), duplicadas


def importar_dividas_lista(lista: list[dict]) -> int:
    if not lista:
        return 0
    existentes = carregar_dividas()
    chaves = set()
    if len(existentes):
        for _, r in existentes.iterrows():
            chaves.add((normalizar(r.get("credor")), round(float(r.get("saldo_negociado") or 0), 2)))
    gravados = 0
    for d in lista:
        chave = (normalizar(d.get("credor")), round(float(d.get("saldo_negociado") or 0), 2))
        if chave in chaves:
            continue
        salvar_divida(
            d.get("data"),
            d.get("credor"),
            d.get("tipo"),
            d.get("saldo_original") or 0,
            d.get("desconto") or 0,
            d.get("saldo_negociado") or 0,
            d.get("parcela_possivel") or 0,
            d.get("vencimento"),
            d.get("prioridade") or "Média",
            d.get("consequencia") or "",
            d.get("status") or "Mapear",
            d.get("proxima_acao") or "",
            d.get("anotacoes") or "Importado da planilha",
        )
        chaves.add(chave)
        gravados += 1
    return gravados


def importar_metas_lista(lista: list[dict]) -> int:
    if not lista:
        return 0
    existentes = carregar_metas()
    nomes = set(normalizar(x) for x in existentes["nome"].tolist()) if len(existentes) and "nome" in existentes.columns else set()
    gravados = 0
    for m in lista:
        if normalizar(m.get("nome")) in nomes:
            continue
        salvar_meta(
            m.get("data") or date.today().isoformat(),
            m.get("nome"),
            m.get("valor_meta") or 0,
            m.get("valor_atual") or 0,
            m.get("aporte_mensal") or 0,
            m.get("prazo") or "",
            m.get("status") or "Em andamento",
            m.get("anotacoes") or "",
        )
        nomes.add(normalizar(m.get("nome")))
        gravados += 1
    return gravados


# ====================== RELATÓRIO PDF ======================
def gerar_pdf(df, investimentos, dividas, metas) -> bytes:
    """Relatório visual no modelo aprovado (capa, fluxo, categorias, dívidas, metas, alertas)."""
    from io import BytesIO
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.lib.colors import Color, white
        from reportlab.pdfgen import canvas as pdf_canvas
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ModuleNotFoundError as e:
        raise ModuleNotFoundError(
            "Biblioteca reportlab não instalada. No Streamlit Cloud, inclua "
            "'reportlab>=4.0.0' no requirements.txt e faça o redeploy."
        ) from e

    # Fonte com suporte a português
    try:
        pdfmetrics.registerFont(TTFont("DejaVu", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"))
        pdfmetrics.registerFont(TTFont("DejaVuBold", "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"))
        FONT, FONT_B = "DejaVu", "DejaVuBold"
    except Exception:
        FONT, FONT_B = "Helvetica", "Helvetica-Bold"

    # ---- métricas ----
    entradas = float(df[df["valor"] > 0]["valor"].sum()) if len(df) else 0.0
    saidas = float(abs(df[df["valor"] < 0]["valor"].sum())) if len(df) else 0.0
    saldo = float(df["valor"].sum()) if len(df) else 0.0
    total_inv = float(investimentos["valor"].sum()) if len(investimentos) else 0.0
    taxa_sobra = (saldo / entradas * 100) if entradas > 0 else 0.0

    total_div = 0.0
    parcelas_div = 0.0
    prioritarias = 0
    if len(dividas):
        for _, d in dividas.iterrows():
            sb = d.get("saldo_negociado") or d.get("saldo_original") or 0
            try:
                sb = float(sb)
            except Exception:
                sb = 0.0
            stt = normalizar(d.get("status"))
            if stt not in ("quitada", "paga"):
                total_div += sb
                try:
                    parcelas_div += float(d.get("parcela_possivel") or 0)
                except Exception:
                    pass
            if normalizar(d.get("prioridade")) == "alta" and stt not in ("quitada", "paga"):
                prioritarias += 1

    progresso_metas = 0.0
    if len(metas):
        pcts = []
        for _, m in metas.iterrows():
            alvo = float(m.get("valor_meta") or 0)
            atual = float(m.get("valor_atual") or 0)
            pcts.append(min(100.0, (atual / alvo * 100) if alvo > 0 else 0))
        progresso_metas = sum(pcts) / len(pcts) if pcts else 0

    score = calcular_score_financeiro(
        entradas, saidas, saldo, total_inv, total_div, parcelas_div, progresso_metas
    )
    score_label = "Saudável" if score >= 70 else ("Atenção" if score >= 45 else "Crítico")

    # período
    periodo = "Sem movimentações"
    if len(df):
        datas = pd.to_datetime(df["data"], errors="coerce").dropna()
        if len(datas):
            periodo = f"{datas.min().strftime('%d/%m/%Y')} a {datas.max().strftime('%d/%m/%Y')}"

    # categorias de saída
    cats = []
    if len(df):
        saidas_df = df[df["valor"] < 0].copy()
        if len(saidas_df):
            saidas_df["categoria"] = saidas_df["categoria"].fillna("Outros").replace("", "Outros")
            g = (
                saidas_df.groupby("categoria", as_index=False)["valor"]
                .sum()
            )
            g["valor_abs"] = g["valor"].abs()
            g = g.sort_values("valor_abs", ascending=False).head(5)
            total_s = g["valor_abs"].sum() or 1
            for _, r in g.iterrows():
                cats.append((str(r["categoria"]), float(r["valor_abs"]), float(r["valor_abs"]) / total_s))

    # fluxo mensal
    fluxo = preparar_fluxo_mensal(df) if len(df) else pd.DataFrame()

    frases = [
        "Cada real de hoje constrói o amanhã que você escolhe.",
        "O futuro agradece quem organiza o presente.",
        "Pequenos passos financeiros hoje viram liberdade depois.",
        "Cuidar do agora é o jeito mais simples de cuidar do futuro.",
    ]
    frase = frases[date.today().toordinal() % len(frases)]

    if saldo >= 0:
        leitura = f"Você terminou no azul. Sobraram {brl(saldo)} no período."
    else:
        leitura = f"O período fechou no vermelho em {brl(abs(saldo))}. Vale revisar gastos flexíveis."

    # cores
    ink = Color(28 / 255, 31 / 255, 38 / 255)
    muted = Color(107 / 255, 114 / 255, 128 / 255)
    lime = Color(217 / 255, 255 / 255, 0 / 255)
    mint = Color(184 / 255, 240 / 255, 216 / 255)
    mint_soft = Color(232 / 255, 250 / 255, 243 / 255)
    bg = Color(243 / 255, 244 / 255, 247 / 255)
    line = Color(0.85, 0.86, 0.88)
    danger = Color(225 / 255, 29 / 255, 72 / 255)
    ok = Color(13 / 255, 159 / 255, 110 / 255)
    warn = Color(138 / 255, 90 / 255, 0)

    W, H = A4
    buf = BytesIO()
    c = pdf_canvas.Canvas(buf, pagesize=A4)
    margin = 16 * mm

    def rr(x, y, w, h, r=10, fill=None, stroke=None, sw=0.7):
        c.saveState()
        if fill:
            c.setFillColor(fill)
        if stroke:
            c.setStrokeColor(stroke)
            c.setLineWidth(sw)
        c.roundRect(x, y, w, h, r, fill=1 if fill else 0, stroke=1 if stroke else 0)
        c.restoreState()

    def T(x, y, s, size=10, color=ink, bold=False, right=False):
        c.setFillColor(color)
        c.setFont(FONT_B if bold else FONT, size)
        (c.drawRightString if right else c.drawString)(x, y, str(s))

    def page_bg():
        c.setFillColor(bg)
        c.rect(0, 0, W, H, fill=1, stroke=0)
        c.setFillColor(Color(217 / 255, 255 / 255, 0 / 255, alpha=0.14))
        c.circle(30, H - 30, 80, fill=1, stroke=0)
        c.setFillColor(Color(184 / 255, 240 / 255, 216 / 255, alpha=0.20))
        c.circle(W - 20, H - 90, 95, fill=1, stroke=0)
        c.setFillColor(Color(168 / 255, 196 / 255, 255 / 255, alpha=0.14))
        c.circle(W - 50, 60, 85, fill=1, stroke=0)

    def footer(page, total=3):
        T(margin, 10 * mm, f"Relatório financeiro pessoal  ·  {date.today().strftime('%d/%m/%Y')}  ·  Página {page}/{total}", 7, muted)

    # ========== PÁGINA 1 ==========
    page_bg()
    rr(margin, H - 20 * mm, 38 * mm, 6.5 * mm, 8, white, line)
    T(margin + 2.5 * mm, H - 17.8 * mm, "Financeiro pessoal", 8, muted, True)
    score_bg = mint_soft if score >= 70 else (Color(1, 0.94, 0.76) if score >= 45 else Color(1, 0.89, 0.91))
    score_fg = ok if score >= 70 else (warn if score >= 45 else danger)
    rr(margin + 40 * mm, H - 20 * mm, 36 * mm, 6.5 * mm, 8, score_bg)
    T(margin + 42.5 * mm, H - 17.8 * mm, f"{score_label} · {score}/100", 8, score_fg, True)

    T(margin, H - 32 * mm, "Relatório do meu dinheiro", 18, ink, True)
    T(margin, H - 38 * mm, f"Período: {periodo}", 9, muted)

    rr(W - margin - 58 * mm, H - 46 * mm, 58 * mm, 26 * mm, 12, white, line)
    T(W - margin - 53 * mm, H - 28 * mm, '"', 14, lime, True)
    # quebra frase
    palavras = frase.split()
    linha1, linha2 = [], []
    acc = ""
    for p in palavras:
        test = (acc + " " + p).strip()
        if len(test) < 28:
            acc = test
        else:
            linha1 = acc.split() if not linha1 else linha1
            if not linha1:
                linha1 = acc.split()
            acc = p
    if not linha1:
        mid = max(1, len(palavras) // 2)
        linha1, linha2 = palavras[:mid], palavras[mid:]
    else:
        linha2 = acc.split() if acc else []
        if not linha2 and len(palavras) > len(linha1):
            linha2 = palavras[len(linha1):]
    T(W - margin - 53 * mm, H - 35 * mm, " ".join(linha1) if isinstance(linha1, list) else linha1, 8, ink, True)
    if linha2:
        T(W - margin - 53 * mm, H - 39 * mm, " ".join(linha2) if isinstance(linha2, list) else linha2, 8, ink, True)

    kpis = [
        (lime, "Saldo do período", brl(saldo), "Resultado acumulado"),
        (mint_soft, "Entradas", brl(entradas), "Tudo que entrou"),
        (white, "Saídas", brl(saidas), "Tudo que saiu"),
        (white, "Taxa de sobra", pct(taxa_sobra), "Parte que sobrou"),
    ]
    cw = (W - 2 * margin - 9 * mm) / 4
    for i, (col, lab, val, foot) in enumerate(kpis):
        x = margin + i * (cw + 3 * mm)
        y = H - 74 * mm
        rr(x, y, cw, 24 * mm, 12, col, None if col != white else line)
        T(x + 3 * mm, y + 16 * mm, lab, 7.5, muted, True)
        T(x + 3 * mm, y + 9 * mm, val, 11, ink, True)
        T(x + 3 * mm, y + 3.5 * mm, foot, 7, muted)

    T(margin, H - 86 * mm, "1. Resumo em 30 segundos", 12, ink, True)
    rr(margin, H - 112 * mm, W - 2 * margin, 22 * mm, 12, white, line)
    T(margin + 5 * mm, H - 95 * mm, "Leitura do período", 8, muted, True)
    T(margin + 5 * mm, H - 102 * mm, leitura[:90], 10, ink, True)
    extra = f"Movimentações: {len(df)}  ·  Dívidas ativas: {sum(1 for _,d in dividas.iterrows() if normalizar(d.get('status')) not in ('quitada','paga')) if len(dividas) else 0}  ·  Metas: {len(metas)}"
    T(margin + 5 * mm, H - 108 * mm, extra, 8, muted)

    # Fluxo mensal
    T(margin, H - 122 * mm, "2. Fluxo de caixa (visão mensal)", 12, ink, True)
    rr(margin, H - 168 * mm, W - 2 * margin, 40 * mm, 12, white, line)
    T(margin + 5 * mm, H - 132 * mm, "Verde = entradas   ·   Preto = saídas", 8, muted)
    if len(fluxo):
        n = min(len(fluxo), 8)
        max_v = max(float(fluxo["entradas"].max() or 1), float(fluxo["saidas"].max() or 1), 1)
        base_y = H - 162 * mm
        gap = (W - 2 * margin - 20 * mm) / max(n, 1)
        for i in range(n):
            row = fluxo.iloc[i]
            x = margin + 12 * mm + i * gap
            he = 26 * mm * (float(row.get("entradas") or 0) / max_v)
            hs = 26 * mm * (float(row.get("saidas") or 0) / max_v)
            c.setFillColor(mint)
            c.rect(x, base_y, 6 * mm, max(he, 0.5), fill=1, stroke=0)
            c.setFillColor(ink)
            c.rect(x + 7 * mm, base_y, 6 * mm, max(hs, 0.5), fill=1, stroke=0)
            mes_val = row.get("mes")
            if hasattr(mes_val, "strftime"):
                rotulo = mes_val.strftime("%m/%y")
            else:
                rotulo = str(mes_val or f"M{i+1}")[:7]
            T(x, base_y - 5 * mm, rotulo, 6.5, muted)
    else:
        T(margin + 5 * mm, H - 150 * mm, "Sem dados mensais suficientes para o gráfico.", 9, muted)

    # Categorias
    T(margin, H - 180 * mm, "3. Para onde foi o dinheiro", 12, ink, True)
    rr(margin, H - 230 * mm, W - 2 * margin, 44 * mm, 12, white, line)
    if cats:
        cores_cat = [mint, lime, Color(0.66, 0.77, 1), Color(0.78, 0.80, 0.84), Color(0.72, 0.74, 0.78)]
        for i, (nome, val, share) in enumerate(cats):
            y = H - 190 * mm - i * 7.5 * mm
            T(margin + 5 * mm, y, nome[:18], 9, ink, True)
            T(margin + 48 * mm, y, brl(val), 8, muted)
            c.setFillColor(Color(0.9, 0.91, 0.93))
            c.roundRect(margin + 85 * mm, y - 1, 90 * mm, 3.5 * mm, 2, fill=1, stroke=0)
            c.setFillColor(cores_cat[i % len(cores_cat)])
            c.roundRect(margin + 85 * mm, y - 1, max(2, 90 * mm * share), 3.5 * mm, 2, fill=1, stroke=0)
            T(margin + 180 * mm, y, f"{share * 100:.0f}%".replace(".", ","), 8, muted)
    else:
        T(margin + 5 * mm, H - 200 * mm, "Sem saídas categorizadas no período.", 9, muted)

    footer(1)
    c.showPage()

    # ========== PÁGINA 2 ==========
    page_bg()
    T(margin, H - 16 * mm, "4. Dívidas em aberto", 12, ink, True)
    T(margin, H - 22 * mm, "Saldo, parcela e próxima ação — não só o número.", 8, muted)

    dk = [
        ("Total aberto", brl(total_div)),
        ("Parcelas/mês", brl(parcelas_div)),
        ("Prioridade alta", str(prioritarias)),
        ("Cadastradas", str(len(dividas))),
    ]
    dw = (W - 2 * margin - 9 * mm) / 4
    for i, (lab, val) in enumerate(dk):
        x = margin + i * (dw + 3 * mm)
        rr(x, H - 44 * mm, dw, 16 * mm, 10, white, line)
        T(x + 3 * mm, H - 34 * mm, lab, 7, muted, True)
        T(x + 3 * mm, H - 40 * mm, val, 10, ink, True)

    if len(dividas):
        y0 = H - 54 * mm
        shown = 0
        for _, d in dividas.iterrows():
            if shown >= 6:
                break
            stt = normalizar(d.get("status"))
            sb = d.get("saldo_negociado") or d.get("saldo_original") or 0
            try:
                sb = float(sb)
            except Exception:
                sb = 0.0
            parc = d.get("parcela_possivel") or 0
            try:
                parc = float(parc)
            except Exception:
                parc = 0.0
            y = y0 - shown * 20 * mm
            rr(margin, y - 12 * mm, W - 2 * margin, 18 * mm, 10, white, line)
            T(margin + 4 * mm, y, str(d.get("credor") or "Credor")[:32], 10, ink, True)
            meta = f"{d.get('prioridade') or '—'} · {d.get('status') or '—'}"
            T(margin + 4 * mm, y - 5 * mm, meta[:50], 8, muted)
            acao = str(d.get("proxima_acao") or "Sem próxima ação definida")[:55]
            T(margin + 4 * mm, y - 9.5 * mm, f"Próxima ação: {acao}", 8, ink)
            T(W - margin - 4 * mm, y, brl(sb), 10, ink, True, True)
            T(W - margin - 4 * mm, y - 6 * mm, f"Parcela {brl(parc)}", 8, muted, False, True)
            shown += 1
    else:
        rr(margin, H - 70 * mm, W - 2 * margin, 16 * mm, 10, mint_soft)
        T(margin + 5 * mm, H - 62 * mm, "Nenhuma dívida cadastrada. Bom sinal.", 10, ok, True)

    T(margin, H - 185 * mm if len(dividas) >= 5 else H - 150 * mm, "5. Metas e reserva", 12, ink, True)
    y_meta_title = H - 185 * mm if len(dividas) >= 5 else H - 150 * mm

    # reserva card
    rr(margin, y_meta_title - 18 * mm, W - 2 * margin, 14 * mm, 10, mint_soft)
    T(margin + 4 * mm, y_meta_title - 10 * mm, f"Patrimônio / investimentos registrados: {brl(total_inv)}", 9, ink, True)

    if len(metas):
        for i, (_, mrow) in enumerate(metas.iterrows()):
            if i >= 5:
                break
            y = y_meta_title - 28 * mm - i * 15 * mm
            alvo = float(mrow.get("valor_meta") or 0)
            atual = float(mrow.get("valor_atual") or 0)
            p = min(1.0, atual / alvo) if alvo > 0 else 0
            rr(margin, y - 7 * mm, W - 2 * margin, 13 * mm, 10, white, line)
            T(margin + 4 * mm, y + 1 * mm, str(mrow.get("nome") or "Meta")[:36], 9, ink, True)
            T(margin + 4 * mm, y - 4 * mm, f"{brl(atual)} de {brl(alvo)}", 8, muted)
            c.setFillColor(Color(0.9, 0.91, 0.93))
            c.roundRect(W - margin - 52 * mm, y - 1.5 * mm, 46 * mm, 3.5 * mm, 2, fill=1, stroke=0)
            c.setFillColor(lime if p > 0.8 else mint)
            c.roundRect(W - margin - 52 * mm, y - 1.5 * mm, max(1, 46 * mm * p), 3.5 * mm, 2, fill=1, stroke=0)
            T(W - margin - 4 * mm, y + 1 * mm, f"{int(p * 100)}%", 9, ink, True, True)
    else:
        T(margin + 4 * mm, y_meta_title - 28 * mm, "Nenhuma meta cadastrada ainda.", 9, muted)

    footer(2)
    c.showPage()

    # ========== PÁGINA 3 ==========
    page_bg()
    T(margin, H - 16 * mm, "6. Projeção dos próximos meses", 12, ink, True)
    T(margin, H - 22 * mm, "Cenário simples com base na média de sobra atual.", 8, muted)
    rr(margin, H - 70 * mm, W - 2 * margin, 42 * mm, 12, white, line)

    media_sobra = 0.0
    if len(fluxo) and "saldo" in fluxo.columns:
        media_sobra = float(fluxo["saldo"].mean())
    elif entradas or saidas:
        media_sobra = (entradas - saidas) / max(len(fluxo), 1) if len(fluxo) else (entradas - saidas) / 6

    saldo_proj = saldo
    pts = []
    labels_m = []
    from datetime import timedelta as _td
    base = date.today().replace(day=1)
    for i in range(6):
        # avança mês
        month = base.month + i
        year = base.year + (month - 1) // 12
        month = (month - 1) % 12 + 1
        labels_m.append(f"{month:02d}/{str(year)[2:]}")
        saldo_proj = saldo + media_sobra * (i + 1)
        pts.append(saldo_proj)

    if pts:
        min_p, max_p = min(pts), max(pts)
        span = max(max_p - min_p, 1)
        ox, oy = margin + 18 * mm, H - 62 * mm
        sx = (W - 2 * margin - 36 * mm) / 5
        sy = 28 * mm
        c.setStrokeColor(ink)
        c.setLineWidth(2)
        path = c.beginPath()
        for i, val in enumerate(pts):
            x = ox + i * sx
            y = oy + ((val - min_p) / span) * sy
            if i == 0:
                path.moveTo(x, y)
            else:
                path.lineTo(x, y)
        c.drawPath(path, stroke=1, fill=0)
        for i, val in enumerate(pts):
            x = ox + i * sx
            y = oy + ((val - min_p) / span) * sy
            c.setFillColor(lime)
            c.circle(x, y, 3.2, fill=1, stroke=0)
            T(x - 6, oy - 6 * mm, labels_m[i], 7, muted)
        T(margin + 5 * mm, H - 34 * mm, f"Saldo atual {brl(saldo)}  ·  Média mensal de sobra {brl(media_sobra)}", 9, ink)

    T(margin, H - 82 * mm, "7. Alertas e próximos passos", 12, ink, True)
    alerts = []
    if prioritarias > 0:
        alerts.append((Color(1, 0.89, 0.91), danger, "Prioridade", f"{prioritarias} dívida(s) de prioridade alta em aberto — manter parcelas em dia."))
    if taxa_sobra < 10 and entradas > 0:
        alerts.append((Color(1, 0.97, 0.86), warn, "Atenção", "Taxa de sobra baixa. Revise gastos flexíveis antes de novos compromissos."))
    if saldo > 0 and total_div > 0:
        alerts.append((mint_soft, ok, "Oportunidade", "Sobra positiva: dá para reforçar reserva ou antecipar dívida prioritária."))
    if cats:
        top = cats[0]
        if top[2] > 0.35:
            alerts.append((Color(0.95, 1, 0.78), ink, "Concentração", f"Categoria {top[0]} concentra {top[2]*100:.0f}% das saídas — vale olhar de perto."))
    if not alerts:
        alerts.append((mint_soft, ok, "Estável", "Sem alertas críticos. Continue registrando e revisando o mês."))
    if total_inv > 0 and saidas > 0:
        meses_colchao = total_inv / (saidas / max(len(fluxo), 1)) if saidas else 0
        if meses_colchao < 1:
            alerts.append((Color(1, 0.97, 0.86), warn, "Reserva", "Reserva ainda cobre pouco mais do que poucos dias de gasto. Meta: 1–3 meses."))

    for i, (bgc, accent, tag, msg) in enumerate(alerts[:5]):
        y = H - 96 * mm - i * 17 * mm
        rr(margin, y - 7 * mm, W - 2 * margin, 15 * mm, 10, bgc)
        rr(margin + 3 * mm, y - 1.5 * mm, 24 * mm, 6 * mm, 6, white)
        T(margin + 5 * mm, y, tag, 7, accent, True)
        # wrap msg roughly
        T(margin + 30 * mm, y, msg[:78], 8, ink)

    T(margin, H - 190 * mm, "8. Como usar este relatório", 12, ink, True)
    tips = [
        "Compare este mês com o anterior nas entradas e saídas.",
        "Nas dívidas, siga a próxima ação antes de aceitar novos acordos.",
        "Nas metas, celebre progresso e ajuste o aporte se a sobra mudar.",
        "Exporte de novo todo mês para montar seu histórico de decisões.",
    ]
    for i, t in enumerate(tips):
        y = H - 202 * mm - i * 7 * mm
        c.setFillColor(lime)
        c.circle(margin + 2.2 * mm, y + 1.5, 2.8, fill=1, stroke=0)
        T(margin + 7 * mm, y, t, 9, ink)

    rr(margin, 18 * mm, W - 2 * margin, 20 * mm, 12, white, line)
    T(margin + 5 * mm, 30 * mm, "Dashboard Financeiro · Uso pessoal · Visão clara e simples", 8, muted)
    T(margin + 5 * mm, 23 * mm, f"Gerado em {date.today().strftime('%d/%m/%Y')} com seus dados do app.", 8, muted)

    footer(3)
    c.save()
    buf.seek(0)
    return buf.getvalue()


# ====================== INICIALIZAÇÃO ======================
init_db()

# ---- Restaura login do cookie/token (sobrevive a refresh e sono do app) ----
restaurar_sessao_se_houver()

# ---- Login obrigatório ----
if "usuario" not in st.session_state or not st.session_state.get("usuario"):
    render_login_page()
    st.stop()

_usuario = st.session_state["usuario"]

df = carregar_dados()
investimentos = carregar_investimentos()
dividas = carregar_dividas()
metas = carregar_metas()

# Métricas principais
entradas = df[df["valor"] > 0]["valor"].sum() if len(df) else 0.0
saidas = abs(df[df["valor"] < 0]["valor"].sum()) if len(df) else 0.0
saldo = df["valor"].sum() if len(df) else 0.0
total_investido = investimentos["valor"].sum() if len(investimentos) else 0.0

if len(dividas):
    d_base = dividas.copy()
    d_base["saldo_base"] = d_base["saldo_negociado"].where(d_base["saldo_negociado"] > 0, d_base["saldo_original"])
    total_dividas_abertas = d_base[d_base["status"] != "Quitada"]["saldo_base"].sum()
    parcelas_dividas = d_base[d_base["status"] != "Quitada"]["parcela_possivel"].sum()
else:
    total_dividas_abertas = 0.0
    parcelas_dividas = 0.0

progresso_metas = 0.0
if len(metas):
    total_meta = metas["valor_meta"].sum()
    total_atual = metas["valor_atual"].sum()
    progresso_metas = (total_atual / total_meta * 100) if total_meta > 0 else 0.0

score = calcular_score_financeiro(
    entradas, saidas, saldo, total_investido, total_dividas_abertas, parcelas_dividas, progresso_metas
)
taxa_sobra = (saldo / entradas * 100) if entradas > 0 else 0.0
comprometimento = ((saidas + parcelas_dividas) / entradas * 100) if entradas > 0 else 0.0

# ====================== CABEÇALHO ======================
_score_cls = "good" if score >= 70 else ("mid" if score >= 45 else "bad")
_score_label = "Saudável" if score >= 70 else ("Atenção" if score >= 45 else "Crítico")
# Frase motivacional rotativa simples (baseada no dia)
_frases_futuro = [
    "Cada real de hoje constrói o amanhã que você escolhe.",
    "O futuro agradece quem organiza o presente.",
    "Pequenos passos financeiros hoje viram liberdade depois.",
    "Cuidar do agora é o jeito mais simples de cuidar do futuro.",
    "Disciplina no bolso é paz no amanhã.",
    "Seu futuro financeiro começa na próxima decisão.",
]
_frase = _frases_futuro[date.today().toordinal() % len(_frases_futuro)]

st.markdown(
    f"""
<div class="page-header">
    <div class="page-header-main">
        <div class="page-header-top">
            <span class="page-kicker">Financeiro pessoal</span>
            <span class="badge-score {_score_cls}">{_score_label} · {score}/100</span>
        </div>
        <h1>Meu dinheiro</h1>
        <p class="page-sub">Saldo, gastos, metas e dívidas em um só lugar.</p>
    </div>
    <div class="page-quote">
        <span class="page-quote-mark">“</span>
        <span class="page-quote-text">{_frase}</span>
    </div>
</div>
""",
    unsafe_allow_html=True,
)

# Cards principais
st.markdown(
    f"""
<div class="metric-grid">
    <div class="metric-card accent">
        <div class="card-icon" aria-hidden="true">↗</div>
        <div class="metric-label">Saldo</div>
        <div class="metric-value">{brl(saldo)}</div>
        <div class="metric-foot">Resultado do período</div>
    </div>
    <div class="metric-card mint">
        <div class="card-icon" aria-hidden="true">↑</div>
        <div class="metric-label">Entradas</div>
        <div class="metric-value">{brl(entradas)}</div>
        <div class="metric-foot">Tudo que entrou</div>
    </div>
    <div class="metric-card">
        <div class="card-icon" aria-hidden="true">↓</div>
        <div class="metric-label">Saídas</div>
        <div class="metric-value">{brl(saidas)}</div>
        <div class="metric-foot">Tudo que saiu</div>
    </div>
    <div class="metric-card">
        <div class="card-icon" aria-hidden="true">◆</div>
        <div class="metric-label">Investido</div>
        <div class="metric-value">{brl(total_investido)}</div>
        <div class="metric-foot">Patrimônio registrado</div>
    </div>
</div>
""",
    unsafe_allow_html=True,
)

# ====================== MENU LATERAL (única instância) ======================
MENU_OPCOES = ["Nova", "Dashboard", "Metas", "Dívidas", "Investir", "Histórico"]

# Garante que o menu não seja montado duas vezes no mesmo run
if "menu_mounted" not in st.session_state:
    st.session_state.menu_mounted = True

with st.sidebar:
    st.markdown(
        """
        <div class="side-brand">
            <div class="side-menu-label">
                <span class="side-spark">✦</span>
                <span>Menu</span>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    pagina = st.radio(
        "Navegação",
        options=MENU_OPCOES,
        index=MENU_OPCOES.index(st.session_state.get("pagina_atual", "Nova")) if st.session_state.get("pagina_atual") in MENU_OPCOES else 0,
        label_visibility="collapsed",
        key="menu_nav_v2",
    )
    st.session_state.pagina_atual = pagina

    st.markdown(
        f"""
        <div class="side-score">
            <div class="side-score-label">Saúde financeira</div>
            <div class="side-score-value">{score}/100</div>
            <div class="side-score-note">{_score_label}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.markdown(
        f"""
        <div style="margin-top:0.85rem;padding:0.7rem 0.85rem;border-radius:12px;
            background:#fff;border:1px solid rgba(28,31,38,0.06);">
            <div style="font-size:0.72rem;font-weight:700;color:#8a90a0;">Conta</div>
            <div style="font-size:0.9rem;font-weight:700;color:#1c1f26;margin-top:0.15rem;">{escape(_usuario.get('nome',''))}</div>
            <div style="font-size:0.75rem;color:#6b7280;">{escape(_usuario.get('email',''))}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    if banco_eh_remoto():
        st.caption("Banco remoto ativo · dados persistentes")
    else:
        st.caption("Banco local · configure Turso nos Secrets para não perder dados")

    if st.button("Sair da conta", use_container_width=True, key="btn_logout"):
        fazer_logout()
        st.rerun()

# ---------- PÁGINA: Nova Movimentação ----------
if pagina == "Nova":
    st.subheader("Registrar movimentação")
    tipo_sel = st.radio(
        "Tipo da movimentação",
        ["Entrada", "Saída"],
        horizontal=True,
        label_visibility="visible",
        key="tipo_mov_radio",
    )
    st.caption("Entrada = dinheiro que entra · Saída = dinheiro que sai")

    with st.form("form_mov", clear_on_submit=False):
        c1, c2 = st.columns(2)
        with c1:
            data_mov = st.date_input("Data", value=date.today(), format="DD/MM/YYYY")
            descricao = st.text_input("Descrição", placeholder="Ex.: Mercado, salário, Uber...")
            if tipo_sel == "Entrada":
                categoria = st.selectbox(
                    "Categoria",
                    [
                        "Salário",
                        "Freelance",
                        "Rendas extras",
                        "Comissão",
                        "13º / Férias",
                        "Reembolso",
                        "Transferência recebida",
                        "Rendimentos / Investimentos",
                        "Presente / Doação",
                        "Outro",
                    ],
                )
            else:
                categoria = st.selectbox(
                    "Categoria",
                    [
                        "Moradia / Aluguel",
                        "Condomínio",
                        "Energia",
                        "Água",
                        "Internet / Telefone",
                        "Gás",
                        "Mercado / Feira",
                        "Delivery / Restaurante",
                        "Transporte / Combustível",
                        "Uber / App",
                        "Saúde / Farmácia",
                        "Plano de saúde",
                        "Educação / Cursos",
                        "Assinaturas",
                        "Lazer / Entretenimento",
                        "Streaming",
                        "Roupas / Calçados",
                        "Beleza / Cuidados pessoais",
                        "Pet",
                        "Filhos / Família",
                        "Casa / Manutenção",
                        "Móveis / Eletro",
                        "Cartão de crédito",
                        "Empréstimo / Financiamento",
                        "Dívidas",
                        "Impostos / Taxas",
                        "Seguros",
                        "Viagem",
                        "Presentes",
                        "Doações",
                        "Investimento / Reserva",
                        "Trabalho / Home office",
                        "Outro",
                    ],
                )
        with c2:
            valor = st.number_input(
                "Valor (R$)",
                min_value=0.0,
                value=0.0,
                step=1.0,
                format="%.2f",
            )
            cartao = st.text_input("Forma de pagamento", placeholder="Ex.: Pix, débito, crédito...")

        if st.form_submit_button("Salvar movimentação", use_container_width=True):
            if not descricao.strip():
                st.error("Informe uma descrição.")
            elif valor is None or float(valor) <= 0:
                st.error("Informe um valor maior que zero.")
            else:
                valor_final = float(valor) if tipo_sel == "Entrada" else -abs(float(valor))
                try:
                    salvar_transacao(data_mov, descricao.strip(), categoria, valor_final, tipo_sel, cartao.strip())
                    st.success("Pronto! Movimentação registrada.")
                    st.rerun()
                except Exception as e:
                    st.error(mensagem_erro_usuario(e))

# ---------- ABA 2: Dashboard ----------
elif pagina == "Dashboard":
    st.subheader("Como está meu dinheiro?")
    with st.expander("Baixar relatório em PDF", expanded=False):
        st.caption("Gera o relatório completo no modelo visual (3 páginas).")
        try:
            _pdf_bytes = gerar_pdf(df, investimentos, dividas, metas)
            st.download_button(
                "Baixar relatório completo em PDF",
                data=_pdf_bytes,
                file_name=f"relatorio-financeiro-{date.today().strftime('%d-%m-%Y')}.pdf",
                mime="application/pdf",
                use_container_width=True,
                key="pdf_dashboard",
            )
        except ModuleNotFoundError:
            st.error(
                "Para gerar o PDF no Streamlit Cloud, adicione no requirements.txt:\n"
                "`reportlab>=4.0.0`\n"
                "Depois faça **Reboot** / **Redeploy** do app."
            )
        except Exception as e:
            st.error(f"Não foi possível gerar o PDF: {mensagem_erro_usuario(e)}")

    # Indicadores
    st.markdown(
        f"""
<div class="indicator-grid">
    <div class="indicator-card">
        <div class="indicator-top">Saúde financeira</div>
        <div class="indicator-value">{score}/100</div>
        <div class="indicator-note">{_score_label} · saldo, dívidas, reserva e metas.</div>
        <div class="progress-track"><span style="width:{limitar_percentual(score)}%;background:{'#0d9f6e' if score >= 70 else ('#e6b400' if score >= 45 else '#e04b5a')}"></span></div>
    </div>
    <div class="indicator-card">
        <div class="indicator-top">Taxa de sobra</div>
        <div class="indicator-value">{pct(taxa_sobra)}</div>
        <div class="indicator-note">Parte das entradas que virou saldo.</div>
        <div class="progress-track"><span style="width:{limitar_percentual(taxa_sobra)}%"></span></div>
    </div>
    <div class="indicator-card">
        <div class="indicator-top">Comprometimento</div>
        <div class="indicator-value">{pct(comprometimento)}</div>
        <div class="indicator-note">Saídas + parcelas vs entradas.</div>
        <div class="progress-track"><span style="width:{limitar_percentual(comprometimento)}%"></span></div>
    </div>
    <div class="indicator-card">
        <div class="indicator-top">Dívidas abertas</div>
        <div class="indicator-value">{brl(total_dividas_abertas)}</div>
        <div class="indicator-note">Saldo a negociar ou pagar.</div>
    </div>
</div>
""",
        unsafe_allow_html=True,
    )

    # Respostas simples
    df_mes, nome_mes = resumo_mes_recente(df)
    fluxo = preparar_fluxo_mensal(df)
    entradas_mes = df_mes[df_mes["valor"] > 0]["valor"].sum() if len(df_mes) else 0
    saidas_mes = abs(df_mes[df_mes["valor"] < 0]["valor"].sum()) if len(df_mes) else 0
    saldo_mes = entradas_mes - saidas_mes
    media_sobra = fluxo["saldo"].mean() if len(fluxo) else saldo

    if saldo_mes >= 0:
        status_mes, classe_mes = "No azul", "answer-good"
        acao_mes = f"Em {nome_mes} sobrou {brl(saldo_mes)}. Separe uma parte antes de gastar."
    else:
        status_mes, classe_mes = "Atenção", "answer-risk"
        acao_mes = f"Em {nome_mes} faltou {brl(abs(saldo_mes))}. Reduza o maior gasto primeiro."

    despesas_mes = df_mes[df_mes["valor"] < 0].copy() if len(df_mes) else pd.DataFrame()
    if len(despesas_mes):
        despesas_mes["abs"] = despesas_mes["valor"].abs()
        top = despesas_mes.groupby("categoria")["abs"].sum().sort_values(ascending=False)
        maior_gasto = f"{top.index[0]} · {brl(top.iloc[0])}" if len(top) else "Sem gastos"
    else:
        maior_gasto = "Sem gastos no mês"

    st.markdown(
        f"""
<div class="chart-intro"><strong>O que preciso saber agora?</strong><br>Respostas simples com base nos seus números.</div>
<div class="answer-grid">
    <div class="answer-card {classe_mes}">
        <div class="answer-question">Como está meu mês?</div>
        <div class="answer-value">{status_mes}</div>
        <div class="answer-action">{acao_mes}</div>
    </div>
    <div class="answer-card answer-care">
        <div class="answer-question">Maior gasto</div>
        <div class="answer-value">{escape(maior_gasto)}</div>
        <div class="answer-action">Revise essa categoria primeiro.</div>
    </div>
    <div class="answer-card {'answer-good' if media_sobra > 0 else 'answer-risk'}">
        <div class="answer-question">Quanto sobra em média?</div>
        <div class="answer-value">{brl(media_sobra)}</div>
        <div class="answer-action">Esse valor pode virar reserva ou meta.</div>
    </div>
    <div class="answer-card answer-care">
        <div class="answer-question">Próxima ação</div>
        <div class="answer-value">{"Guardar primeiro" if media_sobra > 0 else "Cortar vazamento"}</div>
        <div class="answer-action">{"Reserve uma parte ao receber." if media_sobra > 0 else "Escolha uma categoria para reduzir."}</div>
    </div>
</div>
""",
        unsafe_allow_html=True,
    )

    # Projeção futura
    with st.expander("🔮 Meu dinheiro futuro", expanded=False):
        st.caption("Simule os próximos 6 meses. A projeção usa sua média mensal atual.")
        c1, c2, c3 = st.columns(3)
        with c1:
            reduzir = st.slider("E se eu gastar menos por mês?", 0, 3000, 0, 50, format="R$ %d")
        with c2:
            extra = st.slider("E se eu receber um extra por mês?", 0, 5000, 0, 50, format="R$ %d")
        with c3:
            meta_mes = st.slider("Quero separar para metas por mês", 0, 5000, 0, 50, format="R$ %d")

        media_entradas = fluxo["entradas"].mean() if len(fluxo) else entradas
        media_saidas = fluxo["saidas"].mean() if len(fluxo) else saidas
        fluxo_normal = media_sobra + reduzir + extra - meta_mes
        ajuste_bom = max(media_entradas * 0.05, 100) if media_entradas else 100
        ajuste_apertado = max(media_saidas * 0.08, 100) if media_saidas else 100

        inicio = pd.Timestamp(date.today()).to_period("M").to_timestamp()
        proj = []
        s_normal = s_bom = s_apertado = saldo
        for i in range(1, 7):
            mes = inicio + pd.DateOffset(months=i)
            s_normal += fluxo_normal
            s_bom += fluxo_normal + ajuste_bom
            s_apertado += fluxo_normal - ajuste_apertado
            proj.extend([
                {"Mês": mes, "Cenário": "Normal", "Saldo projetado": s_normal},
                {"Mês": mes, "Cenário": "Bom", "Saldo projetado": s_bom},
                {"Mês": mes, "Cenário": "Apertado", "Saldo projetado": s_apertado},
            ])
        df_proj = pd.DataFrame(proj)

        neg = df_proj[(df_proj["Cenário"] == "Normal") & (df_proj["Saldo projetado"] < 0)]
        if neg.empty:
            resp = "Pelo cenário normal, seu saldo continua positivo nos próximos 6 meses."
        else:
            resp = f"No cenário normal, seu saldo pode ficar negativo em {neg.iloc[0]['Mês'].strftime('%m/%Y')}."

        st.markdown(f"""<div class="history-summary"><strong>Vou ficar sem dinheiro?</strong><br>{resp}</div>""", unsafe_allow_html=True)

        fig = px.line(
            df_proj, x="Mês", y="Saldo projetado", color="Cenário",
            title="E nos próximos meses?",
            markers=True,
            color_discrete_map={"Bom": "#0d9f6e", "Normal": "#d9ff00", "Apertado": "#e04b5a"},
            labels={"Saldo projetado": "Saldo projetado", "Mês": "Mês"},
        )
        fig.update_traces(
            line=dict(width=3),
            marker=dict(size=8),
            hovertemplate="<b>%{fullData.name}</b><br>%{x|%m/%Y}<br>R$ %{y:,.2f}<extra></extra>",
        )
        fig.update_yaxes(tickprefix="R$ ")
        fig.update_xaxes(tickformat="%m/%Y")
        fig = style_plot(fig, height=340)
        fig.update_layout(
            legend=dict(orientation="h", y=-0.18, x=0.5, xanchor="center", title_text=""),
            margin=dict(l=28, r=12, t=44, b=70),
        )
        st.plotly_chart(fig, use_container_width=True, config=PLOTLY_CONFIG)

    # Importação
    with st.expander("📤 Subir planilha (Financeira Inteligente ou modelo simples)", expanded=False):
        st.caption(
            "Aceita a Planilha Financeira Inteligente (abas Janeiro–Dezembro, Dívidas e Metas) "
            "ou uma planilha simples com colunas Data/Descrição/Valor. Duplicados são ignorados."
        )
        st.download_button(
            "Baixar modelo Excel simples",
            data=gerar_modelo_excel(),
            file_name="modelo-dashboard-financeiro.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        arquivo = st.file_uploader("Escolha a planilha", type=["csv", "xlsx", "xls"], key="upload_planilha_intel")
        if arquivo is not None:
            try:
                dados = ler_planilha(arquivo)
                pacote = preparar_importacao(dados)
                df_imp = pacote.get("movimentacoes", pd.DataFrame())
                if not isinstance(df_imp, pd.DataFrame):
                    df_imp = pd.DataFrame()
                ignoradas = pacote.get("ignoradas", 0)
                lista_div = pacote.get("dividas", []) or []
                lista_meta = pacote.get("metas", []) or []
                abas = pacote.get("abas_lidas", []) or []
                modelo = pacote.get("modelo", "")

                df_novo, duplicadas = conciliar_movimentacoes(df_imp, df) if len(df_imp) else (pd.DataFrame(), 0)

                if modelo == "planilha_inteligente":
                    st.success(
                        "Planilha Financeira Inteligente reconhecida"
                        + (f" — abas: {', '.join(abas)}" if abas else "")
                    )

                c1, c2, c3, c4 = st.columns(4)
                c1.metric("Novas entradas", brl(df_novo[df_novo["valor"] > 0]["valor"].sum() if len(df_novo) else 0))
                c2.metric("Novas saídas", brl(abs(df_novo[df_novo["valor"] < 0]["valor"].sum() if len(df_novo) else 0)))
                c3.metric("Dívidas lidas", len(lista_div))
                c4.metric("Metas lidas", len(lista_meta))

                st.caption(
                    f"Movimentações novas: {len(df_novo)} · Duplicadas: {duplicadas} · "
                    f"Ignoradas: {ignoradas} · Impacto no saldo: {brl(df_novo['valor'].sum() if len(df_novo) else 0)}"
                )

                if len(df_novo):
                    previa = df_novo.rename(columns={
                        "data": "Data", "descricao": "Descrição", "categoria": "Categoria",
                        "valor": "Valor", "tipo": "Tipo", "cartao": "Pagamento",
                    }).copy()
                    previa["Data"] = previa["Data"].map(data_br)
                    previa["Valor"] = previa["Valor"].map(brl)
                    st.dataframe(previa.head(80), use_container_width=True, hide_index=True)
                    if len(previa) > 80:
                        st.caption(f"Mostrando 80 de {len(previa)} lançamentos.")

                if lista_div:
                    st.markdown("**Dívidas encontradas**")
                    st.dataframe(
                        pd.DataFrame(lista_div)[
                            [c for c in ["credor", "tipo", "saldo_negociado", "parcela_possivel", "prioridade", "status"] if c in pd.DataFrame(lista_div).columns]
                        ],
                        use_container_width=True,
                        hide_index=True,
                    )

                if lista_meta:
                    st.markdown("**Metas encontradas**")
                    st.dataframe(
                        pd.DataFrame(lista_meta)[
                            [c for c in ["nome", "valor_meta", "valor_atual", "prazo", "status"] if c in pd.DataFrame(lista_meta).columns]
                        ],
                        use_container_width=True,
                        hide_index=True,
                    )

                pode_integrar = len(df_novo) > 0 or len(lista_div) > 0 or len(lista_meta) > 0
                if pode_integrar:
                    if st.button("Integrar tudo ao dashboard", type="primary", key="btn_integrar_planilha", use_container_width=True):
                        total, dup = importar_movimentacoes(df_imp) if len(df_imp) else (0, 0)
                        n_div = importar_dividas_lista(lista_div)
                        n_meta = importar_metas_lista(lista_meta)
                        st.success(
                            f"Integrado: {total} movimentações ({dup} duplicadas ignoradas), "
                            f"{n_div} dívidas, {n_meta} metas."
                        )
                        st.rerun()
                else:
                    st.info("Nada novo para importar — tudo já está no dashboard ou não foi reconhecido.")
            except Exception as e:
                st.error(f"Não consegui importar: {mensagem_erro_usuario(e)}")

    # Gráficos
    if len(df):
        df_chart = df.copy()
        df_chart["categoria"] = df_chart["categoria"].replace("", "Sem categoria").fillna("Sem categoria")
        df_chart["cartao"] = df_chart["cartao"].replace("", "Não informado").fillna("Não informado")
        df_chart["valor_abs"] = df_chart["valor"].abs()
        df_chart["data_convertida"] = pd.to_datetime(df_chart["data"], errors="coerce")

        df_tl = df_chart.dropna(subset=["data_convertida"]).sort_values("data_convertida")
        if len(df_tl):
            df_tl["saldo_acumulado"] = df_tl["valor"].cumsum()
            fig_saldo = px.area(
                df_tl, x="data_convertida", y="saldo_acumulado",
                title="Meu saldo ao longo do tempo",
                labels={"data_convertida": "Data", "saldo_acumulado": "Saldo"},
            )
            fig_saldo.update_traces(
                line=dict(color="#1c1f26", width=3),
                fillcolor="rgba(126,217,176,0.28)",
                hovertemplate="<b>%{x|%d/%m/%Y}</b><br>Saldo: R$ %{y:,.2f}<extra></extra>",
            )
            fig_saldo.update_yaxes(tickprefix="R$ ")
            fig_saldo.update_xaxes(tickformat="%d/%m/%Y")
            st.plotly_chart(style_plot(fig_saldo, height=320, show_legend=False), use_container_width=True, config=PLOTLY_CONFIG)

        cat_total = (
            df_chart.groupby("categoria", as_index=False)["valor_abs"]
            .sum()
            .sort_values("valor_abs", ascending=False)
        )
        fluxo_tipo = df_chart.groupby(["categoria", "tipo"], as_index=False)["valor_abs"].sum()

        # --- Donut + Barras (lado a lado) ---
        c1, c2 = st.columns(2)
        with c1:
            # Top 5 + "Outras" (evita labels cortados no mobile)
            if len(cat_total) > 5:
                top8 = cat_total.head(5).copy()
                outros_valor = float(cat_total.iloc[5:]["valor_abs"].sum())
                if outros_valor > 0:
                    top8 = pd.concat([
                        top8,
                        pd.DataFrame([{"categoria": "Outras", "valor_abs": outros_valor}]),
                    ], ignore_index=True)
            else:
                top8 = cat_total.copy()

            fig = px.pie(
                top8,
                names="categoria",
                values="valor_abs",
                hole=0.58,
                title="Para onde vai o dinheiro?",
                color_discrete_sequence=[
                    "#d9ff00", "#1c1f26", "#7ed9b0", "#a8c4ff",
                    "#b8f0d8", "#c5c9d4", "#f3ff9a",
                ],
            )
            fig.update_traces(
                textposition="inside",
                textinfo="percent",
                textfont=dict(size=11, color="#1c1f26"),
                insidetextorientation="horizontal",
                marker=dict(line=dict(color="rgba(255,255,255,0.95)", width=2)),
                pull=[0.01] * len(top8),
                hovertemplate="<b>%{label}</b><br>R$ %{value:,.2f}<br>%{percent}<extra></extra>",
                rotation=20,
            )
            fig = style_plot(fig, height=320)
            fig.update_layout(
                margin=dict(l=10, r=10, t=44, b=72),
                legend=dict(
                    orientation="h",
                    yanchor="top",
                    y=-0.12,
                    xanchor="center",
                    x=0.5,
                    font=dict(size=11),
                    itemwidth=40,
                ),
            )
            st.plotly_chart(fig, use_container_width=True, config=PLOTLY_CONFIG)

        with c2:
            fig2 = px.bar(
                fluxo_tipo,
                x="categoria",
                y="valor_abs",
                color="tipo",
                title="O que entra e o que sai",
                color_discrete_map={"Entrada": "#d9ff00", "Saída": "#1c1f26"},
                labels={"categoria": "Categoria", "valor_abs": "Valor", "tipo": "Tipo"},
                barmode="group",
            )
            fig2.update_traces(
                marker_line_color="rgba(255,255,255,0.9)",
                marker_line_width=1,
                hovertemplate="<b>%{x}</b><br>R$ %{y:,.2f}<extra></extra>",
            )
            fig2.update_yaxes(tickprefix="R$ ")
            fig2.update_xaxes(tickangle=-35)
            fig2 = style_plot(fig2, height=340)
            fig2.update_layout(
                margin=dict(l=28, r=12, t=44, b=80),
                legend=dict(orientation="h", y=-0.22, x=0.5, xanchor="center"),
            )
            st.plotly_chart(fig2, use_container_width=True, config=PLOTLY_CONFIG)

        # --- Barras horizontais + Formas de pagamento ---
        c3, c4 = st.columns(2)
        with c3:
            top = cat_total.head(8).sort_values("valor_abs", ascending=True)
            fig3 = px.bar(
                top,
                x="valor_abs",
                y="categoria",
                orientation="h",
                title="O que mais mexe com o bolso",
                color="valor_abs",
                color_continuous_scale=["#e8faf3", "#7ed9b0", "#1c1f26"],
                labels={"valor_abs": "Valor", "categoria": ""},
                text="valor_abs",
            )
            fig3.update_traces(
                texttemplate="R$ %{x:,.0f}",
                textposition="outside",
                textfont=dict(size=11, color="#1c1f26"),
                hovertemplate="<b>%{y}</b><br>R$ %{x:,.2f}<extra></extra>",
                marker_line_color="rgba(255,255,255,0.9)",
                marker_line_width=1,
                cliponaxis=False,
            )
            fig3.update_layout(coloraxis_showscale=False)
            max_v = float(top["valor_abs"].max()) if len(top) else 1.0
            fig3.update_xaxes(tickprefix="R$ ", range=[0, max_v * 1.25 if max_v > 0 else 1])
            fig3 = style_plot(fig3, height=320, show_legend=False)
            fig3.update_layout(margin=dict(l=8, r=40, t=44, b=36))
            st.plotly_chart(fig3, use_container_width=True, config=PLOTLY_CONFIG)

        with c4:
            pag = (
                df_chart.groupby("cartao", as_index=False)["valor_abs"]
                .sum()
                .sort_values("valor_abs", ascending=False)
            )
            if len(pag) > 5:
                top_pag = pag.head(4).copy()
                rest_pag = float(pag.iloc[4:]["valor_abs"].sum())
                if rest_pag > 0:
                    top_pag = pd.concat([
                        top_pag,
                        pd.DataFrame([{"cartao": "Outras", "valor_abs": rest_pag}]),
                    ], ignore_index=True)
                pag = top_pag
            fig4 = px.pie(
                pag,
                names="cartao",
                values="valor_abs",
                hole=0.58,
                title="Como eu pago?",
                color_discrete_sequence=["#1c1f26", "#d9ff00", "#7ed9b0", "#a8c4ff", "#b8f0d8", "#c5c9d4"],
            )
            fig4.update_traces(
                textposition="inside",
                textinfo="percent",
                textfont=dict(size=11, color="#1c1f26"),
                insidetextorientation="horizontal",
                marker=dict(line=dict(color="rgba(255,255,255,0.95)", width=2)),
                pull=[0.01] * len(pag),
                hovertemplate="<b>%{label}</b><br>R$ %{value:,.2f}<br>%{percent}<extra></extra>",
            )
            fig4 = style_plot(fig4, height=320)
            fig4.update_layout(
                margin=dict(l=10, r=10, t=44, b=72),
                legend=dict(
                    orientation="h",
                    yanchor="top",
                    y=-0.12,
                    xanchor="center",
                    x=0.5,
                    font=dict(size=11),
                ),
            )
            st.plotly_chart(fig4, use_container_width=True, config=PLOTLY_CONFIG)

        if len(fluxo):
            fig_m = px.bar(
                fluxo,
                x="mes",
                y=["entradas", "saidas"],
                barmode="group",
                title="Mês a mês",
                color_discrete_map={"entradas": "#d9ff00", "saidas": "#1c1f26"},
                labels={"mes": "Mês", "value": "Valor", "variable": "Tipo"},
            )
            fig_m.for_each_trace(
                lambda t: t.update(
                    marker_line_color="rgba(255,255,255,0.9)",
                    marker_line_width=1,
                    hovertemplate="<b>%{x|%m/%Y}</b><br>R$ %{y:,.2f}<extra></extra>",
                )
            )
            fig_m.update_xaxes(tickformat="%m/%Y")
            fig_m.update_yaxes(tickprefix="R$ ")
            fig_m = style_plot(fig_m, height=340)
            fig_m.update_layout(
                legend=dict(
                    orientation="h",
                    y=-0.18,
                    x=0.5,
                    xanchor="center",
                    title_text="",
                ),
                margin=dict(l=28, r=12, t=44, b=70),
            )
            st.plotly_chart(fig_m, use_container_width=True, config=PLOTLY_CONFIG)
    else:
        st.info("Cadastre ou importe movimentações para ver os gráficos.")

# ---------- ABA 3: Metas ----------
elif pagina == "Metas":
    st.subheader("Metas")
    st.markdown(
        """<div class="chart-intro"><strong>Minhas metas</strong><br>
        Defina objetivos simples, acompanhe quanto falta e veja em quanto tempo chega lá.</div>""",
        unsafe_allow_html=True,
    )

    with st.form("form_meta"):
        st.markdown("#### Adicionar meta")
        c1, c2, c3 = st.columns(3)
        with c1:
            data_m = st.date_input("Data da meta", value=date.today(), format="DD/MM/YYYY")
            nome_m = st.text_input("Nome da meta", placeholder="Ex.: Reserva, reforma, viagem")
            status_m = st.selectbox("Status", ["Em andamento", "Planejada", "Concluída"])
        with c2:
            valor_m = st.number_input("Valor necessário (R$)", value=0.0, min_value=0.0, step=0.01)
            atual_m = st.number_input("Quanto já tenho (R$)", value=0.0, min_value=0.0, step=0.01)
            aporte_m = st.number_input("Aporte mensal (R$)", value=0.0, min_value=0.0, step=0.01)
        with c3:
            prazo_m = st.text_input("Prazo desejado", placeholder="Ex.: Dezembro/2026")
            anot_m = st.text_area("Anotações")

        if st.form_submit_button("Salvar meta"):
            if not nome_m.strip():
                st.error("Informe o nome da meta.")
            elif valor_m <= 0:
                st.error("Informe o valor necessário.")
            else:
                try:
                    salvar_meta(data_m, nome_m.strip(), valor_m, atual_m, aporte_m, prazo_m.strip(), status_m, anot_m.strip())
                    st.success("Meta salva!")
                    st.rerun()
                except Exception as e:
                    st.error(mensagem_erro_usuario(e))

    if len(metas):
        st.markdown("#### Progresso das metas")
        for _, m in metas.iterrows():
            vm = max(float(m["valor_meta"]), 0)
            va = max(float(m["valor_atual"]), 0)
            am = max(float(m["aporte_mensal"]), 0)
            perc = min((va / vm) * 100, 100) if vm > 0 else 0
            falta = max(vm - va, 0)
            meses = math.ceil(falta / am) if falta > 0 and am > 0 else 0
            prev = f"Chega em {texto_meses(meses)}" if falta > 0 and am > 0 else ("Meta concluída" if falta <= 0 else "Defina um aporte mensal")

            st.markdown(
                f"""
<div class="goal-card">
    <div style="font-weight:850; font-size:1.05rem;">{escape(str(m['nome'] or 'Meta'))}</div>
    <div class="goal-meta">Status: {escape(str(m['status']))} • Prazo: {escape(str(m['prazo'] or '—'))}</div>
    <div class="goal-progress"><span style="width:{perc:.1f}%"></span></div>
    <div class="goal-meta">
        <strong>{perc:.0f}% concluída</strong><br>
        Tenho {brl(va)} de {brl(vm)}. Falta {brl(falta)}.<br>
        {prev}. Aporte: {brl(am)}.
    </div>
</div>
""",
                unsafe_allow_html=True,
            )
            if st.button("Apagar meta", key=f"del_meta_{m['id']}", type="secondary"):
                excluir_meta(m["id"])
                st.rerun()
    else:
        st.info("Nenhuma meta cadastrada ainda.")

# ---------- ABA 4: Dívidas ----------
elif pagina == "Dívidas":
    st.subheader("Dívidas e Negociação")
    st.markdown(
        """<div class="chart-intro"><strong>Controle de dívidas</strong><br>
        Registre credor, saldo, parcela, prioridade e próxima ação.</div>""",
        unsafe_allow_html=True,
    )

    with st.form("form_divida"):
        st.markdown("#### Adicionar dívida ou acordo")
        c1, c2, c3 = st.columns(3)
        with c1:
            data_d = st.date_input("Data do registro", value=date.today(), format="DD/MM/YYYY")
            credor = st.text_input("Credor", placeholder="Ex.: Banco, cartão, loja")
            tipo_d = st.selectbox("Tipo", ["Cartão de crédito", "Empréstimo", "Conta atrasada", "Financiamento", "Acordo", "Outro"])
            prioridade = st.selectbox("Prioridade", ["Alta", "Média", "Baixa"])
        with c2:
            saldo_orig = st.number_input("Saldo original (R$)", value=0.0, min_value=0.0, step=0.01)
            desconto = st.number_input("Desconto (R$)", value=0.0, min_value=0.0, step=0.01)
            saldo_neg = st.number_input("Saldo negociado (R$)", value=0.0, min_value=0.0, step=0.01)
            parcela = st.number_input("Parcela possível (R$)", value=0.0, min_value=0.0, step=0.01)
        with c3:
            venc = st.date_input("Vencimento / próximo prazo", value=date.today(), format="DD/MM/YYYY")
            status_d = st.selectbox("Status", ["Mapear", "Negociar", "Acordada", "Em pagamento", "Quitada"])
            proxima = st.text_input("Próxima ação", placeholder="Ex.: ligar, pedir desconto")
            anot_d = st.text_area("Anotações")

        if st.form_submit_button("Salvar dívida"):
            saldo_final = saldo_neg if saldo_neg > 0 else max(saldo_orig - desconto, 0)
            if not credor.strip():
                st.error("Informe o credor.")
            elif saldo_final <= 0:
                st.error("Informe o saldo.")
            else:
                try:
                    salvar_divida(
                        data_d, credor.strip(), tipo_d, saldo_orig, desconto, saldo_final,
                        parcela, venc, prioridade, "", status_d, proxima.strip(), anot_d.strip(),
                    )
                    st.success("Dívida salva!")
                    st.rerun()
                except Exception as e:
                    st.error(mensagem_erro_usuario(e))

    if len(dividas):
        d_view = dividas.copy()
        d_view["saldo_base"] = d_view["saldo_negociado"].where(d_view["saldo_negociado"] > 0, d_view["saldo_original"])
        abertas = d_view[d_view["status"] != "Quitada"]
        total_aberto = abertas["saldo_base"].sum()
        parcelas_acord = abertas["parcela_possivel"].sum()
        economia = (d_view["saldo_original"] - d_view["saldo_base"]).clip(lower=0).sum()
        prioritarias = len(d_view[d_view["prioridade"] == "Alta"])

        st.markdown(
            f"""
<div class="debt-grid">
    <div class="metric-card"><div class="metric-label">Total aberto</div><div class="metric-value">{brl(total_aberto)}</div><div class="metric-foot">Saldo para negociar ou pagar</div></div>
    <div class="metric-card"><div class="metric-label">Parcelas acordadas</div><div class="metric-value">{brl(parcelas_acord)}</div><div class="metric-foot">Compromisso mensal possível</div></div>
    <div class="metric-card"><div class="metric-label">Economia prevista</div><div class="metric-value">{brl(economia)}</div><div class="metric-foot">Descontos registrados</div></div>
    <div class="metric-card"><div class="metric-label">Prioridade alta</div><div class="metric-value">{prioritarias}</div><div class="metric-foot">Dívidas que exigem atenção</div></div>
</div>
""",
            unsafe_allow_html=True,
        )

        with st.expander("⚡ E se eu quiser pagar mais rápido?", expanded=False):
            base = parcelas_acord if parcelas_acord > 0 else 0
            extra_d = st.slider("Quero pagar a mais por mês", 0, 5000, 0, 50, format="R$ %d")
            total_pag = base + extra_d
            if total_aberto > 0 and total_pag > 0:
                meses_atuais = math.ceil(total_aberto / base) if base > 0 else 0
                meses_novos = math.ceil(total_aberto / total_pag)
                ganho = max(meses_atuais - meses_novos, 0) if meses_atuais else 0
                st.markdown(
                    f"""
<div class="answer-grid">
    <div class="answer-card answer-care">
        <div class="answer-question">Prazo atual</div>
        <div class="answer-value">{texto_meses(meses_atuais) if meses_atuais else "Sem parcela"}</div>
        <div class="answer-action">No ritmo atual.</div>
    </div>
    <div class="answer-card answer-good">
        <div class="answer-question">Novo prazo</div>
        <div class="answer-value">{texto_meses(meses_novos)}</div>
        <div class="answer-action">Com {brl(total_pag)} por mês.</div>
    </div>
    <div class="answer-card answer-good">
        <div class="answer-question">Tempo ganho</div>
        <div class="answer-value">{texto_meses(ganho)}</div>
        <div class="answer-action">Estimativa simples (sem juros futuros).</div>
    </div>
</div>
""",
                    unsafe_allow_html=True,
                )
            else:
                st.info("Cadastre saldo e parcela para simular.")

        for _, d in d_view.iterrows():
            sb = d["saldo_base"]
            st.markdown(
                f"""
<div class="debt-item">
    <div>
        <div style="font-weight:850; font-size:1.05rem;">{escape(str(d['credor'] or 'Credor'))}</div>
        <div class="debt-meta">{escape(str(d['tipo']))} • Venc: {data_br(d['vencimento'])} • {escape(str(d['status']))} • Prioridade: {escape(str(d['prioridade']))}</div>
        <div class="debt-meta"><strong>Próxima ação:</strong> {escape(str(d['proxima_acao'] or '—'))}<br><strong>Anotações:</strong> {escape(str(d['anotacoes'] or '—'))}</div>
    </div>
    <div style="text-align:right;">
        <div class="negative" style="font-size:1.1rem; font-weight:900;">{brl(sb)}</div>
        <div class="debt-meta">Parcela: {brl(d['parcela_possivel'])}</div>
    </div>
</div>
""",
                unsafe_allow_html=True,
            )
            if st.button("Apagar dívida", key=f"del_div_{d['id']}", type="secondary"):
                excluir_divida(d["id"])
                st.rerun()
    else:
        st.info("Nenhuma dívida cadastrada ainda.")

# ---------- ABA 5: Investimentos ----------
elif pagina == "Investir":
    st.subheader("Investimentos")
    with st.form("form_inv"):
        st.markdown("#### Adicionar investimento")
        c1, c2, c3 = st.columns(3)
        with c1:
            data_i = st.date_input("Data", value=date.today(), format="DD/MM/YYYY", key="di")
            tipo_i = st.selectbox(
                "Tipo",
                ["Reserva de emergência", "Tesouro Direto", "CDB", "LCI / LCA", "Fundo", "Ações", "FII", "Previdência", "Cripto", "Outro"],
            )
        with c2:
            valor_i = st.number_input("Valor investido (R$)", value=0.0, min_value=0.0, step=0.01)
            rent_i = st.text_input("Rentabilidade", placeholder="Ex.: 12% a.a. ou 105% CDI")
        with c3:
            desc_i = st.text_input("Descrição", placeholder="Ex.: CDB Banco X")
            status_i = st.selectbox("Status", ["Ativo", "Planejado", "Resgatado"])

        if st.form_submit_button("Salvar investimento"):
            if valor_i <= 0:
                st.error("Informe um valor maior que zero.")
            else:
                try:
                    salvar_investimento(data_i, tipo_i, valor_i, rent_i, desc_i, status_i)
                    st.success("Investimento salvo!")
                    st.rerun()
                except Exception as e:
                    st.error(mensagem_erro_usuario(e))

    if len(investimentos):
        total_ativos = investimentos[investimentos["status"] == "Ativo"]["valor"].sum()
        total_plan = investimentos[investimentos["status"] == "Planejado"]["valor"].sum()
        st.markdown(
            f"""
<div class="investment-grid">
    <div class="metric-card"><div class="metric-label">Patrimônio registrado</div><div class="metric-value">{brl(total_investido)}</div></div>
    <div class="metric-card"><div class="metric-label">Ativos</div><div class="metric-value">{brl(total_ativos)}</div></div>
    <div class="metric-card"><div class="metric-label">Planejados</div><div class="metric-value">{brl(total_plan)}</div></div>
</div>
""",
            unsafe_allow_html=True,
        )

        tipo_inv = investimentos.groupby("tipo", as_index=False)["valor"].sum().sort_values("valor", ascending=False)
        c1, c2 = st.columns(2)
        with c1:
            fig = px.pie(
                tipo_inv.head(5) if len(tipo_inv) > 5 else tipo_inv,
                names="tipo",
                values="valor",
                hole=0.58,
                title="Onde está investido?",
                color_discrete_sequence=["#1c1f26", "#d9ff00", "#7ed9b0", "#a8c4ff", "#b8f0d8", "#c5c9d4"],
            )
            if len(tipo_inv) > 5:
                # rebuild with Outras
                top_inv = tipo_inv.head(5).copy()
                rest = float(tipo_inv.iloc[5:]["valor"].sum())
                if rest > 0:
                    top_inv = pd.concat([top_inv, pd.DataFrame([{"tipo": "Outras", "valor": rest}])], ignore_index=True)
                fig = px.pie(
                    top_inv,
                    names="tipo",
                    values="valor",
                    hole=0.58,
                    title="Onde está investido?",
                    color_discrete_sequence=["#1c1f26", "#d9ff00", "#7ed9b0", "#a8c4ff", "#b8f0d8", "#c5c9d4"],
                )
            fig.update_traces(
                textposition="inside",
                textinfo="percent",
                textfont=dict(size=11, color="#1c1f26"),
                insidetextorientation="horizontal",
                marker=dict(line=dict(color="rgba(255,255,255,0.95)", width=2)),
                hovertemplate="<b>%{label}</b><br>R$ %{value:,.2f}<br>%{percent}<extra></extra>",
            )
            fig = style_plot(fig, height=320)
            fig.update_layout(
                margin=dict(l=10, r=10, t=44, b=72),
                legend=dict(orientation="h", y=-0.15, x=0.5, xanchor="center", font=dict(size=11)),
            )
            st.plotly_chart(fig, use_container_width=True, config=PLOTLY_CONFIG)
        with c2:
            fig2 = px.bar(
                investimentos.groupby("status", as_index=False)["valor"].sum(),
                x="status", y="valor", color="status",
                title="Status dos investimentos",
                color_discrete_map={"Ativo": "#0d9f6e", "Planejado": "#7ed9b0", "Resgatado": "#a8c4ff"},
                labels={"status": "Status", "valor": "Valor"},
            )
            fig2.update_traces(
                texttemplate="R$ %{y:,.0f}",
                textposition="outside",
                marker_line_color="rgba(255,255,255,0.9)",
                marker_line_width=1,
                hovertemplate="<b>%{x}</b><br>R$ %{y:,.2f}<extra></extra>",
            )
            fig2.update_yaxes(tickprefix="R$ ")
            fig2 = style_plot(fig2, height=320, show_legend=False)
            fig2.update_layout(margin=dict(l=40, r=20, t=60, b=40))
            st.plotly_chart(fig2, use_container_width=True, config=PLOTLY_CONFIG)

        for _, inv in investimentos.iterrows():
            st.markdown(
                f"""
<div class="investment-item">
    <div>
        <div style="font-weight:850;">{escape(str(inv['descricao'] or 'Investimento'))}</div>
        <div class="investment-meta">{escape(str(inv['tipo']))} • {data_br(inv['data'])} • {escape(str(inv['status']))}</div>
    </div>
    <div style="text-align:right;">
        <div class="positive" style="font-weight:900;">{brl(inv['valor'])}</div>
        <div class="investment-meta">Rentab.: {escape(str(inv['rentabilidade'] or '—'))}</div>
    </div>
</div>
""",
                unsafe_allow_html=True,
            )
            if st.button("Apagar investimento", key=f"del_inv_{inv['id']}", type="secondary"):
                excluir_investimento(inv["id"])
                st.rerun()
    else:
        st.info("Nenhum investimento cadastrado ainda.")

# ---------- ABA 6: Histórico ----------
elif pagina == "Histórico":
    st.subheader("Histórico")
    st.markdown(
        f"""<div class="history-summary"><strong>Relatório detalhado</strong><br>
        {len(df)} movimentações • {len(investimentos)} investimentos • {len(dividas)} dívidas • {len(metas)} metas</div>""",
        unsafe_allow_html=True,
    )

    try:
        _pdf_hist = gerar_pdf(df, investimentos, dividas, metas)
        st.download_button(
            "Baixar relatório completo em PDF",
            data=_pdf_hist,
            file_name=f"relatorio-financeiro-{date.today().strftime('%d-%m-%Y')}.pdf",
            mime="application/pdf",
            use_container_width=True,
            key="pdf_historico",
        )
        st.caption("Capa, fluxo, categorias, dívidas, metas, projeção e alertas — modelo visual aprovado.")
    except ModuleNotFoundError:
        st.error(
            "Para gerar o PDF no Streamlit Cloud, adicione `reportlab>=4.0.0` no requirements.txt e faça redeploy."
        )
    except Exception as e:
        st.error(f"Não foi possível gerar o PDF: {mensagem_erro_usuario(e)}")


    if len(df):
        with st.expander("🧹 Limpar histórico completo", expanded=False):
            st.caption("Apaga todas as movimentações. Investimentos, dívidas e metas não são alterados.")
            conf = st.checkbox("Confirmo que quero apagar todas as movimentações")
            if st.button("Apagar todo o histórico", disabled=not conf):
                limpar_historico()
                st.success("Histórico limpo.")
                st.rerun()

        c1, c2, c3 = st.columns([2, 1, 1])
        busca = c1.text_input("Buscar", placeholder="Descrição, categoria ou pagamento")
        tipo_f = c2.selectbox("Tipo", ["Todos", "Entrada", "Saída"])
        cats_raw = df["categoria"].fillna("Sem categoria").replace("", "Sem categoria")
        cats = sorted({str(c) for c in cats_raw.unique() if str(c).lower() not in ("nan", "none", "")})
        cat_f = c3.selectbox("Categoria", ["Todas", *cats])

        df_h = df.copy()
        if busca.strip():
            termo = normalizar(busca)
            def _bate(r):
                texto = " ".join([
                    limpar_texto(r.get("descricao") if hasattr(r, "get") else r["descricao"]),
                    limpar_texto(r.get("categoria") if hasattr(r, "get") else r["categoria"]),
                    limpar_texto(r.get("cartao") if hasattr(r, "get") else r["cartao"]),
                ])
                return termo in normalizar(texto)
            try:
                df_h = df_h[df_h.apply(_bate, axis=1)]
            except Exception:
                pass
        if tipo_f != "Todos":
            df_h = df_h[df_h["tipo"] == tipo_f]
        if cat_f != "Todas":
            df_h = df_h[df_h["categoria"].fillna("Sem categoria").replace("", "Sem categoria") == cat_f]

        cols_show = ["data", "descricao", "categoria", "tipo", "cartao", "valor"]
        tabela = df_h[cols_show].copy()
        tabela = tabela.rename(columns={
            "data": "Data",
            "descricao": "Descricao",
            "categoria": "Categoria",
            "tipo": "Tipo",
            "cartao": "Pagamento",
            "valor": "Valor",
        })
        tabela["Data"] = tabela["Data"].astype(str).map(lambda v: data_br(v))
        tabela["Valor"] = tabela["Valor"].apply(lambda v: brl(v))
        st.dataframe(tabela, use_container_width=True, hide_index=True)
        st.caption(f"{len(df_h)} de {len(df)} movimentacoes exibidas.")

        records = df_h.to_dict(orient="records")
        for rec in records:
            try:
                valor_row = float(rec.get("valor") or 0)
            except Exception:
                valor_row = 0.0
            if valor_row != valor_row:
                valor_row = 0.0
            classe = "positive" if valor_row >= 0 else "negative"
            desc = limpar_texto(rec.get("descricao"), "Sem descricao")
            cat = limpar_texto(rec.get("categoria"), "Sem categoria")
            cart = limpar_texto(rec.get("cartao"), "-")
            data_txt = data_br(rec.get("data"))
            valor_txt = brl(valor_row)
            html_item = (
                '<div class="history-item">'
                "<div>"
                f'<div style="font-weight:850;">{escape(desc)}</div>'
                f'<div style="color:#8a90a0; font-size:0.88rem;">{escape(cat)} | {data_txt} | {escape(cart)}</div>'
                "</div>"
                f'<div class="{classe}">{valor_txt}</div>'
                "</div>"
            )
            st.markdown(html_item, unsafe_allow_html=True)
            try:
                rid = int(rec.get("id") or 0)
            except Exception:
                rid = 0
            if rid and st.button("Apagar", key=f"del_t_{rid}", type="secondary"):
                excluir_transacao(rid)
                st.rerun()

    else:
        st.info("Nenhum registro ainda.")

st.caption("Dashboard Financeiro • Visão clara e simples • Uso pessoal")
